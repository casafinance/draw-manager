"""
Draw Bulk Upload --- FCI
========================
Reads one or more 'today' sheets + the 'Draw' sheet from a live Google Sheet,
AND/OR parses one or more scanned/exported draw-bulk PDFs, optionally enriches
each property with Loan# + Investor Account# from FCI (via the FCI API or a
Chrome remote-debug attach), and appends rows to a local balance-update .xlsx
using the original column layout:

    A: Loan Number
    B: Investor Account Number
    C: Amount
    D: Date Received
    E: Address
    F: CF Number

Rows whose 6 fields are all filled are skipped on re-runs.
"""

import csv
import io
import json
import os
import platform
import re
import shutil
import subprocess
import sys
import threading
import time
from datetime import date, datetime
from pathlib import Path
from urllib.parse import quote

import requests
import tkinter as tk
from tkinter import filedialog, messagebox, scrolledtext, simpledialog, ttk

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

# Selenium is optional - only needed for the FCI / Chrome-attach features.
try:
    from selenium import webdriver
    from selenium.webdriver.chrome.options import Options
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support import expected_conditions as EC
    from selenium.webdriver.support.ui import WebDriverWait
    HAS_SELENIUM = True
except ImportError:
    HAS_SELENIUM = False

# PyMuPDF (fitz) for rendering PDF pages, pytesseract + Pillow for OCR of
# scanned draw-bulk sheets. All optional — only needed if you parse PDFs.
try:
    import fitz  # PyMuPDF
    HAS_FITZ = True
except ImportError:
    HAS_FITZ = False

try:
    from PIL import Image
    import pytesseract
    HAS_OCR = True
except ImportError:
    HAS_OCR = False


# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------

APP_DIR = Path(__file__).resolve().parent
CONFIG_PATH = APP_DIR / "config.json"

DEFAULT_CONFIG = {
    "sheet_url": "",
    "today_sheet_name": "",           # blank -> auto MMDD
    "draw_sheet_name": "Draw",
    "output_xlsx": str(APP_DIR / "balance_update.xlsx"),
    "chrome_path": "",                # blank -> auto-detect
    "chrome_debug_port": 9222,
    "chrome_user_data_dir": str(APP_DIR / "chrome-profile"),
    "fci_url": "https://fciweb.myfci.com/loanSearch",
    "fci_mode": "manual",             # legacy
    "fci_auto": True,
    "fci_api_key": "",
    "fci_loan_address_field":  "propertyStreet",
    "fci_funding_amount_field": "originalBalance",
}


def load_config() -> dict:
    if CONFIG_PATH.exists():
        try:
            cfg = json.loads(CONFIG_PATH.read_text())
            for k, v in DEFAULT_CONFIG.items():
                cfg.setdefault(k, v)
            return cfg
        except Exception:
            pass
    return dict(DEFAULT_CONFIG)


def save_config(cfg: dict) -> None:
    CONFIG_PATH.write_text(json.dumps(cfg, indent=2))


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

ABBREV = {
    "street": "st", "avenue": "ave", "road": "rd", "boulevard": "blvd",
    "drive": "dr", "court": "ct", "place": "pl", "lane": "ln",
    "terrace": "ter", "circle": "cir", "highway": "hwy", "parkway": "pkwy",
    "square": "sq", "trail": "trl",
    "north": "n", "south": "s", "east": "e", "west": "w",
    "northeast": "ne", "northwest": "nw", "southeast": "se", "southwest": "sw",
}
DIRS = {"n", "s", "e", "w", "ne", "nw", "se", "sw"}


def addr_tokens(s):
    if s is None:
        return []
    s = str(s).replace("\xa0", " ").lower()
    s = re.sub(r"[^\w\s]", " ", s)
    return [ABBREV.get(t, t) for t in s.split() if t]


def addr_key(s):
    """Return (street-number, first non-directional street word) — a strong
    fuzzy match key that survives 'Ave' vs 'Avenue Dr W' style differences."""
    toks = addr_tokens(s)
    num = None
    street = None
    for t in toks:
        if num is None and any(c.isdigit() for c in t):
            num = t
        elif num is not None and t not in DIRS and street is None:
            street = t
            break
    if num and street:
        return (num, street)
    return None


def addr_norm(s):
    return " ".join(addr_tokens(s))


def parse_amount(val):
    if val is None or val == "":
        return None
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).replace("$", "").replace(",", "").strip()
    if not s:
        return None
    try:
        return float(s)
    except ValueError:
        return None


def parse_date(val):
    if val is None or val == "":
        return None
    if isinstance(val, datetime):
        return val
    if isinstance(val, date):
        return datetime(val.year, val.month, val.day)
    s = str(val).strip()
    if not s:
        return None
    for fmt in ("%m/%d/%Y", "%m/%d/%y", "%Y-%m-%d", "%m-%d-%Y", "%m/%d"):
        try:
            d = datetime.strptime(s, fmt)
            if fmt == "%m/%d":
                d = d.replace(year=datetime.now().year)
            return d
        except ValueError:
            continue
    return None


def extract_sheet_id(url):
    m = re.search(r"/spreadsheets/d/([a-zA-Z0-9_-]+)", url or "")
    return m.group(1) if m else None


def extract_gid(url):
    m = re.search(r"[#?&]gid=(\d+)", url or "")
    return m.group(1) if m else None


def today_sheet_default():
    return datetime.now().strftime("%m%d")


# ---------------------------------------------------------------------------
# Google Sheets fetching
# ---------------------------------------------------------------------------

def _gviz_url(sheet_id, sheet_name=None, gid=None):
    base = f"https://docs.google.com/spreadsheets/d/{sheet_id}/gviz/tq?tqx=out:csv"
    if gid:
        return base + f"&gid={gid}"
    return base + f"&sheet={quote(sheet_name or '', safe='')}"


def list_sheet_tabs(driver, sheet_id):
    """Use the attached Chrome to read tab names from a Google Sheet.
    Returns list of names in display order, or [] if not found."""
    url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/edit"
    main = driver.current_window_handle
    driver.switch_to.new_window("tab")
    try:
        driver.get(url)
        # Tabs render after the sheet loads; poll up to ~15s
        for _ in range(30):
            time.sleep(0.5)
            names = driver.execute_script("""
                // Google Sheets tab buttons - try the obvious selectors
                const sels = ['.docs-sheet-tab-name',
                              '.docs-sheet-tab-caption',
                              '[role="tab"]'];
                for (const s of sels) {
                    const els = document.querySelectorAll(s);
                    if (els.length) {
                        return Array.from(els)
                                    .map(t => (t.textContent || '').trim())
                                    .filter(n => n);
                    }
                }
                return [];
            """)
            if names:
                # de-dup while preserving order
                seen = set()
                out  = []
                for n in names:
                    if n not in seen:
                        seen.add(n); out.append(n)
                return out
        return []
    finally:
        try: driver.close()
        except Exception: pass
        driver.switch_to.window(main)


# ---------------------------------------------------------------------------
# Google Sheets — proper API via OAuth (preferred when set up)
# ---------------------------------------------------------------------------

GOOGLE_SCOPES = ['https://www.googleapis.com/auth/spreadsheets.readonly']

def google_creds_path():
    return APP_DIR / "credentials.json"

def google_token_path():
    return APP_DIR / "token.json"


def get_google_creds(interactive=True):
    """Return authenticated Google credentials.

    Priority:
      1. Service-account key (google-credentials.json) referenced in config —
         the same key Draw Manager uses. No browser popup ever needed.
      2. OAuth token cached from a previous 'Sign in to Google' (token.json).
      3. Fresh OAuth flow via credentials.json (browser popup, one-time).

    Path 1 means most users never need to set up OAuth at all.
    """
    try:
        from google.auth.transport.requests import Request
    except ImportError:
        raise RuntimeError(
            "Google API libraries not installed. "
            "Run: pip install google-auth google-auth-oauthlib google-api-python-client"
        )

    # ---- Path 1: service-account key (same as Draw Manager) ----
    sa_path = _service_account_path()
    if sa_path and sa_path.exists():
        try:
            from google.oauth2 import service_account
            creds = service_account.Credentials.from_service_account_file(
                str(sa_path), scopes=GOOGLE_SCOPES)
            return creds
        except Exception as e:
            pass  # fall through to OAuth paths

    # ---- Path 2: cached OAuth token ----
    try:
        from google.oauth2.credentials import Credentials
        from google_auth_oauthlib.flow import InstalledAppFlow
    except ImportError:
        raise RuntimeError(
            "Google API libraries not installed. "
            "Run: pip install google-auth google-auth-oauthlib google-api-python-client"
        )

    creds = None
    tp = google_token_path()
    if tp.exists():
        try:
            creds = Credentials.from_authorized_user_file(str(tp), GOOGLE_SCOPES)
        except Exception:
            creds = None

    if creds and creds.valid:
        return creds

    if creds and creds.expired and creds.refresh_token:
        try:
            creds.refresh(Request())
            tp.write_text(creds.to_json())
            return creds
        except Exception:
            pass  # fall through to fresh flow

    if not interactive:
        raise RuntimeError(
            "Google credentials not found. Options:\n"
            "  A) Point config.json 'sa_credentials_path' at your "
            "google-credentials.json service-account key (same one Draw Manager uses).\n"
            "  B) Set up OAuth: follow the 'Sign in to Google' setup instructions."
        )

    # ---- Path 3: fresh OAuth flow ----
    cp = google_creds_path()
    if not cp.exists():
        raise RuntimeError(
            f"No Google credentials found.\n\n"
            f"Easiest fix — reuse Draw Manager's service account:\n"
            f"  1. Open settings (gear icon in Draw Manager's Casa card)\n"
            f"  2. The service-account key is already configured there.\n"
            f"     Casa will find it automatically on next Pull.\n\n"
            f"Or set up OAuth (one-time, ~10 min):\n"
            f"  1. Go to https://console.cloud.google.com/\n"
            f"  2. APIs & Services → Credentials → Create Credentials\n"
            f"     → OAuth client ID → Desktop app → Download JSON\n"
            f"  3. Rename to credentials.json, place at: {cp}\n"
            f"  4. Click 'Sign in to Google' in the app."
        )

    flow = InstalledAppFlow.from_client_secrets_file(str(cp), GOOGLE_SCOPES)
    creds = flow.run_local_server(port=0, open_browser=True,
                                  authorization_prompt_message="")
    tp.write_text(creds.to_json())
    return creds


def _service_account_path():
    """Find the service-account key: check config.json for sa_credentials_path,
    then fall back to google-credentials.json next to this script."""
    # Try config.json's sa_credentials_path first
    try:
        cfg = load_config()
        p = cfg.get("sa_credentials_path") or ""
        if p and Path(p).exists():
            return Path(p)
    except Exception:
        pass
    # Fall back to google-credentials.json next to the app
    fallback = APP_DIR / "google-credentials.json"
    if fallback.exists():
        return fallback
    return None


def sheets_api_service(creds):
    from googleapiclient.discovery import build
    return build('sheets', 'v4', credentials=creds, cache_discovery=False)


def list_sheet_tabs_via_api(creds, sheet_id):
    svc = sheets_api_service(creds)
    meta = svc.spreadsheets().get(spreadsheetId=sheet_id,
                                  fields="sheets.properties.title").execute()
    return [s['properties']['title'] for s in meta.get('sheets', [])]


def fetch_sheet_via_api(creds, sheet_id, tab_name):
    svc = sheets_api_service(creds)
    # Use the tab name as a range — returns all values
    resp = svc.spreadsheets().values().get(
        spreadsheetId=sheet_id,
        range=f"'{tab_name}'"
    ).execute()
    return resp.get('values', [])


def fetch_sheet_via_driver(driver, sheet_id, sheet_name=None, gid=None):
    """Fetch a sheet tab as CSV by running fetch() inside the Chrome window
    that the user is signed in to. No cookie extraction, no requests-library
    auth. The CSV comes back through the same session the browser is using.

    Robustness: we land the helper tab on the *actual spreadsheet* URL first
    (guaranteeing the docs.google.com origin + this sheet's auth context),
    wait until the origin really is docs.google.com, then run the fetch with
    a few retries. This avoids the instant "TypeError: Failed to fetch" that
    happens when docs.google.com/ root bounces to another origin."""
    url = _gviz_url(sheet_id, sheet_name=sheet_name, gid=gid)
    sheet_home = f"https://docs.google.com/spreadsheets/d/{sheet_id}/edit"

    result = {"status": 0, "text": ""}
    main = driver.current_window_handle
    driver.switch_to.new_window("tab")
    try:
        # Land on the real spreadsheet so origin + auth are correct.
        try:
            driver.get(sheet_home)
        except Exception:
            driver.get("https://docs.google.com/")

        # Wait until we're genuinely on docs.google.com (not a redirect to
        # accounts/consent on another origin).
        for _ in range(20):
            try:
                host = driver.execute_script("return location.host || '';")
            except Exception:
                host = ""
            if "docs.google.com" in host:
                break
            time.sleep(0.5)

        driver.set_script_timeout(45)
        last_err = ""
        for attempt in range(3):
            time.sleep(0.4 + 0.4 * attempt)
            try:
                result = driver.execute_async_script("""
                    const cb = arguments[arguments.length - 1];
                    fetch(arguments[0], {credentials: 'include'})
                      .then(r => r.text().then(t => cb({status: r.status, text: t})))
                      .catch(e => cb({status: 0, text: String(e)}));
                """, url)
            except Exception as e:
                result = {"status": 0, "text": str(e)}
            if result.get("status", 0) != 0:
                break                       # got a real HTTP response
            last_err = result.get("text", "")
            # Re-land on the sheet before the next try (in case the tab drifted).
            try:
                driver.get(sheet_home)
                time.sleep(0.6)
            except Exception:
                pass
        if result.get("status", 0) == 0 and last_err:
            result["text"] = last_err
    finally:
        try:
            driver.close()
        except Exception:
            pass
        driver.switch_to.window(main)

    status = result.get("status", 0)
    text   = result.get("text", "")

    if status == 0:
        raise RuntimeError(
            f"Browser fetch failed: {text}\n"
            f"Make sure the Chrome window this app opened is signed in to "
            f"Google AND has this exact sheet open at least once "
            f"(open it in that window, let it load, then retry).")
    if status == 401 or status == 403:
        raise RuntimeError(
            f"Google returned HTTP {status}. The Chrome window isn't signed "
            f"in to a Google account that has access to this sheet.\n"
            f"Open the Chrome window this app launched, click in to "
            f"docs.google.com, sign in there, open the sheet to confirm "
            f"it loads, then try again."
        )
    if status == 404:
        raise RuntimeError(
            f"Tab '{sheet_name or gid}' wasn't found in this workbook. "
            f"Check the tab name (it's case-sensitive). Tip: you can paste "
            f"the URL with #gid=… to address a tab by its gid instead."
        )
    if status != 200:
        raise RuntimeError(f"HTTP {status} fetching sheet: {text[:200]}")
    if text.lstrip().startswith("<"):
        raise RuntimeError(
            "Google returned HTML instead of CSV — your session isn't authed. "
            "Sign in to Google in the Chrome window this app opened, then retry."
        )
    return list(csv.reader(io.StringIO(text)))


def fetch_sheet_csv_anon(sheet_id, sheet_name=None, gid=None):
    """Unauthenticated fetch for 'Anyone with the link' sheets."""
    url = _gviz_url(sheet_id, sheet_name=sheet_name, gid=gid)
    r = requests.get(url, timeout=30, allow_redirects=True)
    if r.status_code != 200:
        raise RuntimeError(f"HTTP {r.status_code} (sheet is private — need Chrome sign-in)")
    if r.text.lstrip().startswith("<"):
        raise RuntimeError("Got HTML — sheet is private")
    return list(csv.reader(io.StringIO(r.text)))


# ---------------------------------------------------------------------------
# Sheet parsers
# ---------------------------------------------------------------------------

def parse_today_sheet(rows):
    """The 'today' tab contains an info table that looks like:

        Draw # | Name | Address | AA | BL | Last 4 | Draw Fee |
        Wire Received | OUT | Going out | Released On | Notes

    People often leave some of the header cells blank in real sheets
    (only 'Name', 'Address', 'Last 4' etc. get labels). So we find the
    header row by the position of 'address', then fill in the rest by
    convention: Draw# is two columns left of Address, Name is one left,
    Amount is one right."""
    header_row_idx = None
    addr_col = None
    for i, row in enumerate(rows):
        lowered = [str(c).strip().lower() for c in row[:12]]
        if "address" in lowered and ("name" in lowered or "borrower" in lowered):
            header_row_idx = i
            addr_col = lowered.index("address")
            break

    if header_row_idx is None:
        raise RuntimeError(
            "Couldn't find the info-table header row (need a row with both "
            "'Name' and 'Address' cells in the first 12 columns)."
        )

    header = [str(c).strip().lower() for c in rows[header_row_idx]]

    def col_for(*names):
        for n in names:
            if n in header:
                return header.index(n)
        return None

    c_addr = addr_col
    c_name = col_for("name", "borrower")
    c_draw = col_for("draw #", "draw#", "loan #", "draw")
    c_amt  = col_for("aa", "amount", "approved amount", "aa amount", "aa $", "$")
    c_released = col_for("released on", "released", "release date", "date released")

    # Positional fallback — order is Draw#, Name, Address, AA, BL, Last4,
    # Draw Fee, Wire Received, OUT, Going out, Released On, Notes
    if c_name is None and c_addr >= 1:     c_name     = c_addr - 1
    if c_draw is None and c_addr >= 2:     c_draw     = c_addr - 2
    if c_amt  is None:                     c_amt      = c_addr + 1
    if c_released is None:                 c_released = c_addr + 8

    properties = []
    blanks = 0
    for row in rows[header_row_idx + 1:]:
        if not any(str(c).strip() for c in row):
            blanks += 1
            if blanks >= 4:        # likely end of table
                break
            continue
        blanks = 0
        addr = row[c_addr] if c_addr < len(row) else ""
        amt  = row[c_amt]  if c_amt  is not None and c_amt  < len(row) else ""
        draw = row[c_draw] if c_draw is not None and c_draw < len(row) else ""
        name = row[c_name] if c_name is not None and c_name < len(row) else ""
        released = (row[c_released] if c_released is not None and c_released < len(row)
                    else "")

        addr_str = str(addr or "").strip()
        if not addr_str:
            continue
        if addr_str.lower() in {"reminders", "notes", "summary"}:
            break

        amt_val = parse_amount(amt)
        if amt_val is None:
            continue
        try:
            draw_num = int(float(str(draw).strip())) if str(draw).strip() else None
        except ValueError:
            draw_num = None

        properties.append({
            "draw_num":    draw_num,
            "name":        str(name or "").strip(),
            "address":     addr_str.replace("\xa0", " "),
            "amount":      amt_val,
            "released_on": parse_date(released),
        })
    return properties


def build_draw_index(rows):
    """Build address → (cf_number, draw_row_data) index from the Draw sheet.

    The Draw sheet layout:
        Row 1: section group headers (Draw One/Two/...)
        Row 2: column headers (Loan#, Name, Address, …, Funded for each draw)
        Row 3+: data rows

    For each draw N (1-based), the 'Funded' column sits at index 6 + 8*N
    (1-based) — i.e. col N for draw 1, V for draw 2, AD for draw 3, …
    """
    if len(rows) < 3:
        return {}, []

    header = [str(c).strip().lower() for c in rows[1]]

    def find_col(name):
        for i, h in enumerate(header):
            if h == name:
                return i
        return None

    cf_col   = find_col("loan #")     # col A — CF number
    addr_col = find_col("address")    # col C
    if cf_col is None or addr_col is None:
        # Fallback to the literal positions documented above
        cf_col, addr_col = 0, 2

    index = {}
    flat = []
    for row in rows[2:]:
        if len(row) <= max(cf_col, addr_col):
            continue
        cf   = str(row[cf_col]   if row[cf_col]   is not None else "").strip()
        addr = str(row[addr_col] if row[addr_col] is not None else "").strip().replace("\xa0", " ")
        if not cf or not addr:
            continue
        k = addr_key(addr)
        entry = {"cf": cf, "address": addr, "row": row}
        flat.append(entry)
        if k:
            index.setdefault(k, []).append(entry)
    return index, flat


def lookup_draw(index, flat, query_addr):
    """Return matching entry from the Draw sheet, or None."""
    k = addr_key(query_addr)
    qtoks = set(addr_tokens(query_addr))
    candidates = index.get(k, []) if k else []
    if len(candidates) == 1:
        return candidates[0]
    if len(candidates) > 1:
        return max(candidates,
                   key=lambda c: len(qtoks & set(addr_tokens(c["address"]))))
    # Last-ditch fallback: substring search of the normalized query
    qnorm = addr_norm(query_addr)
    for entry in flat:
        if qnorm and qnorm in addr_norm(entry["address"]):
            return entry
    return None


def funded_value(draw_row, draw_num=None):
    """Return the LATEST funded date across all of this property's draws.
    Funded cells sit at 1-based columns 14, 22, 30, 38, ... — i.e. 6 + 8*N
    for draw N (1..11). draw_num is accepted for backwards compatibility
    but ignored: per the user's spec, we always want the most recent date,
    not the one for a specific draw."""
    latest = None
    for n in range(1, 12):
        col_idx0 = (6 + 8 * n) - 1
        if col_idx0 >= len(draw_row):
            break
        d = parse_date(draw_row[col_idx0])
        if d and (latest is None or d > latest):
            latest = d
    return latest


# ---------------------------------------------------------------------------
# PDF parsing (scanned / exported draw-bulk sheets)
# ---------------------------------------------------------------------------

_MONEY_RE = re.compile(r"\$?\s?\d[\d,]*\.\d{2}")
_DATE_RE  = re.compile(r"\b\d{1,2}/\d{1,2}(?:/\d{2,4})?\b")
_NAME_SUFFIX = {"llc", "inc", "corp", "co", "trust", "pa", "ltd", "group",
                "company", "homes", "solutions", "properties", "investments",
                "ventures", "marketing", "consulting", "builders", "services"}


def _pdf_check_deps():
    """Raise a friendly error if the PDF libraries aren't installed."""
    missing = []
    if not HAS_FITZ:
        missing.append("PyMuPDF")
    if not HAS_OCR:
        missing.append("pytesseract + Pillow")
    if missing:
        raise RuntimeError(
            "PDF parsing needs: " + ", ".join(missing) + ".\n"
            "Run:  pip install PyMuPDF pytesseract pillow\n"
            "Scanned PDFs also need the Tesseract OCR engine installed:\n"
            "  Windows: https://github.com/UB-Mannheim/tesseract/wiki\n"
            "  (after install, Tesseract must be on your PATH)."
        )
    # Confirm the tesseract binary is reachable for scanned PDFs.
    try:
        pytesseract.get_tesseract_version()
    except Exception:
        raise RuntimeError(
            "Tesseract OCR engine not found on PATH.\n"
            "Install it (Windows: https://github.com/UB-Mannheim/tesseract/wiki) "
            "then restart this app. Needed for scanned PDFs."
        )


def _pdf_render(page, dpi=300):
    mat = fitz.Matrix(dpi / 72, dpi / 72)
    pix = page.get_pixmap(matrix=mat)
    return Image.open(io.BytesIO(pix.tobytes("png")))


def _ocr_score(t):
    return (len(_MONEY_RE.findall(t)) + len(_DATE_RE.findall(t)) +
            sum(t.lower().count(k) for k in ("address", "draw", "released", "going")))


def _pdf_best_orientation(img):
    """Scanned faxes/scans come in any rotation. Try 0/90/180/270 and keep
    whichever yields the most money/date/keyword tokens."""
    best, best_s = img, -1
    for ang in (0, 90, 180, 270):
        r = img.rotate(ang, expand=True)
        s = _ocr_score(pytesseract.image_to_string(r, config="--psm 6"))
        if s > best_s:
            best, best_s = r, s
    return best


def _ocr_lines(img):
    """Return list of lines; each line is a list of {t, cx} word dicts in x order."""
    d = pytesseract.image_to_data(img, config="--psm 6",
                                  output_type=pytesseract.Output.DICT)
    groups = {}
    for i in range(len(d["text"])):
        t = d["text"][i].strip()
        if not t:
            continue
        key = (d["block_num"][i], d["par_num"][i], d["line_num"][i])
        groups.setdefault(key, []).append(
            {"t": t, "cx": d["left"][i] + d["width"][i] / 2})
    lines = []
    for key in sorted(groups):
        words = sorted(groups[key], key=lambda w: w["cx"])
        lines.append(words)
    return lines


def _hdr_anchor(hdr, *names):
    for w in hdr:
        if w["t"].lower().strip(":#") in names:
            return w["cx"]
    return None


def _deglue(s):
    """Best-effort: re-insert spaces OCR dropped (e.g. 'NW29thAve' ->
    'NW 29th Ave', '2021NW' -> '2021 NW'), then re-merge ordinals that the
    split broke apart ('29 th' -> '29th', '171 st' -> '171st')."""
    s = re.sub(r"(?<=[A-Za-z])(?=\d)", " ", s)     # letter -> digit
    s = re.sub(r"(?<=\d)(?=[A-Za-z])", " ", s)     # digit  -> letter
    s = re.sub(r"(?<=[a-z])(?=[A-Z])", " ", s)     # camelCase
    s = re.sub(r"\s+", " ", s).strip()
    s = re.sub(r"\b(\d+)\s+(st|nd|rd|th)\b", r"\1\2", s, flags=re.IGNORECASE)
    return s


def _clean_addr(words):
    toks = [w for w in words if re.sub(r"[^\w]", "", w)]
    # drop leading name-suffix / punctuation tokens until the street number
    while toks and not any(c.isdigit() for c in toks[0]):
        if toks[0].lower().strip(".,|") in _NAME_SUFFIX or not re.search(r"\w", toks[0]):
            toks.pop(0)
        else:
            break
    for i, t in enumerate(toks):
        if any(c.isdigit() for c in t):
            toks = toks[i:]
            break
    s = _deglue(" ".join(toks).replace("|", " "))
    # Drop stray OCR punctuation (—, «, $, ., etc.); keep alphanumerics/spaces.
    s = re.sub(r"[^0-9A-Za-z ]+", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    # Re-merge ordinals once more in case punctuation removal exposed them.
    s = re.sub(r"\b(\d+)\s+(st|nd|rd|th)\b", r"\1\2", s, flags=re.IGNORECASE)
    return s


def parse_pdf_draw_sheet(path, log=None, progress=None):
    """Parse a draw-bulk PDF (the FCI/Toshiba-scanned layout) into the same
    property dicts that parse_today_sheet returns.

    Columns expected (left->right):
        Draw # | Name | Address | AA | BL | Last 4 | Draw Fee |
        Wire Received | OUT | Going out | Released On

    'AA' is taken as the Amount and 'Released On' as the date. The result is
    a best-effort OCR draft — review it in the preview table before saving.
    """
    _pdf_check_deps()

    def _log(m):
        if log:
            log(m)

    doc = fitz.open(path)
    n_pages = len(doc)
    props = []
    fname = os.path.basename(path)

    for pi, page in enumerate(doc, 1):
        if progress:
            progress(pi - 1, n_pages, f"OCR '{fname}' page {pi}/{n_pages}")
        # Prefer a real text layer if present (digital PDFs); else OCR.
        raw = page.get_text().strip()
        img = _pdf_render(page)
        if len(raw) < 40:                # essentially no text layer -> scanned
            img = _pdf_best_orientation(img)
        lines = _ocr_lines(img)

        # Find the header row.
        hdr = hidx = None
        for i, ln in enumerate(lines):
            j = " ".join(w["t"].lower() for w in ln)
            if "address" in j and "released" in j and ("draw" in j or "name" in j):
                hdr, hidx = ln, i
                break
        if hdr is None:
            _log(f"  '{fname}' p{pi}: no header row found — skipped.")
            continue

        a_addr = _hdr_anchor(hdr, "address")
        a_aa   = _hdr_anchor(hdr, "aa")
        a_name = _hdr_anchor(hdr, "name")
        a_rel  = _hdr_anchor(hdr, "released", "on")
        xs = sorted(w["cx"] for w in hdr)
        if a_addr is None and len(xs) >= 3:
            a_addr = xs[2]
        if a_aa is None and a_addr is not None:
            a_aa = a_addr + 190

        page_rows = 0
        for ln in lines[hidx + 1:]:
            if a_addr is None:
                break
            left_b  = (a_name + a_addr) / 2 if a_name else a_addr - 160
            right_b = (a_addr + a_aa) / 2 if a_aa else a_addr + 160
            addr = _clean_addr([w["t"] for w in ln if left_b < w["cx"] < right_b])

            monies = [(w["cx"], _MONEY_RE.search(w["t"]).group())
                      for w in ln if _MONEY_RE.search(w["t"])]
            aa = (min(monies, key=lambda m: abs(m[0] - a_aa))[1]
                  if (monies and a_aa) else None)

            dates = [(w["cx"], _DATE_RE.search(w["t"]).group())
                     for w in ln if _DATE_RE.search(w["t"])]
            rel = None
            if dates:
                rel = (min(dates, key=lambda d: abs(d[0] - a_rel))[1]
                       if a_rel else dates[-1][1])

            if not addr or not re.search(r"\d", addr) or not aa:
                continue
            props.append({
                "draw_num":    None,
                "name":        "",
                "address":     addr,
                "amount":      parse_amount(aa),
                "released_on": parse_date(rel),
                "_source_pdf": fname,
            })
            page_rows += 1
        _log(f"  '{fname}' p{pi}: {page_rows} row(s).")

    if progress:
        progress(n_pages, n_pages, f"Parsed '{fname}'")
    _log(f"  '{fname}': {len(props)} total row(s) parsed.")
    return props


# ---------------------------------------------------------------------------
# Date tagging (for per-date output files)
# ---------------------------------------------------------------------------

def date_tag_from(value):
    """Return an 'MMDD' tag from a date/datetime/str, or None."""
    d = parse_date(value) if not isinstance(value, (date, datetime)) else value
    if isinstance(value, (date, datetime)):
        d = value
    if d is None:
        return None
    if isinstance(d, datetime):
        return d.strftime("%m%d")
    if isinstance(d, date):
        return datetime(d.year, d.month, d.day).strftime("%m%d")
    return None


def tab_date_tag(tab_name):
    """Turn a tab name like '06/18' or '6-3' into '0618' / '0603'."""
    if not tab_name:
        return None
    m = re.search(r"(\d{1,2})\s*[/\-.]\s*(\d{1,2})", str(tab_name))
    if m:
        return f"{int(m.group(1)):02d}{int(m.group(2)):02d}"
    return None


# ---------------------------------------------------------------------------
# Excel I/O
# ---------------------------------------------------------------------------

EXCEL_HEADER = ["Loan Number", "Investor Account Number", "Amount",
                "Date Received", "Address", "CF Number"]


def load_existing_rows(path):
    """Return (list_of_dicts, max_row_used). Missing/empty/corrupt file is fine."""
    if not os.path.exists(path) or os.path.getsize(path) == 0:
        return [], 0
    try:
        wb = load_workbook(path)
    except Exception as e:
        # Corrupt/half-written xlsx — back it up so the next run starts fresh.
        try:
            bak = path + ".corrupt.bak"
            os.replace(path, bak)
            print(f"[load_existing_rows] {path} unreadable ({e}); moved to {bak}")
        except Exception:
            pass
        return [], 0
    ws = wb.active
    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not any(c is not None and str(c).strip() != "" for c in r):
            continue
        row = list(r) + [None] * (6 - len(r))
        rows.append({
            "loan_number": row[0],
            "account_number": row[1],
            "amount": row[2],
            "date_received": row[3],
            "address": row[4],
            "cf_number": row[5],
        })
    return rows, ws.max_row


def is_row_complete(row: dict) -> bool:
    keys = ("loan_number", "account_number", "amount",
            "date_received", "address", "cf_number")
    return all(row.get(k) not in (None, "") for k in keys)


def write_rows(path, rows, *, append=True):
    """Write rows to the balance-update xlsx.  If `append` and file exists,
    the existing sheet is preserved and new rows are added at the bottom."""
    if append and os.path.exists(path) and os.path.getsize(path) > 0:
        try:
            wb = load_workbook(path)
            ws = wb.active
            start_row = ws.max_row + 1
        except Exception:
            wb = Workbook()
            ws = wb.active
            ws.append(EXCEL_HEADER)
            start_row = 2
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(EXCEL_HEADER)
        start_row = 2

    for i, r in enumerate(rows):
        rr = start_row + i
        ws.cell(row=rr, column=1, value=r.get("loan_number"))
        ws.cell(row=rr, column=2, value=r.get("account_number"))
        ws.cell(row=rr, column=3, value=r.get("amount"))
        d = r.get("date_received")
        if isinstance(d, datetime):
            ws.cell(row=rr, column=4, value=d).number_format = "m/d/yyyy"
        else:
            ws.cell(row=rr, column=4, value=d)
        ws.cell(row=rr, column=5, value=r.get("address"))
        ws.cell(row=rr, column=6, value=r.get("cf_number"))

    for col, width in zip("ABCDEF", (14, 18, 12, 14, 32, 12)):
        ws.column_dimensions[col].width = width

    wb.save(path)


# ---------------------------------------------------------------------------
# Chrome / Selenium
# ---------------------------------------------------------------------------

def detect_chrome_path():
    if platform.system() == "Windows":
        candidates = [
            r"C:\Program Files\Google\Chrome\Application\chrome.exe",
            r"C:\Program Files (x86)\Google\Chrome\Application\chrome.exe",
            os.path.expandvars(r"%LOCALAPPDATA%\Google\Chrome\Application\chrome.exe"),
        ]
    elif platform.system() == "Darwin":
        candidates = ["/Applications/Google Chrome.app/Contents/MacOS/Google Chrome"]
    else:
        candidates = ["/usr/bin/google-chrome", "/usr/bin/chromium", "/usr/bin/chromium-browser"]
    for c in candidates:
        if os.path.exists(c):
            return c
    found = shutil.which("chrome") or shutil.which("google-chrome") or shutil.which("chromium")
    return found or ""


def _port_listening(host, port, timeout=0.5):
    import socket
    try:
        s = socket.create_connection((host, port), timeout=timeout)
        s.close()
        return True
    except OSError:
        return False


def _wait_for_port(host, port, total_timeout=20):
    deadline = time.time() + total_timeout
    while time.time() < deadline:
        if _port_listening(host, port):
            return True
        time.sleep(0.3)
    return False


def launch_chrome_debug(cfg, extra_tabs=()):
    chrome = cfg.get("chrome_path") or detect_chrome_path()
    if not chrome:
        raise RuntimeError("Could not find Chrome on this system. "
                           "Set chrome_path in config.json.")
    port = int(cfg.get("chrome_debug_port", 9222))
    headless = bool(cfg.get("chrome_headless"))

    # If something is already on the port, reuse it — no second instance.
    if _port_listening("127.0.0.1", port):
        return None

    profile = cfg.get("chrome_user_data_dir") or str(APP_DIR / "chrome-profile")
    os.makedirs(profile, exist_ok=True)

    initial_url = cfg.get("fci_url", "https://fciweb.myfci.com/loanSearch")
    args = [chrome,
            f"--remote-debugging-port={port}",
            f"--user-data-dir={profile}",
            "--no-first-run", "--no-default-browser-check",
            "--disable-background-timer-throttling",
            "--disable-backgrounding-occluded-windows"]
    if headless:
        # Headless reuses whatever session is saved in `profile`. If FCI/Google
        # aren't already logged in there, headless lookups will fail — sign in
        # once in visible mode first.
        args += ["--headless=new", "--disable-gpu",
                 "--window-size=1600,1100", "--no-sandbox"]
    args += [initial_url, *extra_tabs]

    if platform.system() == "Windows":
        flags = (subprocess.CREATE_NEW_PROCESS_GROUP
                 | subprocess.DETACHED_PROCESS
                 | 0x08000000)  # CREATE_NO_WINDOW — no console flash
        proc = subprocess.Popen(args, creationflags=flags, close_fds=True)
    else:
        proc = subprocess.Popen(args, start_new_session=True)

    if not _wait_for_port("127.0.0.1", port, total_timeout=25):
        raise RuntimeError(
            f"Chrome opened but never bound to debug port {port}.\n"
            f"Most common cause: your normal Chrome is already running and is "
            f"using the automation profile.\n"
            f"Fix options:\n"
            f"  1) Fully quit your regular Chrome (check the system tray), then "
            f"retry; or\n"
            f"  2) This tool uses its own profile at:\n     {profile}\n"
            f"     so it shouldn't collide — if it still fails, a leftover "
            f"Chrome may be holding that folder. Open Task Manager, end all "
            f"chrome.exe processes, and retry."
        )
    return proc


def attach_to_chrome(port=9222):
    if not HAS_SELENIUM:
        raise RuntimeError("selenium is not installed. `pip install selenium`")
    if not _port_listening("127.0.0.1", port):
        raise RuntimeError(
            f"Nothing is listening on 127.0.0.1:{port}. Click 'Launch Chrome' "
            f"first, then make sure the Chrome window stays open."
        )
    opts = Options()
    opts.add_experimental_option("debuggerAddress", f"127.0.0.1:{port}")
    return webdriver.Chrome(options=opts)


def open_url_in_new_tab(driver, url):
    driver.switch_to.new_window("tab")
    driver.get(url)


# ---------------------------------------------------------------------------
# FCI GraphQL API
# ---------------------------------------------------------------------------

FCI_API_URL = "https://fapi.myfci.com/graphql"


def fci_api_call(api_key, query, variables=None, timeout=30):
    """POST a GraphQL query to FCI. Returns the `data` block on success.
    Raises RuntimeError with a useful message on any failure."""
    if not api_key:
        raise RuntimeError("FCI API key is empty — set it in the app.")
    headers = {
        "Authorization": f"Bearer {api_key}",
        "Content-Type":  "application/json",
        "Accept":        "application/json",
    }
    payload = {"query": query}
    if variables:
        payload["variables"] = variables

    r = requests.post(FCI_API_URL, headers=headers, json=payload, timeout=timeout)

    if r.status_code == 401:
        raise RuntimeError("FCI API 401: token rejected. Either it's wrong, "
                           "expired, or this endpoint isn't enabled for your account.")
    if r.status_code == 429:
        raise RuntimeError("FCI API 429: rate-limited. Slow down or wait.")
    if r.status_code >= 500:
        raise RuntimeError(f"FCI API {r.status_code}: server error. Try again.")
    if r.status_code != 200:
        raise RuntimeError(f"FCI API HTTP {r.status_code}: {r.text[:300]}")

    try:
        body = r.json()
    except ValueError:
        raise RuntimeError(f"FCI API non-JSON response: {r.text[:300]}")

    if isinstance(body, dict) and body.get("errors"):
        msgs = "; ".join(e.get("message", str(e)) for e in body["errors"])
        raise RuntimeError(f"FCI GraphQL errors: {msgs}")

    return body.get("data", body) if isinstance(body, dict) else body


def fci_introspect(api_key):
    """Lightweight connectivity + auth check for the 'Test API' button.

    The FCI server has GraphQL schema introspection disabled (HC0046), so we
    can't ask for __schema. Instead we run a real, minimal query against the
    same endpoint the actual lookups use — getLoanInformation, requesting only
    loanAccount — which confirms the key authenticates and the API answers.
    Returns a small dict the summary function describes."""
    q = "query{getLoanInformation(includeNoProperty:false){loanAccount}}"
    data = fci_api_call(api_key, q)
    loans = data.get("getLoanInformation", []) if isinstance(data, dict) else []
    if not isinstance(loans, list):
        raise RuntimeError("Unexpected response: getLoanInformation did not "
                           "return a list. The key authenticated but the API "
                           "shape was unexpected.")
    return {"loan_count": len(loans)}


def fci_introspect_summary(result):
    """Human-readable result of the connectivity check above."""
    if isinstance(result, dict) and "loan_count" in result:
        n = result["loan_count"]
        return (f"  Connected. Your key can read the FCI portfolio "
                f"({n} loan{'s' if n != 1 else ''} visible).")
    return "  Connected."


def fci_load_portfolio(api_key, address_field="propertyAddress", log=None):
    """Pull the loan portfolio. Returns list of {loanAccount, address} dicts.
    The exact GraphQL field name for the property address must be passed in
    — discoverable via Test API. Default 'propertyAddress' is a guess."""
    q = (f"query{{getLoanInformation(includeNoProperty:false)"
         f"{{loanAccount {address_field}}}}}")
    data = fci_api_call(api_key, q)
    loans = data.get("getLoanInformation", [])
    if not isinstance(loans, list):
        raise RuntimeError(f"getLoanInformation returned non-list: "
                           f"{json.dumps(data, default=str)[:200]}")
    if log: log(f"      portfolio: {len(loans)} loans loaded")
    return [{"loanAccount": str(l.get("loanAccount") or ""),
             "address":     str(l.get(address_field) or "")}
            for l in loans]


def fci_funding_for_loan(api_key, loan_account, amount_field="amountFunded", log=None):
    """Returns (best_account, best_amount, rows) for a single loan.

    Picks the PRINCIPAL LENDER — the investor that holds the most of the loan.

    FCI's getFundingHistory exposes three real fields per row:
        lenderAccount, principalBalance, originalBalance
    A single loan often has MULTIPLE rows per investor (each funding tranche
    is a separate row). The correct ownership signal is the SUM of
    principalBalance across all rows for each lender, not any single row.

    Algorithm:
      1. Group by lenderAccount, sum principalBalance per account.
      2. The account with the largest total principal wins.
      3. If every principal is zero, fall back to summed originalBalance.
      4. The configured `amount_field` (legacy: defaults to originalBalance)
         is no longer used to pick — but we still return an amount for
         logging purposes.
    """
    q = ("query Q($la:String!){getFundingHistory(loanaccount:$la)"
         "{lenderAccount principalBalance originalBalance}}")
    data = fci_api_call(api_key, q, {"la": str(loan_account)})
    rows = data.get("getFundingHistory", [])
    if not isinstance(rows, list):
        raise RuntimeError(f"getFundingHistory returned non-list: "
                           f"{json.dumps(data, default=str)[:200]}")
    if not rows:
        return None, 0.0, []

    def _num(v):
        n = parse_amount(v)
        return n if isinstance(n, (int, float)) else 0.0

    # Sum principal + original per lenderAccount.
    by_acct = {}  # acct -> {"principal": float, "original": float, "rows": int}
    for r in rows:
        acct = r.get("lenderAccount")
        if not acct:
            continue
        agg = by_acct.setdefault(acct, {"principal": 0.0, "original": 0.0, "rows": 0})
        agg["principal"] += _num(r.get("principalBalance"))
        agg["original"]  += _num(r.get("originalBalance"))
        agg["rows"]      += 1

    if not by_acct:
        return None, 0.0, rows

    # Pick by total principal; if all zero, fall back to total original.
    have_principal = any(a["principal"] > 0 for a in by_acct.values())
    sort_key = (lambda kv: kv[1]["principal"]) if have_principal \
               else (lambda kv: kv[1]["original"])
    ranked = sorted(by_acct.items(), key=sort_key, reverse=True)
    winner_acct, winner = ranked[0]
    winner_amt = winner["principal"] if have_principal else winner["original"]

    if log and len(by_acct) > 1:
        log(f"      {len(by_acct)} distinct investor(s) on {loan_account}; "
            f"picked {winner_acct} (principal ${winner['principal']:,.2f} "
            f"across {winner['rows']} row(s))")
        for acct, agg in ranked:
            mark = "→" if acct == winner_acct else " "
            log(f"        {mark} acct={acct}  principal=${agg['principal']:,.2f}  "
                f"original=${agg['original']:,.2f}  rows={agg['rows']}")

    return winner_acct, winner_amt, rows


def fci_api_lookup_auto(api_key, address, portfolio_cache, log=None,
                        amount_field="amountFunded"):
    """address -> loan # (from cached portfolio) -> account # (from funding)."""
    result = {"loan_number": None, "account_number": None, "error": None}

    target_key = addr_key(address)
    target_tokens = set(addr_tokens(address))
    match = None
    fallbacks = []
    for loan in portfolio_cache:
        addr = loan.get("address") or ""
        if not addr:
            continue
        k = addr_key(addr)
        if target_key and k == target_key:
            match = loan; break
        overlap = len(target_tokens & set(addr_tokens(addr)))
        if overlap >= 2:
            fallbacks.append((overlap, loan))
    if match is None and fallbacks:
        fallbacks.sort(key=lambda x: -x[0])
        match = fallbacks[0][1]

    if match is None:
        result["error"] = f"no loan in portfolio matched '{address}'"
        return result

    loan_account = match.get("loanAccount")
    if not loan_account:
        result["error"] = "matched loan has no loanAccount"
        return result

    result["loan_number"] = loan_account
    if log: log(f"      loan # = {loan_account}")

    try:
        acct, amt, _ = fci_funding_for_loan(api_key, loan_account,
                                            amount_field=amount_field, log=log)
    except RuntimeError as e:
        result["error"] = f"funding query failed: {e}"
        return result

    if not acct:
        result["error"] = "no investor rows on funding for this loan"
        return result
    result["account_number"] = acct
    if log: log(f"      account # = {acct}  (amount funded ${amt:,.2f})")
    return result


def fci_api_lookup_many(api_key, properties, log=None, on_result=None,
                        address_field="propertyAddress",
                        amount_field="amountFunded", workers=4):
    """Batch lookup. Loads portfolio once, then runs the per-property funding
    queries in parallel (they're independent HTTP calls). `workers` caps
    concurrency; on_result is called as each finishes (thread-safe via lock).
    Each property retries once on a transient error (e.g. 429/timeout)."""
    import threading, time as _t
    from concurrent.futures import ThreadPoolExecutor, as_completed

    if log: log("      loading FCI portfolio…")
    portfolio = fci_load_portfolio(api_key, address_field=address_field, log=log)

    results = []
    lock = threading.Lock()

    def _safe_log(msg):
        if log:
            with lock:
                log(msg)

    def work(p):
        last_err = None
        for attempt in (1, 2):  # one retry on transient failure
            try:
                res = fci_api_lookup_auto(api_key, p["address"], portfolio,
                                          log=None, amount_field=amount_field)
                # retry only if the funding query itself transiently failed
                err = (res or {}).get("error") or ""
                if attempt == 1 and ("429" in err or "timed out" in err.lower()
                                     or "server error" in err.lower()):
                    last_err = err
                    _t.sleep(1.2)
                    continue
                return p, res
            except Exception as e:
                last_err = f"{type(e).__name__}: {e}"
                if attempt == 1:
                    _t.sleep(1.2)
                    continue
                return p, {"loan_number": None, "account_number": None,
                           "error": last_err}
        return p, {"loan_number": None, "account_number": None, "error": last_err}

    n = max(1, min(int(workers or 1), 12))
    if log: log(f"      looking up {len(properties)} propert"
                f"{'y' if len(properties)==1 else 'ies'} ({n} in parallel)…")
    with ThreadPoolExecutor(max_workers=n) as ex:
        futures = [ex.submit(work, p) for p in properties]
        for fut in as_completed(futures):
            p, res = fut.result()
            # per-property visibility (thread-safe)
            addr = (p.get("address") or "")[:38]
            if res.get("loan_number") and res.get("account_number"):
                _safe_log(f"      ✓ {addr}  loan {res['loan_number']} / acct {res['account_number']}")
            else:
                _safe_log(f"      ✗ {addr}  — {res.get('error') or 'no result'}")
            with lock:
                results.append((p, res))
                if on_result:
                    try: on_result(p, res)
                    except Exception: pass
    return results


# Field-name candidate lists used by the Test API discovery step.
# The first hit wins — known-good fields are listed first.
LOAN_ADDRESS_CANDIDATES = [
    "propertyStreet",  # confirmed working as of Jun 2026
    "propertyAddress", "address", "PropertyAddress", "Address",
    "streetAddress", "loanAddress", "Property", "propAddress",
    "propertyaddress", "fullAddress", "addressLine1", "addressLine",
    "address1", "street", "propStreetAddr", "property",
]
FUNDING_AMOUNT_CANDIDATES = [
    "originalBalance",   # = "Amount Funded" in FCI's UI; confirmed Jun 2026
    "principalBalance",  # current outstanding (also confirmed but not what we want)
    "amountFunded", "amount", "fundedAmount", "investmentAmount",
    "purchasedAmount", "balance", "currentBalance", "originalAmount",
    "fundingAmount", "principal", "investorAmount", "accountBalance",
    "amountfunded", "AmountFunded",
    "fundedBalance", "balanceFunded",
]


# ---------------------------------------------------------------------------
# FCI lookup (browser, Selenium-based — fallback if API isn't configured)
# ---------------------------------------------------------------------------

def fci_lookup_auto(driver, fci_url, address, log=None, timeout=20, debug_dir=None):
    """Search FCI for an address and walk INFO -> Loan Details -> Funding,
    returning {loan_number, account_number, error}.  Even when `error` is
    set, any data we DID manage to extract (e.g. just the loan #) is still
    returned — partial successes are preferable to dropped data.
    """
    from selenium.webdriver.common.keys import Keys
    from selenium.webdriver.common.by import By
    from selenium.webdriver.common.action_chains import ActionChains
    from selenium.webdriver.support import expected_conditions as EC
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.common.exceptions import (
        TimeoutException, NoSuchElementException, StaleElementReferenceException,
        ElementClickInterceptedException
    )

    def _log(m):
        if log: log(m)

    addr_token = re.sub(r"\s+", " ", address).strip()
    parts = [p for p in re.split(r"\s+", addr_token.lower()) if p]
    num   = next((p for p in parts if any(c.isdigit() for c in p)), parts[0] if parts else "")
    dirs  = {"n","s","e","w","ne","nw","se","sw"}
    street = next((p for p in parts[parts.index(num)+1:] if p not in dirs), "") if num in parts else ""
    addr_match_substr = (num + " " + street).strip().lower() if street else num.lower()

    # Result that we mutate as we make progress; whatever we got stays even
    # if a later step blows up.
    result = {"loan_number": None, "account_number": None, "error": None}

    def dump_debug(tag):
        if debug_dir is None:
            return
        try:
            os.makedirs(debug_dir, exist_ok=True)
            stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            safe  = re.sub(r"[^a-zA-Z0-9]+", "_", address)[:40]
            base  = os.path.join(debug_dir, f"{stamp}_{tag}_{safe}")
            try: driver.save_screenshot(base + ".png")
            except Exception: pass
            try:
                with open(base + ".html", "w", encoding="utf-8") as f:
                    f.write(driver.page_source)
            except Exception: pass
            _log(f"      [debug] saved {base}.html and .png")
        except Exception:
            pass

    try:
        driver.get(fci_url)

        addr_in = WebDriverWait(driver, timeout).until(
            EC.element_to_be_clickable((By.ID, "Address"))
        )
        addr_in.clear()
        addr_in.send_keys(address)
        addr_in.send_keys(Keys.RETURN)

        for sel in ("button[type='submit']",
                    "button.k-button.k-primary",
                    "button.btn-primary",
                    "button[aria-label*='Search' i]"):
            try:
                btn = driver.find_element(By.CSS_SELECTOR, sel)
                if btn.is_displayed() and btn.is_enabled():
                    btn.click()
                    break
            except Exception:
                continue

        # ---- Wait until a result row contains our address (stale-safe) ----
        def address_present(d):
            try:
                rows = d.find_elements(By.CSS_SELECTOR, "tr.k-master-row")
            except StaleElementReferenceException:
                return False
            for r in rows:
                try:
                    addr_cell = r.find_element(By.CSS_SELECTOR, "td[aria-colindex='4']")
                    if addr_match_substr in addr_cell.text.lower():
                        return True
                except (NoSuchElementException, StaleElementReferenceException):
                    continue
            return False

        try:
            WebDriverWait(driver, timeout,
                          ignored_exceptions=(StaleElementReferenceException,
                                              NoSuchElementException)
                          ).until(address_present)
        except TimeoutException:
            dump_debug("no_match")
            result["error"] = (f"no result row matching '{addr_match_substr}' "
                               f"appeared within {timeout}s")
            return result

        # ---- Read loan # from the matching row (retry on stale) ----
        loan_number = None
        for _ in range(5):
            try:
                for r in driver.find_elements(By.CSS_SELECTOR, "tr.k-master-row"):
                    try:
                        addr_cell = r.find_element(By.CSS_SELECTOR, "td[aria-colindex='4']")
                        if addr_match_substr not in addr_cell.text.lower():
                            continue
                        text = r.find_element(By.CSS_SELECTOR, "td[aria-colindex='2']").text.strip()
                        if text:
                            loan_number = text
                            break
                    except (NoSuchElementException, StaleElementReferenceException):
                        continue
                if loan_number:
                    break
                time.sleep(0.25)
            except StaleElementReferenceException:
                time.sleep(0.25)

        if not loan_number:
            dump_debug("blank_loan")
            result["error"] = "couldn't extract loan # from matching row"
            return result

        result["loan_number"] = loan_number   # SAVE NOW so partial-success is preserved
        _log(f"      loan # = {loan_number}")

        # ---- Click INFO via JavaScript so DOM lookup + click are atomic ----
        clicked = driver.execute_script("""
            const want = (arguments[0] || '').trim();
            const rows = document.querySelectorAll('tr.k-master-row');
            for (const r of rows) {
                const cell = r.querySelector("td[aria-colindex='2']");
                if (!cell) continue;
                if (cell.textContent.trim() === want) {
                    const btn = r.querySelector('button.dropdown-toggle, button.btn-success');
                    if (btn) { btn.click(); return true; }
                }
            }
            return false;
        """, loan_number)

        if not clicked:
            dump_debug("no_info_button")
            result["error"] = f"couldn't click INFO for loan {loan_number}"
            return result

        # ---- 'Loan Details' menu item ----
        ld_xpath = ("//*[self::span or self::a or self::div or self::li or self::button]"
                    "[normalize-space(translate(text(),"
                    " 'ABCDEFGHIJKLMNOPQRSTUVWXYZ',"
                    " 'abcdefghijklmnopqrstuvwxyz'))='loan details']")
        try:
            loan_details = WebDriverWait(driver, timeout,
                ignored_exceptions=(StaleElementReferenceException,)
                ).until(EC.element_to_be_clickable((By.XPATH, ld_xpath)))
        except TimeoutException:
            dump_debug("no_loan_details")
            result["error"] = "'Loan Details' didn't appear after INFO click"
            return result
        try:
            loan_details.click()
        except (ElementClickInterceptedException, StaleElementReferenceException):
            try:
                ActionChains(driver).move_to_element(loan_details).click().perform()
            except Exception:
                # last resort — JS click
                driver.execute_script("arguments[0].click()", loan_details)

        # ---- 'Funding' link ----
        funding_xpath = (
            "//a[@href='/LenFunding' or contains(@href,'Funding')]"
            " | //*[self::a or self::button or self::span][normalize-space(translate(text(),"
            " 'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz'))='funding']"
        )
        try:
            funding = WebDriverWait(driver, timeout,
                ignored_exceptions=(StaleElementReferenceException,)
                ).until(EC.element_to_be_clickable((By.XPATH, funding_xpath)))
        except TimeoutException:
            dump_debug("no_funding_link")
            result["error"] = "no 'Funding' link after Loan Details"
            return result
        try:
            funding.click()
        except (ElementClickInterceptedException, StaleElementReferenceException):
            try:
                ActionChains(driver).move_to_element(funding).click().perform()
            except Exception:
                driver.execute_script("arguments[0].click()", funding)

        # ---- Funding table: pick row with max Amount Funded ----
        try:
            WebDriverWait(driver, timeout,
                ignored_exceptions=(StaleElementReferenceException,)
                ).until(EC.presence_of_element_located((By.CSS_SELECTOR, "tr.k-master-row")))
        except TimeoutException:
            dump_debug("no_funding_rows")
            result["error"] = "no rows on Funding page"
            return result
        time.sleep(0.4)

        best_acct, best_amt = None, -1.0
        for _ in range(3):
            try:
                for row in driver.find_elements(By.CSS_SELECTOR, "tr.k-master-row"):
                    try:
                        cells = row.find_elements(By.TAG_NAME, "td")
                        if len(cells) < 6:
                            continue
                        acct = cells[2].text.strip()
                        amt  = parse_amount(cells[4].text.strip()) or 0.0
                        if acct and amt > best_amt:
                            best_amt, best_acct = amt, acct
                    except StaleElementReferenceException:
                        continue
                break
            except StaleElementReferenceException:
                time.sleep(0.25)

        if best_acct is None:
            dump_debug("no_account_rows")
            result["error"] = "no investor rows on Funding page"
            return result

        result["account_number"] = best_acct
        _log(f"      account # = {best_acct}  (amount funded ${best_amt:,.2f})")
        return result

    except Exception as e:
        dump_debug("exception")
        result["error"] = f"{type(e).__name__}: {str(e)[:200]}"
        return result


def _to_int_or_keep(s):
    """Helper: '399609356' -> 399609356, '$1,234' -> 1234, 'abc' -> 'abc'."""
    if s is None or s == "":
        return s
    if isinstance(s, (int, float)):
        return int(s)
    clean = re.sub(r"[^\d-]", "", str(s))
    try:
        return int(clean) if clean else s
    except ValueError:
        return s


def fci_lookup_many(properties, port, fci_url, *, workers=1, log=None,
                    debug_dir=None, on_result=None):
    """Look up several properties in FCI.

    workers=1  : sequential, using a single tab (the main driver).
    workers>1  : open `workers` extra Chrome tabs, one per parallel worker.
                 Each worker gets its own WebDriver session attached to the
                 same Chrome — separate session state per worker means each
                 can have its own active window. WARNING: if FCI keeps the
                 'active loan' in server-side session state (the /LenFunding
                 URL has no loan ID), parallel workers may corrupt each
                 other's flow. Try workers=2 first to verify.

    `on_result(prop, res)` is called from worker threads as each lookup
    finishes — use it to update the UI incrementally.
    """
    import queue as _queue
    import threading as _th

    work = _queue.Queue()
    for p in properties:
        work.put(p)
    results = []
    results_lock = _th.Lock()

    def _run(drv, owns_tab):
        while True:
            try:
                p = work.get_nowait()
            except _queue.Empty:
                return
            try:
                if log: log(f"  FCI > {p['address']}")
                res = fci_lookup_auto(drv, fci_url, p["address"],
                                       log=log, debug_dir=debug_dir)
            except Exception as e:
                res = {"loan_number": None, "account_number": None,
                       "error": f"worker exception: {e}"}
            with results_lock:
                results.append((p, res))
            try:
                if on_result:
                    on_result(p, res)
            except Exception:
                pass

    # Sequential — reuse the existing driver/tab.
    if workers <= 1:
        drv = attach_to_chrome(port)
        _run(drv, owns_tab=False)
        return results

    # Parallel — extra drivers + tabs.
    extra_drivers = []
    extra_tabs    = []
    try:
        for _ in range(workers):
            d = attach_to_chrome(port)
            d.switch_to.new_window("tab")
            extra_drivers.append(d)
            extra_tabs.append(d.current_window_handle)

        threads = []
        for d in extra_drivers:
            t = _th.Thread(target=_run, args=(d, True), daemon=True)
            t.start()
            threads.append(t)
        for t in threads:
            t.join()
    finally:
        # Close each parallel worker's tab.
        for d, h in zip(extra_drivers, extra_tabs):
            try:
                d.switch_to.window(h)
                d.close()
            except Exception:
                pass
    return results


def fci_lookup_manual(driver, fci_url, prop, parent_window):
    """Open FCI loan search in Chrome, prefill the search box with the
    borrower name/address, and pop up a dialog for the user to paste the
    Loan# + Investor Account# from FCI."""
    if driver is not None:
        try:
            driver.get(fci_url)
        except Exception:
            pass

    dlg = tk.Toplevel(parent_window)
    dlg.title(f"FCI: {prop['address']}")
    dlg.transient(parent_window)
    dlg.grab_set()
    dlg.resizable(False, False)

    info = (f"Borrower:  {prop.get('name') or '(unknown)'}\n"
            f"Address:   {prop['address']}\n"
            f"Amount:    {prop.get('amount')}\n"
            f"CF #:      {prop.get('cf_number') or '(not found)'}")
    tk.Label(dlg, text=info, justify="left", font=("Consolas", 10)).pack(
        padx=14, pady=(14, 8), anchor="w")

    frame = ttk.Frame(dlg)
    frame.pack(padx=14, pady=4, fill="x")

    ttk.Label(frame, text="Loan Number:").grid(row=0, column=0, sticky="w", pady=2)
    e_loan = ttk.Entry(frame, width=24)
    e_loan.grid(row=0, column=1, padx=8, pady=2)

    ttk.Label(frame, text="Investor Account #:").grid(row=1, column=0, sticky="w", pady=2)
    e_acct = ttk.Entry(frame, width=24)
    e_acct.grid(row=1, column=1, padx=8, pady=2)

    ttk.Label(frame, text="Date Received (m/d/yyyy):").grid(row=2, column=0, sticky="w", pady=2)
    e_date = ttk.Entry(frame, width=24)
    if isinstance(prop.get("date_received"), datetime):
        e_date.insert(0, prop["date_received"].strftime("%m/%d/%Y"))
    e_date.grid(row=2, column=1, padx=8, pady=2)

    result = {"action": "skip"}

    def on_ok():
        result["loan"] = e_loan.get().strip()
        result["acct"] = e_acct.get().strip()
        result["date"] = e_date.get().strip()
        result["action"] = "ok"
        dlg.destroy()

    def on_skip():
        result["action"] = "skip"
        dlg.destroy()

    def on_stop():
        result["action"] = "stop"
        dlg.destroy()

    btn_frame = ttk.Frame(dlg)
    btn_frame.pack(pady=10)
    ttk.Button(btn_frame, text="Save & Next",   command=on_ok).pack(side="left", padx=4)
    ttk.Button(btn_frame, text="Skip",          command=on_skip).pack(side="left", padx=4)
    ttk.Button(btn_frame, text="Stop",          command=on_stop).pack(side="left", padx=4)

    e_loan.focus_set()
    dlg.wait_window()
    return result


# ---------------------------------------------------------------------------
# GUI
# ---------------------------------------------------------------------------

class CasaApp:
    def __init__(self, root):
        self.root = root
        self.cfg = load_config()
        self.driver = None
        self.properties = []          # rows pulled from Sheets and/or PDFs
        self.existing = []            # rows already in the output xlsx
        self.pdf_paths = []           # PDFs queued for parsing
        self.detected_tabs = []       # tabs detected from the workbook

        root.title("Draw Bulk Upload --- FCI")
        root.geometry("1040x820")
        root.minsize(900, 700)

        self._build_ui()
        self._refresh_today_default()

    # ---- UI ---------------------------------------------------------------

    def _build_ui(self):
        pad = {"padx": 8, "pady": 4}

        top = ttk.LabelFrame(self.root, text="Source & destination")
        top.pack(fill="x", padx=10, pady=(10, 6))

        ttk.Label(top, text="Google Sheet URL:").grid(row=0, column=0, sticky="e", **pad)
        self.v_sheet = tk.StringVar(value=self.cfg.get("sheet_url", ""))
        ttk.Entry(top, textvariable=self.v_sheet, width=80).grid(row=0, column=1, columnspan=3, sticky="we", **pad)

        ttk.Label(top, text="Draw sheet name:").grid(row=1, column=0, sticky="e", **pad)
        self.v_draw = tk.StringVar(value=self.cfg.get("draw_sheet_name", "Draw"))
        ttk.Entry(top, textvariable=self.v_draw, width=14).grid(row=1, column=1, sticky="w", **pad)
        ttk.Label(top, text="(used only for CF# lookup)").grid(
            row=1, column=2, columnspan=2, sticky="w", **pad)

        ttk.Label(top, text="Output .xlsx:").grid(row=2, column=0, sticky="e", **pad)
        self.v_out = tk.StringVar(value=self.cfg.get("output_xlsx", ""))
        ttk.Entry(top, textvariable=self.v_out, width=70).grid(row=2, column=1, columnspan=2, sticky="we", **pad)
        ttk.Button(top, text="Browse…", command=self._browse_out).grid(row=2, column=3, sticky="w", **pad)

        ttk.Label(top, text="FCI API key:").grid(row=3, column=0, sticky="e", **pad)
        self.v_api_key = tk.StringVar(value=self.cfg.get("fci_api_key", ""))
        ttk.Entry(top, textvariable=self.v_api_key, show="•", width=70).grid(
            row=3, column=1, columnspan=2, sticky="we", **pad)
        ttk.Button(top, text="Test API",
                   command=self.on_test_api).grid(row=3, column=3, sticky="w", **pad)

        top.columnconfigure(1, weight=1)

        # ---- Sources: pick sheet tabs and/or add PDFs ---------------------
        src = ttk.LabelFrame(self.root, text="Sources to parse (sheets and/or PDFs)")
        src.pack(fill="x", padx=10, pady=6)

        # Left: detected sheet tabs (multi-select)
        sheets_f = ttk.Frame(src)
        sheets_f.pack(side="left", fill="both", expand=True, padx=8, pady=6)
        hdr = ttk.Frame(sheets_f); hdr.pack(fill="x")
        ttk.Label(hdr, text="Sheet tabs:").pack(side="left")
        ttk.Button(hdr, text="Detect tabs",
                   command=self.on_detect_tabs).pack(side="right")
        lb_wrap = ttk.Frame(sheets_f); lb_wrap.pack(fill="both", expand=True, pady=(4, 0))
        self.lb_tabs = tk.Listbox(lb_wrap, selectmode="extended", height=5,
                                  exportselection=False)
        self.lb_tabs.pack(side="left", fill="both", expand=True)
        tsb = ttk.Scrollbar(lb_wrap, orient="vertical", command=self.lb_tabs.yview)
        self.lb_tabs.configure(yscrollcommand=tsb.set); tsb.pack(side="right", fill="y")
        ovr = ttk.Frame(sheets_f); ovr.pack(fill="x", pady=(4, 0))
        ttk.Label(ovr, text="…or type tab(s), comma-separated:").pack(side="left")
        self.v_today = tk.StringVar(value=self.cfg.get("today_sheet_name", ""))
        ttk.Entry(ovr, textvariable=self.v_today, width=22).pack(
            side="left", padx=6, fill="x", expand=True)

        ttk.Separator(src, orient="vertical").pack(side="left", fill="y", pady=6)

        # Right: PDF queue
        pdf_f = ttk.Frame(src)
        pdf_f.pack(side="left", fill="both", expand=True, padx=8, pady=6)
        phdr = ttk.Frame(pdf_f); phdr.pack(fill="x")
        ttk.Label(phdr, text="PDFs:").pack(side="left")
        ttk.Button(phdr, text="Remove", command=self.on_remove_pdf).pack(side="right")
        ttk.Button(phdr, text="Add PDF(s)…", command=self.on_add_pdfs).pack(side="right", padx=4)
        pl_wrap = ttk.Frame(pdf_f); pl_wrap.pack(fill="both", expand=True, pady=(4, 0))
        self.lb_pdfs = tk.Listbox(pl_wrap, selectmode="extended", height=5,
                                  exportselection=False)
        self.lb_pdfs.pack(side="left", fill="both", expand=True)
        psb = ttk.Scrollbar(pl_wrap, orient="vertical", command=self.lb_pdfs.yview)
        self.lb_pdfs.configure(yscrollcommand=psb.set); psb.pack(side="right", fill="y")

        actions = ttk.LabelFrame(self.root, text="Actions")
        actions.pack(fill="x", padx=10, pady=6)

        ttk.Button(actions, text="1. Launch Chrome (debug)",
                   command=self.on_launch_chrome).pack(side="left", padx=6, pady=8)
        ttk.Button(actions, text="Sign in to Google (API)",
                   command=self.on_google_signin).pack(side="left", padx=6, pady=8)
        ttk.Button(actions, text="2. Pull / Parse selected",
                   command=self.on_pull).pack(side="left", padx=6, pady=8)
        ttk.Button(actions, text="3. Run FCI lookup",
                   command=self.on_fci).pack(side="left", padx=6, pady=8)
        ttk.Button(actions, text="4. Save / append to Excel",
                   command=self.on_save).pack(side="left", padx=6, pady=8)
        ttk.Separator(actions, orient="vertical").pack(side="left", fill="y", padx=8, pady=8)
        ttk.Button(actions, text="\u25B6  Run everything",
                   command=self.on_run_all).pack(side="left", padx=6, pady=8)
        ttk.Button(actions, text="Open output folder",
                   command=self.on_open_folder).pack(side="right", padx=6, pady=8)
        ttk.Button(actions, text="Open Excel",
                   command=self.on_open_excel).pack(side="right", padx=6, pady=8)
        self.v_auto_fci = tk.BooleanVar(value=bool(self.cfg.get("fci_auto", True)))
        ttk.Checkbutton(actions, text="Auto FCI lookup",
                        variable=self.v_auto_fci).pack(side="right", padx=10)

        # Preview table
        prev = ttk.LabelFrame(self.root, text="Properties")
        prev.pack(fill="both", expand=True, padx=10, pady=6)

        cols = ("status", "draw", "address", "amount", "cf", "loan", "acct", "date")
        self.tree = ttk.Treeview(prev, columns=cols, show="headings", height=12)
        for c, txt, w in [
            ("status", "✓", 26), ("draw", "Draw#", 50), ("address", "Address", 240),
            ("amount", "Amount", 90), ("cf", "CF #", 70), ("loan", "Loan #", 110),
            ("acct", "Acct #", 90), ("date", "Date", 90),
        ]:
            self.tree.heading(c, text=txt)
            self.tree.column(c, width=w, anchor="w")
        self.tree.pack(fill="both", expand=True, side="left")
        sb = ttk.Scrollbar(prev, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=sb.set)
        sb.pack(side="right", fill="y")
        # Double-click a cell to fix OCR/import mistakes before saving.
        self.tree.bind("<Double-1>", self._on_tree_edit)
        ttk.Label(self.root,
                  text="Tip: double-click Address / Amount / CF / Loan / Acct / Date "
                       "to edit a value before saving.",
                  foreground="#666").pack(anchor="w", padx=14)

        # ---- Progress -----------------------------------------------------
        prog_f = ttk.Frame(self.root)
        prog_f.pack(fill="x", padx=12, pady=(6, 0))
        self.v_status = tk.StringVar(value="Ready.")
        ttk.Label(prog_f, textvariable=self.v_status).pack(side="left")
        self.v_progress = tk.DoubleVar(value=0.0)
        self.progress = ttk.Progressbar(prog_f, orient="horizontal",
                                        mode="determinate", maximum=100.0,
                                        variable=self.v_progress, length=320)
        self.progress.pack(side="right")

        # Log
        log_f = ttk.LabelFrame(self.root, text="Log")
        log_f.pack(fill="both", expand=False, padx=10, pady=(6, 10))
        self.log_widget = scrolledtext.ScrolledText(log_f, height=8, wrap="word",
                                                    font=("Consolas", 9))
        self.log_widget.pack(fill="both", expand=True)

    def _refresh_today_default(self):
        # Tabs are now chosen via the multi-select list; nothing to prefill.
        return

    # ---- Logging ---------------------------------------------------------

    def log(self, msg):
        ts = datetime.now().strftime("%H:%M:%S")
        self.log_widget.insert("end", f"[{ts}] {msg}\n")
        self.log_widget.see("end")
        self.root.update_idletasks()

    # ---- Progress (thread-safe) -----------------------------------------

    def set_progress(self, done, total, msg=None):
        def _apply():
            pct = (100.0 * done / total) if total else 0.0
            self.v_progress.set(pct)
            if msg is not None:
                self.v_status.set(f"{msg}  ({done}/{total})" if total else msg)
        try:
            self.root.after(0, _apply)
        except Exception:
            pass

    def status(self, msg):
        try:
            self.root.after(0, lambda: self.v_status.set(msg))
        except Exception:
            pass

    # ---- Sources: tabs + PDFs -------------------------------------------

    def _selected_tabs(self):
        """Tabs the user chose: listbox selection, else the comma override,
        else []."""
        sel = [self.lb_tabs.get(i) for i in self.lb_tabs.curselection()]
        if sel:
            return sel
        override = self.v_today.get().strip()
        if override:
            return [t.strip() for t in override.split(",") if t.strip()]
        return []

    def on_detect_tabs(self):
        threading.Thread(target=self._detect_tabs_thread, daemon=True).start()

    def _detect_tabs_thread(self):
        try:
            self._save_config()
            self.status("Detecting workbook tabs…")
            sheet_id = extract_sheet_id(self.v_sheet.get().strip())
            if not sheet_id:
                raise RuntimeError("Paste a valid Google Sheet URL first.")
            use_api = google_creds_path().exists() and google_token_path().exists()
            tabs = []
            if use_api:
                try:
                    creds = get_google_creds(interactive=False)
                    tabs = list_sheet_tabs_via_api(creds, sheet_id)
                    self.log("Detected tabs via Sheets API.")
                except Exception as e:
                    self.log(f"API tab list failed: {e} — trying Chrome.")
                    use_api = False
            if not use_api:
                drv = self._ensure_driver(auto_launch=True)
                tabs = list_sheet_tabs(drv, sheet_id)
                self.log("Detected tabs via Chrome.")
            self.detected_tabs = tabs

            def _fill():
                self.lb_tabs.delete(0, "end")
                for t in tabs:
                    self.lb_tabs.insert("end", t)
                # Auto-select every tab except obvious non-data ones.
                skip = ("draw", "approved", "template", "summary", "reminders",
                        "notes", "master", "sheet1")
                any_sel = False
                for i, t in enumerate(tabs):
                    if t.strip().lower() not in skip:
                        self.lb_tabs.selection_set(i)
                        any_sel = True
                if not any_sel:                      # fallback: select all
                    self.lb_tabs.selection_set(0, "end")
                self.lb_tabs.see(0)
            self.root.after(0, _fill)
            self.log(f"Found {len(tabs)} tab(s): {tabs}")
            self.status(f"Detected {len(tabs)} tab(s) — auto-selected. "
                        f"Deselect any you don't want, then Pull / Parse.")
        except Exception as e:
            self.log(f"ERROR: {e}")
            self.status("Tab detection failed.")
            messagebox.showerror("Detect tabs failed", str(e))

    def on_add_pdfs(self):
        paths = filedialog.askopenfilenames(
            title="Select PDF(s) to parse",
            filetypes=[("PDF files", "*.pdf"), ("All files", "*.*")])
        for p in paths:
            if p and p not in self.pdf_paths:
                self.pdf_paths.append(p)
                self.lb_pdfs.insert("end", os.path.basename(p))
        if paths:
            self.status(f"{len(self.pdf_paths)} PDF(s) queued.")

    def on_remove_pdf(self):
        for i in reversed(self.lb_pdfs.curselection()):
            self.lb_pdfs.delete(i)
            del self.pdf_paths[i]
        self.status(f"{len(self.pdf_paths)} PDF(s) queued.")

    # ---- Inline preview editing -----------------------------------------

    _EDITABLE = {"address": "address", "amount": "amount", "cf": "cf_number",
                 "loan": "loan_number", "acct": "account_number", "date": "date_received"}

    def _on_tree_edit(self, event):
        if self.tree.identify("region", event.x, event.y) != "cell":
            return
        col_id = self.tree.identify_column(event.x)
        row_id = self.tree.identify_row(event.y)
        if not row_id:
            return
        col_index = int(col_id[1:]) - 1
        cols = ("status", "draw", "address", "amount", "cf", "loan", "acct", "date")
        if col_index < 0 or col_index >= len(cols):
            return
        col = cols[col_index]
        if col not in self._EDITABLE:
            return
        try:
            idx = int(row_id)
        except ValueError:
            return
        p = self.properties[idx]
        field = self._EDITABLE[col]
        cur = p.get(field)
        if isinstance(cur, datetime):
            cur = cur.strftime("%m/%d/%Y")
        x, y, w, h = self.tree.bbox(row_id, col_id)
        var = tk.StringVar(value="" if cur is None else str(cur))
        ent = ttk.Entry(self.tree, textvariable=var)
        ent.place(x=x, y=y, width=w, height=h)
        ent.focus_set()
        ent.select_range(0, "end")

        def commit(_=None):
            val = var.get().strip()
            if field == "amount":
                p[field] = parse_amount(val)
            elif field == "date_received":
                p[field] = parse_date(val) or (val or None)
            elif field in ("loan_number", "account_number"):
                p[field] = _to_int_or_keep(val) if val else None
            else:
                p[field] = val or None
            ent.destroy()
            self._refresh_tree()

        ent.bind("<Return>", commit)
        ent.bind("<FocusOut>", commit)
        ent.bind("<Escape>", lambda e: ent.destroy())


    # ---- Persisting config ----------------------------------------------

    def _save_config(self):
        self.cfg.update({
            "sheet_url":         self.v_sheet.get().strip(),
            "today_sheet_name":  self.v_today.get().strip(),
            "draw_sheet_name":   self.v_draw.get().strip() or "Draw",
            "output_xlsx":       self.v_out.get().strip(),
            "fci_auto":          bool(self.v_auto_fci.get()),
            "fci_api_key":       self.v_api_key.get().strip(),
        })
        save_config(self.cfg)

    # ---- Browse output --------------------------------------------------

    def _browse_out(self):
        initial = self.v_out.get() or str(APP_DIR / "balance_update.xlsx")
        p = filedialog.asksaveasfilename(
            title="Output Excel file",
            initialfile=os.path.basename(initial),
            initialdir=os.path.dirname(initial) or ".",
            defaultextension=".xlsx",
            filetypes=[("Excel workbook", "*.xlsx")],
        )
        if p:
            self.v_out.set(p)

    # ---- Step 1: Chrome --------------------------------------------------

    def on_launch_chrome(self):
        threading.Thread(target=self._launch_chrome_thread, daemon=True).start()

    def _launch_chrome_thread(self):
        try:
            self._save_config()
            sheet_url = self.v_sheet.get().strip()
            extra = [sheet_url] if sheet_url else []
            self.log("Launching Chrome in debug mode…")
            launch_chrome_debug(self.cfg, extra_tabs=extra)
            self.log(f"Chrome is up on port {self.cfg['chrome_debug_port']}.")
            self.log("Sign in to Google + FCI in that Chrome window. "
                     "Open the Google Sheet there too so you know auth works. "
                     "Then come back and click 'Pull from Sheet'.")
            self._ensure_driver()
        except Exception as e:
            self.log(f"ERROR: {e}")
            messagebox.showerror("Chrome launch failed", str(e))

    def _ensure_driver(self, auto_launch=False):
        if self.driver is not None:
            return self.driver
        port = int(self.cfg["chrome_debug_port"])
        if not _port_listening("127.0.0.1", port):
            if not auto_launch:
                raise RuntimeError(
                    "Chrome isn't open yet. Click '1. Launch Chrome' first, "
                    "sign in to Google + FCI in the window that opens, then try again."
                )
            self.log("Chrome isn't running — launching it now…")
            sheet_url = self.v_sheet.get().strip()
            launch_chrome_debug(self.cfg, extra_tabs=[sheet_url] if sheet_url else [])
        self.log("Attaching to Chrome…")
        self.driver = attach_to_chrome(port)
        self.log("Attached.")
        return self.driver

    # ---- Step 2: pull -----------------------------------------------------

    def on_pull(self):
        threading.Thread(target=self._pull_thread, daemon=True).start()

    def _pull_thread(self):
        try:
            self._save_config()
            sheet_url  = self.v_sheet.get().strip()
            draw_name  = self.v_draw.get().strip() or "Draw"
            sel_tabs   = self._selected_tabs()
            pdf_paths  = list(self.pdf_paths)

            if not sel_tabs and not pdf_paths:
                raise RuntimeError(
                    "Nothing selected. Detect & select sheet tab(s) and/or "
                    "add PDF(s), then click Pull / Parse.")

            have_sheets = bool(sel_tabs and sheet_url)
            total_steps = len(sel_tabs) + len(pdf_paths) + (1 if have_sheets else 0)
            step = 0
            self.set_progress(0, total_steps, "Starting…")

            all_properties = []
            idx, flat = {}, []

            # ---- Sheets -----------------------------------------------------
            creds = drv = None
            use_api = False
            if have_sheets:
                sheet_id = extract_sheet_id(sheet_url)
                if not sheet_id:
                    raise RuntimeError("That doesn't look like a Google Sheets URL.")
                use_api = google_creds_path().exists() and google_token_path().exists()
                if use_api:
                    try:
                        creds = get_google_creds(interactive=False)
                        self.log("Using Google Sheets API.")
                    except Exception as e:
                        self.log(f"API creds problem: {e} — falling back to Chrome.")
                        use_api = False
                if not use_api:
                    drv = self._ensure_driver(auto_launch=True)
                    self.log("Using Chrome session.")

                def fetch_tab(name):
                    if use_api:
                        return fetch_sheet_via_api(creds, sheet_id, name)
                    return fetch_sheet_via_driver(drv, sheet_id, sheet_name=name)

                self.log(f"Parsing {len(sel_tabs)} sheet tab(s): {sel_tabs}")
                for tab in sel_tabs:
                    self.set_progress(step, total_steps, f"Sheet tab '{tab}'")
                    try:
                        rows = fetch_tab(tab)
                        props = parse_today_sheet(rows)
                        for p in props:
                            p["_source_tab"] = tab
                        self.log(f"  '{tab}': {len(rows)} rows, {len(props)} properties.")
                        all_properties.extend(props)
                    except Exception as e:
                        self.log(f"  '{tab}': skipped — {e}")
                    step += 1

                # Draw sheet for CF# lookup
                self.set_progress(step, total_steps, f"Fetching '{draw_name}'")
                try:
                    draw_rows = fetch_tab(draw_name)
                    idx, flat = build_draw_index(draw_rows)
                    self.log(f"  '{draw_name}': {len(draw_rows)} rows, "
                             f"{sum(len(v) for v in idx.values())} addresses indexed.")
                except Exception as e:
                    self.log(f"  couldn't fetch Draw sheet for CF#: {e}")
                step += 1
            elif sel_tabs and not sheet_url:
                self.log("Tabs selected but no Sheet URL — ignoring tabs.")

            # ---- PDFs -------------------------------------------------------
            for path in pdf_paths:
                fname = os.path.basename(path)
                self.set_progress(step, total_steps, f"Parsing PDF '{fname}'")

                def pdf_progress(done, tot, msg, _base=step):
                    # fine-grained page status; overall bar advances per-PDF
                    self.status(msg)
                try:
                    pdf_props = parse_pdf_draw_sheet(path, log=self.log,
                                                     progress=pdf_progress)
                    all_properties.extend(pdf_props)
                except Exception as e:
                    self.log(f"  PDF '{fname}' failed: {e}")
                    messagebox.showerror("PDF parse failed", f"{fname}:\n\n{e}")
                step += 1

            self.log(f"Total properties: {len(all_properties)}")
            if not all_properties:
                self.set_progress(total_steps, total_steps, "Nothing parsed.")
                self.properties = []
                self._refresh_tree()
                return

            # ---- Enrich: CF# from Draw (if available); date = released. ----
            for p in all_properties:
                if idx or flat:
                    match = lookup_draw(idx, flat, p["address"])
                    p["cf_number"] = match["cf"] if match else p.get("cf_number")
                else:
                    p.setdefault("cf_number", None)
                p["date_received"] = p.get("released_on")
                p["_date_tag"] = self._date_tag_for(p)

            # ---- Dedupe vs existing master output --------------------------
            self.existing, _ = load_existing_rows(self.v_out.get().strip())
            existing_by_key = {
                (addr_norm(r.get("address") or ""), parse_amount(r.get("amount"))): r
                for r in self.existing
            }
            for p in all_properties:
                key = (addr_norm(p["address"]), p.get("amount"))
                ex = existing_by_key.get(key)
                if ex and is_row_complete(ex):
                    p["_status"] = "skip"
                else:
                    if ex:
                        p.setdefault("loan_number",    ex.get("loan_number"))
                        p.setdefault("account_number", ex.get("account_number"))
                    p["_status"] = "pending"

            self.properties = all_properties
            self.root.after(0, self._refresh_tree)
            self.set_progress(total_steps, total_steps, "Pull / parse complete.")
            self.log("Pull / parse complete. Review the table, then Run FCI / Save.")
        except Exception as e:
            self.log(f"ERROR: {e}")
            self.status("Pull / parse failed.")
            messagebox.showerror("Pull failed", str(e))

    def _refresh_tree(self):
        self.tree.delete(*self.tree.get_children())
        for i, p in enumerate(self.properties):
            status = "✓" if p.get("_status") == "skip" else (
                "●" if p.get("_status") == "done" else "…")
            amt = p.get("amount")
            amt_s = f"${amt:,.2f}" if isinstance(amt, (int, float)) else ""
            d = p.get("date_received")
            d_s = d.strftime("%m/%d/%Y") if isinstance(d, datetime) else (d or "")
            self.tree.insert("", "end", iid=str(i),
                values=(status, p.get("draw_num") or "", p.get("address", ""),
                        amt_s, p.get("cf_number") or "",
                        p.get("loan_number") or "", p.get("account_number") or "",
                        d_s))

    # ---- Step 3: FCI -----------------------------------------------------

    def on_fci(self):
        if not self.properties:
            messagebox.showinfo("Nothing to look up",
                                "Pull properties from the Sheet first.")
            return
        threading.Thread(target=self._fci_thread, daemon=True).start()

    def _fci_thread(self):
        try:
            self._save_config()
            api_key = self.v_api_key.get().strip()
            if not api_key:
                messagebox.showerror(
                    "No FCI API key",
                    "Enter your FCI API key in the field at the top of the "
                    "window, click 'Test API' once, then try again.")
                return

            use_auto   = bool(self.v_auto_fci.get())
            addr_field = self.cfg.get("fci_loan_address_field")  or "propertyStreet"
            amt_field  = self.cfg.get("fci_funding_amount_field") or "originalBalance"

            to_lookup = [p for p in self.properties
                         if p.get("_status") != "skip"
                         and not all(p.get(k) for k in ("loan_number","account_number"))]

            failed = []
            total_lk = len(to_lookup)
            done_lk = [0]
            self.set_progress(0, total_lk or 1, "FCI lookup…")

            def on_result(p, res):
                if res.get("loan_number"):
                    p["loan_number"] = _to_int_or_keep(res["loan_number"])
                if res.get("account_number"):
                    p["account_number"] = _to_int_or_keep(res["account_number"])
                if res.get("error") or not (p.get("loan_number") and p.get("account_number")):
                    failed.append((p, res.get("error") or "missing fields"))
                else:
                    p["_status"] = "done"
                done_lk[0] += 1
                self.set_progress(done_lk[0], total_lk or 1,
                                  f"FCI: {p.get('address','')[:32]}")
                self.root.after(0, self._refresh_tree)

            if use_auto and to_lookup:
                self.log(f"FCI API: {len(to_lookup)} properties "
                         f"(address={addr_field}, amount={amt_field})…")
                try:
                    fci_api_lookup_many(api_key, to_lookup, log=self.log,
                                        on_result=on_result,
                                        address_field=addr_field,
                                        amount_field=amt_field)
                except Exception as e:
                    self.log(f"API call aborted: {e}")
                    failed = [(p, str(e)) for p in to_lookup]

                ok_count = sum(1 for p in to_lookup if p.get("_status") == "done")
                self.log(f"  {ok_count}/{len(to_lookup)} resolved, "
                         f"{len(failed)} need manual entry.")
            elif not use_auto:
                failed = [(p, "auto disabled") for p in to_lookup]

            # Manual fallback only for the unresolved ones.
            for p, why in failed:
                self.log(f"FCI manual > {p['address']}  ({str(why)[:80]})")
                dlg = self._manual_dialog_threadsafe(p)
                if dlg["action"] == "stop":
                    self.log("Stopped by user."); break
                if dlg["action"] == "skip":
                    continue
                if dlg.get("loan"): p["loan_number"]    = _to_int_or_keep(dlg["loan"])
                if dlg.get("acct"): p["account_number"] = _to_int_or_keep(dlg["acct"])
                if dlg.get("date"): p["date_received"]  = parse_date(dlg["date"]) or dlg["date"]
                p["_status"] = "done"
                self.root.after(0, self._refresh_tree)

            self.set_progress(total_lk or 1, total_lk or 1, "FCI lookup complete.")
            self.log("FCI lookup complete.")
        except Exception as e:
            self.log(f"ERROR: {e}")
            messagebox.showerror("FCI lookup failed", str(e))

    def on_test_api(self):
        threading.Thread(target=self._test_api_thread, daemon=True).start()

    def on_google_signin(self):
        threading.Thread(target=self._google_signin_thread, daemon=True).start()

    def _google_signin_thread(self):
        try:
            self.log("Starting Google OAuth flow…")
            creds = get_google_creds(interactive=True)
            self.log("✓ Signed in to Google. Token saved to token.json.")
            self.log("  You can now run Pull — it will use the Sheets API.")
        except Exception as e:
            self.log(f"ERROR: {e}")
            messagebox.showerror("Google sign-in failed", str(e))

    def _test_api_thread(self):
        try:
            self._save_config()
            api_key = self.v_api_key.get().strip()
            if not api_key:
                messagebox.showerror("No API key", "Enter your FCI API key first.")
                return

            self.log("=" * 60)
            self.log("FCI API discovery")
            self.log("=" * 60)

            # ---- Confirm minimal portfolio query works & get a test loan ----
            self.log("[1/3] Confirming minimal getLoanInformation query…")
            try:
                data = fci_api_call(api_key,
                    "query{getLoanInformation(includeNoProperty:false,limit:1)"
                    "{loanAccount}}")
                loans = data.get("getLoanInformation", [])
                if not loans:
                    self.log("  ✗ portfolio query returned 0 loans — aborting")
                    return
                test_loan = str(loans[0].get("loanAccount") or "")
                if not test_loan:
                    self.log("  ✗ no loanAccount in result — aborting")
                    return
                self.log(f"  ✓ test loan = {test_loan}")
            except RuntimeError as e:
                self.log(f"  ✗ failed: {e}")
                return

            # ---- Probe address field on getLoanInformation -----------------
            self.log("")
            self.log("[2/3] Probing for property-address field on getLoanInformation…")
            found_address = None
            for cand in LOAN_ADDRESS_CANDIDATES:
                q = (f"query{{getLoanInformation(includeNoProperty:false,limit:1)"
                     f"{{loanAccount {cand}}}}}")
                try:
                    data = fci_api_call(api_key, q)
                    loans = data.get("getLoanInformation", [])
                    if loans:
                        value = loans[0].get(cand)
                        self.log(f"  ✓ {cand!r}  →  {value!r}")
                        found_address = cand
                        break
                except RuntimeError as e:
                    err = str(e)
                    if "does not exist" in err.lower():
                        self.log(f"  ✗ {cand}")
                    else:
                        self.log(f"  ? {cand}: {err[:120]}")
            if not found_address:
                self.log("  → no candidate worked. The address field has some")
                self.log("    name we haven't guessed. Ask Lucy for the exact")
                self.log("    field list on CustomLoanInformation.")

            # ---- Probe amount field on getFundingHistory -------------------
            self.log("")
            self.log(f"[3/3] Probing for amount-funded field on getFundingHistory "
                     f"(loan {test_loan})…")
            found_amount = None
            for cand in FUNDING_AMOUNT_CANDIDATES:
                q = (f"query Q($la:String!){{getFundingHistory(loanaccount:$la)"
                     f"{{lenderAccount {cand}}}}}")
                try:
                    data = fci_api_call(api_key, q, {"la": test_loan})
                    rows = data.get("getFundingHistory", [])
                    if rows:
                        value = rows[0].get(cand)
                        self.log(f"  ✓ {cand!r}  →  {value!r}")
                        found_amount = cand
                        break
                except RuntimeError as e:
                    err = str(e)
                    if "does not exist" in err.lower():
                        self.log(f"  ✗ {cand}")
                    else:
                        self.log(f"  ? {cand}: {err[:120]}")
            if not found_amount:
                self.log("  → no candidate worked. Same situation — ask for")
                self.log("    the field list on CustomFundingHistory.")

            # ---- Save what we found ---------------------------------------
            self.log("")
            self.log("=" * 60)
            if found_address: self.cfg["fci_loan_address_field"]   = found_address
            if found_amount:  self.cfg["fci_funding_amount_field"] = found_amount
            save_config(self.cfg)

            if found_address and found_amount:
                self.log("✓ DISCOVERY COMPLETE")
                self.log(f"  loan address field   = {found_address}")
                self.log(f"  funding amount field = {found_amount}")
                self.log("  Saved to config.json. The FCI button should work now.")
            else:
                self.log("Partial discovery. Fields saved:")
                self.log(f"  loan address field   = {found_address or '(not found)'}")
                self.log(f"  funding amount field = {found_amount or '(not found)'}")
                self.log("Run again after adjusting candidates, or share the")
                self.log("FCI data-dictionary excerpts.")
            self.log("=" * 60)
        except Exception as e:
            self.log(f"ERROR: {e}")
            messagebox.showerror("API test failed", str(e))

    # ---- Step 4: save ----------------------------------------------------

    def on_save(self):
        threading.Thread(target=lambda: self._do_save(prompt=True), daemon=True).start()

    def _do_save(self, prompt=False):
        try:
            self._save_config()
            out = self.v_out.get().strip()
            if not out:
                messagebox.showerror("No output path", "Pick an output .xlsx file first.")
                return []

            to_write = []
            for p in self.properties:
                if p.get("_status") == "skip":
                    continue
                to_write.append({
                    "loan_number":    p.get("loan_number"),
                    "account_number": p.get("account_number"),
                    "amount":         p.get("amount"),
                    "date_received":  p.get("date_received"),
                    "address":        p.get("address"),
                    "cf_number":      p.get("cf_number"),
                    "_date_tag":      self._date_tag_for(p),
                })

            if not to_write:
                if prompt:
                    messagebox.showinfo("Nothing to save",
                        "Every property is either already complete in the output "
                        "file, or hasn't been processed yet.")
                else:
                    self.log("Nothing new to write.")
                return []

            self.status("Saving…")
            files = self._save_outputs(out, to_write)
            self.log(f"Saved {len(to_write)} row(s) → master + "
                     f"{len(files)-1} per-date file(s).")
            self.status("Saved.")
            if prompt:
                extra = "\n".join("  • " + os.path.basename(f) for f in files)
                if messagebox.askyesno("Saved",
                        f"Wrote {len(to_write)} row(s) to:\n{extra}\n\n"
                        f"Open the master file now?"):
                    self.on_open_excel()
            return files
        except Exception as e:
            self.log(f"ERROR: {e}")
            self.status("Save failed.")
            messagebox.showerror("Save failed", str(e))
            return []

    def _date_tag_for(self, p):
        """MMDD tag for a property: tab date if it parses, else the row's
        date, else today."""
        tag = tab_date_tag(p.get("_source_tab"))
        if tag:
            return tag
        tag = date_tag_from(p.get("date_received") or p.get("released_on"))
        if tag:
            return tag
        return datetime.now().strftime("%m%d")

    def _save_outputs(self, master_path, rows):
        """Write the master workbook (everything, appended) PLUS one
        balance_update_<MMDD>.xlsx per date, in the master's folder. Returns
        the list of file paths written (master first)."""
        def strip(r):
            return {k: v for k, v in r.items() if k != "_date_tag"}

        written = [master_path]
        self.log(f"[master] {os.path.basename(master_path)}")
        self._merge_and_save(master_path, [strip(r) for r in rows])

        folder = os.path.dirname(master_path) or "."
        groups = {}
        for r in rows:
            groups.setdefault(r.get("_date_tag") or "nodate", []).append(strip(r))
        for tag, grp in sorted(groups.items()):
            per_path = os.path.join(folder, f"balance_update_{tag}.xlsx")
            self.log(f"[{tag}] {os.path.basename(per_path)} ({len(grp)} row(s))")
            self._merge_and_save(per_path, grp)
            written.append(per_path)
        return written

    def _merge_and_save(self, path, new_rows):
        new_rows = [{k: v for k, v in r.items() if k != "_date_tag"}
                    for r in new_rows]
        existing, _ = load_existing_rows(path)
        by_key = {(addr_norm(r["address"] or ""), parse_amount(r["amount"])): r
                  for r in existing}

        merged = list(existing)
        merged_by_key = {(addr_norm(r["address"] or ""), parse_amount(r["amount"])): r
                         for r in merged}

        appended = 0
        updated = 0
        for nr in new_rows:
            key = (addr_norm(nr["address"] or ""), parse_amount(nr["amount"]))
            if key in merged_by_key:
                target = merged_by_key[key]
                # Fill only missing fields — never overwrite values the user
                # has already saved.
                for k, v in nr.items():
                    if not target.get(k) and v:
                        target[k] = v
                updated += 1
            else:
                merged.append(nr)
                merged_by_key[key] = nr
                appended += 1

        # Write from scratch using the merged list
        wb = Workbook()
        ws = wb.active
        ws.append(EXCEL_HEADER)
        for r in merged:
            ws.cell(column=1, row=ws.max_row + 1, value=r.get("loan_number"))
            row_idx = ws.max_row
            ws.cell(column=2, row=row_idx, value=r.get("account_number"))
            ws.cell(column=3, row=row_idx, value=r.get("amount"))
            d = r.get("date_received")
            if isinstance(d, datetime):
                cell = ws.cell(column=4, row=row_idx, value=d)
                cell.number_format = "m/d/yyyy"
            else:
                ws.cell(column=4, row=row_idx, value=d)
            ws.cell(column=5, row=row_idx, value=r.get("address"))
            ws.cell(column=6, row=row_idx, value=r.get("cf_number"))
        for col, width in zip("ABCDEF", (14, 18, 12, 14, 32, 12)):
            ws.column_dimensions[col].width = width
        wb.save(path)
        self.log(f"  → {appended} appended, {updated} updated in place.")

    # ---- "Run everything" ------------------------------------------------

    def _manual_dialog_threadsafe(self, prop):
        """Show the manual FCI dialog on the main thread (Tk widgets are not
        thread-safe). Returns the dialog's result dict; blocks the calling
        thread until the user dismisses the dialog."""
        result = [None]
        done = threading.Event()
        fci_url = self.cfg.get("fci_url", DEFAULT_CONFIG["fci_url"])

        def show():
            try:
                result[0] = fci_lookup_manual(self.driver, fci_url, prop, self.root)
            except Exception as e:
                result[0] = {"action": "skip", "_error": str(e)}
            finally:
                done.set()

        if threading.current_thread() is threading.main_thread():
            show()
        else:
            self.root.after(0, show)
            done.wait()
        return result[0] or {"action": "skip"}

    def on_run_all(self):
        threading.Thread(target=self._run_all_thread, daemon=True).start()

    def _run_all_thread(self):
        try:
            self._save_config()
            if not self.v_out.get().strip():
                raise RuntimeError("Pick an output .xlsx file first.")
            if not self._selected_tabs() and not self.pdf_paths:
                raise RuntimeError(
                    "Nothing selected. Detect & select sheet tab(s) and/or "
                    "add PDF(s) first.")

            self.log("=== Run everything: pull/parse → FCI → save ===")
            self._pull_thread()
            if not self.properties:
                self.log("Nothing parsed — stopping.")
                return
            self._fci_thread()
            self._do_save(prompt=False)
            self.set_progress(1, 1, "Run everything: done.")
            self.log("Run everything: done.")
        except Exception as e:
            self.log(f"ERROR: {e}")
            self.status("Run everything failed.")
            messagebox.showerror("Run failed", str(e))

    # ---- misc ------------------------------------------------------------

    def on_open_excel(self):
        out = self.v_out.get().strip()
        if not out:
            messagebox.showerror("No output file", "Pick an output .xlsx first.")
            return
        if not os.path.exists(out):
            messagebox.showerror("File doesn't exist yet",
                                 f"Click 'Save / append to Excel' first.\n\n{out}")
            return
        try:
            if platform.system() == "Windows":
                os.startfile(out)
            elif platform.system() == "Darwin":
                subprocess.Popen(["open", out])
            else:
                subprocess.Popen(["xdg-open", out])
        except Exception as e:
            self.log(f"Could not open Excel file: {e}")
            messagebox.showerror("Open failed", str(e))

    def on_open_folder(self):
        out = self.v_out.get().strip()
        folder = os.path.dirname(out) if out else str(APP_DIR)
        if not folder or not os.path.isdir(folder):
            folder = str(APP_DIR)
        try:
            if platform.system() == "Windows":
                os.startfile(folder)
            elif platform.system() == "Darwin":
                subprocess.Popen(["open", folder])
            else:
                subprocess.Popen(["xdg-open", folder])
        except Exception as e:
            self.log(f"Could not open folder: {e}")


def main():
    root = tk.Tk()
    try:
        # Pleasant theme where available
        style = ttk.Style()
        for theme in ("vista", "clam", "alt", "default"):
            if theme in style.theme_names():
                style.theme_use(theme)
                break
    except Exception:
        pass
    CasaApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()