"""
Draw Manager — pywebview app shell.

v0.3 additions:
  - Settings restructured into connections{} + scripts{}
  - Google Sheets API integration (service account)
  - Background sync thread with interval + manual sync
  - Per-script settings exposed under scripts[slug]
"""

from __future__ import annotations
import csv
import io
import json
import os
import queue
import re
import sqlite3
import subprocess
import sys
import threading
import time
import uuid
from datetime import datetime, timezone
from pathlib import Path

import webview

APP_DIR       = (Path(sys.executable).parent
                 if getattr(sys, "frozen", False)
                 else Path(__file__).parent)
HTML_PATH     = (Path(sys._MEIPASS) / "draw_manager.html"      # type: ignore[attr-defined]
                 if getattr(sys, "frozen", False)
                 else APP_DIR / "draw_manager.html")
ICON_PATH     = (Path(sys._MEIPASS) / "draw_manager.ico"       # type: ignore[attr-defined]
                 if getattr(sys, "frozen", False)
                 else APP_DIR / "draw_manager.ico")
SETTINGS_FILE = APP_DIR / "settings.json"
DB_PATH       = APP_DIR / "draws.db"
PYTHON        = sys.executable

# ---------------------------------------------------------------------------
# Version + auto-update
# ---------------------------------------------------------------------------
VERSION = "0.1.1"   # <-- bump this, push to main; CI tags a release v<VERSION>

# Your GitHub repo, "owner/name". The updater hits the public Releases API.
# If you rename the repo, change this string (it's the only place it lives).
GITHUB_REPO = "casafinance/draw-manager"

# Names of the two release assets CI uploads. Must match build-release.yml.
_ASSET_MAIN   = "Draw Manager.exe"
_ASSET_WORKER = "draw-request.exe"

# Updater status, polled by the UI titlebar indicator.
#   state: "idle" | "checking" | "up_to_date" | "updating" | "error"
_update_status = {"state": "idle", "version": VERSION, "latest": None, "detail": ""}
_update_lock = threading.Lock()


def _set_update_status(**kw):
    with _update_lock:
        _update_status.update(kw)


def _parse_ver(s: str):
    """'v1.2.3' / '1.2.3' -> (1,2,3). Non-numeric parts -> 0. Pads to 3."""
    s = (s or "").strip().lstrip("vV")
    parts = re.split(r"[.\-+]", s)
    out = []
    for p in parts[:3]:
        try:
            out.append(int(p))
        except ValueError:
            out.append(0)
    while len(out) < 3:
        out.append(0)
    return tuple(out)


def _is_newer(latest: str, current: str) -> bool:
    return _parse_ver(latest) > _parse_ver(current)


def _gh_latest_release():
    """Return the latest non-draft, non-prerelease release dict, or None."""
    import urllib.request
    url = f"https://api.github.com/repos/{GITHUB_REPO}/releases/latest"
    req = urllib.request.Request(url, headers={
        "Accept": "application/vnd.github+json",
        "User-Agent": f"DrawManager/{VERSION}",
    })
    with urllib.request.urlopen(req, timeout=15) as resp:
        return json.loads(resp.read().decode("utf-8"))


def _download(url: str, dest: Path):
    import urllib.request
    req = urllib.request.Request(url, headers={
        "Accept": "application/octet-stream",
        "User-Agent": f"DrawManager/{VERSION}",
    })
    with urllib.request.urlopen(req, timeout=120) as resp, open(dest, "wb") as f:
        while True:
            chunk = resp.read(65536)
            if not chunk:
                break
            f.write(chunk)


def _apply_update_and_relaunch(new_main: Path, new_worker: Path | None):
    """Windows locks running exes, so we can't overwrite ourselves directly.
    Strategy: write a .bat that waits for THIS process to exit, copies the
    freshly-downloaded exes over the live ones, relaunches the main exe, then
    deletes itself. We hand off to it and quit."""
    if os.name != "nt":
        return  # only meaningful for the frozen Windows build
    cur_main   = APP_DIR / _ASSET_MAIN
    cur_worker = APP_DIR / _ASSET_WORKER
    pid = os.getpid()
    bat = APP_DIR / "_dm_update.bat"

    lines = [
        "@echo off",
        "setlocal enableextensions",
        "rem --- wait for the running Draw Manager (pid %1) to exit ---",
        ":waitloop",
        'tasklist /FI "PID eq %~1" 2>NUL | find /I "%~1" >NUL',
        "if not errorlevel 1 (",
        "  timeout /t 1 /nobreak >NUL",
        "  goto waitloop",
        ")",
        "rem --- swap in the new binaries ---",
        f'move /Y "{new_main}" "{cur_main}" >NUL',
    ]
    if new_worker is not None:
        lines.append(f'move /Y "{new_worker}" "{cur_worker}" >NUL')
    lines += [
        "rem --- relaunch ---",
        f'start "" "{cur_main}"',
        "rem --- self-delete ---",
        '(goto) 2>nul & del "%~f0"',
    ]
    bat.write_text("\r\n".join(lines), encoding="ascii")

    # Detached so it survives our exit. CREATE_NEW_PROCESS_GROUP|DETACHED_PROCESS.
    subprocess.Popen(
        ["cmd", "/c", str(bat), str(pid)],
        creationflags=0x00000200 | 0x00000008,
        close_fds=True,
    )
    # Give the .bat a beat to start waiting, then quit so it can swap files.
    time.sleep(0.5)
    os._exit(0)


def _check_for_update(blocking_apply: bool = True):
    """Background entrypoint. Hits the Releases API; if a newer version is
    published, downloads both assets to a temp dir and triggers the swap."""
    # Only the frozen build can replace .exe files meaningfully.
    if not getattr(sys, "frozen", False):
        _set_update_status(state="up_to_date", detail="dev build (no self-update)")
        return
    if "CHANGE_ME" in GITHUB_REPO:
        _set_update_status(state="error", detail="GITHUB_REPO not configured")
        return
    try:
        _set_update_status(state="checking")
        rel = _gh_latest_release()
        latest = rel.get("tag_name") or rel.get("name") or ""
        _set_update_status(latest=latest.lstrip("vV"))
        if not _is_newer(latest, VERSION):
            _set_update_status(state="up_to_date")
            return

        assets = {a["name"]: a for a in rel.get("assets", [])}
        a_main = assets.get(_ASSET_MAIN)
        if not a_main:
            _set_update_status(state="error", detail=f"release missing {_ASSET_MAIN}")
            return

        _set_update_status(state="updating", detail=f"downloading {latest}")
        import tempfile
        tmp = Path(tempfile.mkdtemp(prefix="dm_update_"))
        new_main = tmp / _ASSET_MAIN
        _download(a_main["browser_download_url"], new_main)

        new_worker = None
        a_worker = assets.get(_ASSET_WORKER)
        if a_worker:
            new_worker = tmp / _ASSET_WORKER
            _download(a_worker["browser_download_url"], new_worker)

        if blocking_apply:
            _set_update_status(state="updating", detail="relaunching")
            _apply_update_and_relaunch(new_main, new_worker)
    except Exception as e:
        _set_update_status(state="error", detail=repr(e))


def _start_update_check():
    threading.Thread(target=_check_for_update, daemon=True).start()

# ---------------------------------------------------------------------------
# Win32 plumbing (no-op on non-Windows). Set up once, reused everywhere.
# ---------------------------------------------------------------------------
if os.name == "nt":
    import ctypes
    from ctypes import wintypes
    _u32, _k32 = ctypes.windll.user32, ctypes.windll.kernel32
    _MY_PID    = _k32.GetCurrentProcessId()
    _EnumProc  = ctypes.WINFUNCTYPE(ctypes.c_bool, wintypes.HWND, wintypes.LPARAM)
else:
    ctypes = wintypes = _u32 = _k32 = _MY_PID = _EnumProc = None  # type: ignore

# ---------------------------------------------------------------------------
# State machine
# ---------------------------------------------------------------------------
STATE_ORDER = [
    "requested", "inspection_ordered", "appraisal_paid", "appraisal_received",
    "email_sent", "approved", "fci_updated", "funded",
]
STATE_LABEL = {
    "pending":            "Pending",
    "requested":          "Requested",
    "inspection_ordered": "Inspection Ordered",
    "appraisal_paid":     "Appraisal Paid",
    "appraisal_received": "Appraisal Received",
    "email_sent":         "Email Sent",
    "approved":           "Approved",
    "fci_updated":        "FCI Updated",
    "funded":             "Funded",
}
STALE_DAYS = {
    "requested": 3, "inspection_ordered": 7, "appraisal_paid": 5,
    "appraisal_received": 2, "email_sent": 3, "approved": 2, "fci_updated": 3,
}

# Per-draw manual progress flow — user drives this from the app.
# Independent of the sheet-derived state (which reflects sheet columns).
PROGRESS_ORDER = [
    "not_started", "order_inspection", "inspection_received",
    "inspection_approved", "pending_wire_instructions", "wire_sent",
]
PROGRESS_LABEL = {
    "not_started":               "Not Started",
    "order_inspection":          "Order Inspection",
    "inspection_received":       "Inspection Received",
    "inspection_approved":       "Inspection Approved",
    "pending_wire_instructions": "Pending Wire Instructions",
    "wire_sent":                 "Wire Sent",
}
def _default_progress(sheet_state):
    """Sensible default progress_state given the sheet-derived current_state."""
    return {
        "pending":            "not_started",
        "requested":          "not_started",
        "inspection_ordered": "order_inspection",
        "appraisal_paid":     "inspection_received",
        "appraisal_received": "inspection_received",
        "email_sent":         "inspection_approved",
        "approved":           "inspection_approved",
        "fci_updated":        "pending_wire_instructions",
        "funded":             "wire_sent",
    }.get(sheet_state, "not_started")

_tasks: dict[str, dict] = {}
_tasks_lock = threading.Lock()
_db_lock    = threading.Lock()
_sync_state = {"counter": 0, "last_error": None}

# Parallel batch runs (each entry tracks many concurrent address sub-runs).
_batches: dict[str, dict] = {}
_batches_lock = threading.Lock()

# Bounded ring buffer of recent errors — surfaced to the debug HUD.
_ERROR_BUF: list[dict] = []
_ERROR_BUF_MAX = 100
_err_lock = threading.Lock()

def _record_error(where: str, exc: BaseException):
    import traceback
    entry = {
        "at": datetime.now(timezone.utc).isoformat(),
        "where": where,
        "type": type(exc).__name__,
        "msg":  str(exc),
        "tb":   traceback.format_exc()[-1200:],
    }
    with _err_lock:
        _ERROR_BUF.append(entry)
        if len(_ERROR_BUF) > _ERROR_BUF_MAX:
            del _ERROR_BUF[: len(_ERROR_BUF) - _ERROR_BUF_MAX]
    return entry


# ---------------------------------------------------------------------------
# Settings (with migration)
# ---------------------------------------------------------------------------
def _as_bool(v, default=False):
    """Coerce JSON bools and UI strings ('true'/'false') to a real bool.
    Plain bool(\"false\") is True, so selects that send strings need this."""
    if isinstance(v, bool):
        return v
    if v is None:
        return default
    return str(v).strip().lower() in ("true", "1", "yes", "on", "live")

def _migrate(s: dict) -> dict:
    """Build current-shape settings from any prior version."""
    s = s or {}
    conn = s.get("connections", {})
    gs   = conn.get("google_sheets", {})
    bl   = conn.get("baseline", {})
    scrs = s.get("scripts", {})
    dr   = scrs.get("draw-request", {})

    return {
        "scriptDir": s.get("scriptDir", str(APP_DIR)),
        "theme":     s.get("theme", "blueprint"),    # 'blueprint' | 'light'
        "connections": {
            "baseline": {
                "api_key": bl.get("api_key", s.get("baselineKey", "")),
            },
            "google_sheets": {
                "credentials_path": gs.get("credentials_path", ""),
                "spreadsheet_id":   gs.get("spreadsheet_id", ""),
                "sheet_name":       gs.get("sheet_name", "Draw"),
                "range":            gs.get("range", "A:DZ"),
                "auto_sync":        bool(gs.get("auto_sync", False)),
                "interval_minutes": int(gs.get("interval_minutes", 15)),
                "last_sync_at":     gs.get("last_sync_at"),
                # Write-back: when enabled, the app can update the source sheet
                # (requires service-account to have Editor permission on the sheet).
                "write_enabled":    bool(gs.get("write_enabled", False)),
                "auto_write":       bool(gs.get("auto_write", False)),
            },
        },
        "scripts": {
            "draw-request": {
                "inspection_type": dr.get("inspection_type",
                                          s.get("inspType",
                                                "draw inspection without floorplan")),
                "progress_step":   dr.get("progress_step", s.get("step", "25")),
                "contact_choice":  dr.get("contact_choice", s.get("contactChoice", "auto")),
                "draw_items_csv":  dr.get("draw_items_csv", ""),
                # Parallel runs: how many browser workers at once.
                "workers":         int(float(dr.get("workers", 7) or 7)),
                # LIVE actually clicks Submit; TEST stops just before it.
                "live_submit":     _as_bool(dr.get("live_submit", False)),
                # Run browsers without a visible window.
                "headless":        _as_bool(dr.get("headless", False)),
            }
        },
    }

def _load_settings():
    raw = {}
    if SETTINGS_FILE.exists():
        try: raw = json.loads(SETTINGS_FILE.read_text())
        except Exception: raw = {}
    return _migrate(raw)

def _save_settings(d):
    """Deep-merge into existing settings, then write."""
    cur = _load_settings()
    def merge(a, b):
        for k, v in (b or {}).items():
            if isinstance(v, dict) and isinstance(a.get(k), dict):
                merge(a[k], v)
            else:
                a[k] = v
    merge(cur, d or {})
    SETTINGS_FILE.write_text(json.dumps(cur, indent=2))


# ---------------------------------------------------------------------------
# DB
# ---------------------------------------------------------------------------
def _now():
    return datetime.now(timezone.utc).isoformat(timespec="seconds")

def _db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    return conn

def _init_db():
    with _db() as c:
        c.executescript("""
        CREATE TABLE IF NOT EXISTS loans (
          id INTEGER PRIMARY KEY AUTOINCREMENT,
          cf_number TEXT UNIQUE NOT NULL,
          borrower_name TEXT, address TEXT, fci_loan_number TEXT,
          budget_amount REAL, budget_left REAL,
          status TEXT DEFAULT 'active',
          imported_at TEXT, updated_at TEXT
        );
        CREATE TABLE IF NOT EXISTS draws (
          id INTEGER PRIMARY KEY AUTOINCREMENT,
          loan_id INTEGER NOT NULL REFERENCES loans(id) ON DELETE CASCADE,
          draw_number INTEGER NOT NULL,
          requested_amount REAL,
          inspection_ordered INTEGER DEFAULT 0,
          appraisal_paid INTEGER DEFAULT 0,
          appraisal_received INTEGER DEFAULT 0,
          draw_email_sent INTEGER DEFAULT 0,
          approved_amount REAL,
          fci_updated INTEGER DEFAULT 0,
          funded_date TEXT,
          current_state TEXT, state_entered_at TEXT,
          imported_at TEXT, updated_at TEXT,
          UNIQUE(loan_id, draw_number)
        );
        CREATE TABLE IF NOT EXISTS state_transitions (
          id INTEGER PRIMARY KEY AUTOINCREMENT,
          draw_id INTEGER NOT NULL REFERENCES draws(id) ON DELETE CASCADE,
          from_state TEXT, to_state TEXT, at TEXT, source TEXT
        );
        CREATE TABLE IF NOT EXISTS sheet_data (
          id INTEGER PRIMARY KEY AUTOINCREMENT,
          sheet_name TEXT NOT NULL,
          row_index INTEGER NOT NULL,
          cf_number TEXT,
          data_json TEXT NOT NULL,
          imported_at TEXT,
          UNIQUE(sheet_name, row_index)
        );
        CREATE INDEX IF NOT EXISTS idx_draws_state ON draws(current_state);
        CREATE INDEX IF NOT EXISTS idx_draws_loan  ON draws(loan_id);
        CREATE INDEX IF NOT EXISTS idx_sheet_data_cf    ON sheet_data(cf_number);
        CREATE INDEX IF NOT EXISTS idx_sheet_data_sheet ON sheet_data(sheet_name);
        """)
        # Migration: add progress_state column if missing, backfill from current_state
        cols = {r[1] for r in c.execute("PRAGMA table_info(draws)")}
        if "progress_state" not in cols:
            c.execute("ALTER TABLE draws ADD COLUMN progress_state TEXT")
            for state in set(STATE_LABEL):
                c.execute("UPDATE draws SET progress_state=? WHERE current_state=?",
                          (_default_progress(state), state))
            c.execute("UPDATE draws SET progress_state='order_inspection' WHERE progress_state IS NULL")
        # Migration: total_draws column on loans (how many draws are planned for this loan)
        lcols = {r[1] for r in c.execute("PRAGMA table_info(loans)")}
        if "total_draws" not in lcols:
            c.execute("ALTER TABLE loans ADD COLUMN total_draws INTEGER")
            # Default to current draw count per loan
            c.execute("""UPDATE loans SET total_draws = (
                            SELECT COUNT(*) FROM draws WHERE draws.loan_id = loans.id
                         ) WHERE total_draws IS NULL""")
        # Migration: sheet_row_index — 1-based row number in the primary sheet
        # where this loan's CF# was found. Used for Google Sheets write-back.
        if "sheet_row_index" not in lcols:
            c.execute("ALTER TABLE loans ADD COLUMN sheet_row_index INTEGER")
        c.commit()


# ---------------------------------------------------------------------------
# Sheet parser (unchanged)
# ---------------------------------------------------------------------------
_CF_RE    = re.compile(r"^CF\s*\d+$", re.IGNORECASE)
_MONEY_RE = re.compile(r"[^\d.\-]")

def _money(s):
    if s is None: return None
    s = str(s).strip()
    if not s or s in ("-", "—"): return None
    cleaned = _MONEY_RE.sub("", s)
    if not cleaned or cleaned in (".", "-", "-."): return None
    try:
        v = float(cleaned)
        return v if v != 0 else None
    except ValueError:
        return None

def _bool(s):
    if not s: return False
    return str(s).strip().upper() in ("TRUE", "YES", "Y", "1", "X", "✓")

def _state_of(d):
    if d["funded_date"]:        return "funded"
    if d["fci_updated"]:        return "fci_updated"
    if d["approved_amount"]:    return "approved"
    if d["draw_email_sent"]:    return "email_sent"
    if d["appraisal_received"]: return "appraisal_received"
    if d["appraisal_paid"]:     return "appraisal_paid"
    if d["inspection_ordered"]: return "inspection_ordered"
    if d["requested_amount"]:   return "requested"
    return "pending"

def parse_sheet(text):
    if not text or not text.strip():
        return []
    first  = [ln for ln in text.splitlines() if ln.strip()][:5]
    tabs   = sum(ln.count("\t") for ln in first)
    commas = sum(ln.count(",")  for ln in first)
    delim  = "\t" if tabs >= commas else ","

    reader = csv.reader(io.StringIO(text), delimiter=delim)
    loans = []
    # Track 1-based row index so we can write back to the right sheet row.
    for row_idx_zero, row in enumerate(reader):
        row_idx = row_idx_zero + 1
        if len(row) < 6: continue
        cf = (row[0] or "").strip()
        if not _CF_RE.match(cf): continue
        loan = {
            "cf_number":       cf.upper().replace(" ", ""),
            "borrower_name":   (row[1] or "").strip(),
            "address":         (row[2] or "").strip(),
            "fci_loan_number": (row[3] or "").strip(),
            "budget_amount":   _money(row[4]),
            "budget_left":     _money(row[5]),
            "sheet_row_index": row_idx,
            "draws":           [],
        }
        col = 6; draw_num = 1
        while col < len(row):
            chunk = row[col:col + 8]
            while len(chunk) < 8: chunk.append("")
            d = {
                "draw_number":         draw_num,
                "requested_amount":    _money(chunk[0]),
                "inspection_ordered":  _bool(chunk[1]),
                "appraisal_paid":      _bool(chunk[2]),
                "appraisal_received":  _bool(chunk[3]),
                "draw_email_sent":     _bool(chunk[4]),
                "approved_amount":     _money(chunk[5]),
                "fci_updated":         _bool(chunk[6]),
                "funded_date":         ((chunk[7] or "").strip() or None),
            }
            d["current_state"] = _state_of(d)
            if d["current_state"] != "pending":
                loan["draws"].append(d)
            col += 8; draw_num += 1
        loans.append(loan)
    return loans


def import_loans(parsed):
    now = _now()
    stats = {"loans_inserted": 0, "loans_updated": 0,
             "draws_inserted": 0, "draws_updated": 0, "state_changes": 0}
    with _db_lock, _db() as c:
        for loan in parsed:
            row = c.execute("SELECT id FROM loans WHERE cf_number=?",
                            (loan["cf_number"],)).fetchone()
            if row:
                loan_id = row["id"]
                c.execute("""UPDATE loans
                             SET borrower_name=?, address=?, fci_loan_number=?,
                                 budget_amount=?, budget_left=?, sheet_row_index=?,
                                 updated_at=?
                             WHERE id=?""",
                          (loan["borrower_name"], loan["address"], loan["fci_loan_number"],
                           loan["budget_amount"], loan["budget_left"],
                           loan.get("sheet_row_index"), now, loan_id))
                stats["loans_updated"] += 1
            else:
                cur = c.execute("""INSERT INTO loans
                             (cf_number, borrower_name, address, fci_loan_number,
                              budget_amount, budget_left, sheet_row_index,
                              imported_at, updated_at)
                             VALUES (?,?,?,?,?,?,?,?,?)""",
                          (loan["cf_number"], loan["borrower_name"], loan["address"],
                           loan["fci_loan_number"], loan["budget_amount"],
                           loan["budget_left"], loan.get("sheet_row_index"), now, now))
                loan_id = cur.lastrowid
                stats["loans_inserted"] += 1

            for d in loan["draws"]:
                existing = c.execute("""SELECT id, current_state FROM draws
                                        WHERE loan_id=? AND draw_number=?""",
                                     (loan_id, d["draw_number"])).fetchone()
                state_entered_at = None
                if existing:
                    old = existing["current_state"]
                    if old != d["current_state"]:
                        state_entered_at = now
                        c.execute("""INSERT INTO state_transitions
                                     (draw_id, from_state, to_state, at, source)
                                     VALUES (?,?,?,?, 'sync')""",
                                  (existing["id"], old, d["current_state"], now))
                        stats["state_changes"] += 1
                    c.execute("""UPDATE draws SET
                                   requested_amount=?, inspection_ordered=?,
                                   appraisal_paid=?, appraisal_received=?,
                                   draw_email_sent=?, approved_amount=?,
                                   fci_updated=?, funded_date=?,
                                   current_state=?,
                                   state_entered_at=COALESCE(?, state_entered_at),
                                   updated_at=?
                                 WHERE id=?""",
                              (d["requested_amount"], int(d["inspection_ordered"]),
                               int(d["appraisal_paid"]), int(d["appraisal_received"]),
                               int(d["draw_email_sent"]), d["approved_amount"],
                               int(d["fci_updated"]), d["funded_date"], d["current_state"],
                               state_entered_at, now, existing["id"]))
                    stats["draws_updated"] += 1
                else:
                    c.execute("""INSERT INTO draws
                                 (loan_id, draw_number, requested_amount, inspection_ordered,
                                  appraisal_paid, appraisal_received, draw_email_sent,
                                  approved_amount, fci_updated, funded_date, current_state,
                                  state_entered_at, progress_state, imported_at, updated_at)
                                 VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                              (loan_id, d["draw_number"], d["requested_amount"],
                               int(d["inspection_ordered"]), int(d["appraisal_paid"]),
                               int(d["appraisal_received"]), int(d["draw_email_sent"]),
                               d["approved_amount"], int(d["fci_updated"]),
                               d["funded_date"], d["current_state"], now,
                               _default_progress(d["current_state"]), now, now))
                    stats["draws_inserted"] += 1
        c.commit()
    return stats


# ---------------------------------------------------------------------------
# Draw-log parser  (the layout people actually paste/upload to test)
# Columns, in order:
#   Draw # | Name | Address | AA | BL | Last 4 | Draw Fee |
#   Wire Received | OUT | Going Out | Released On
# There's no CF# in this layout, so we synthesize a stable one from the
# address (same address -> same CF#, so re-imports update instead of dupe).
# ---------------------------------------------------------------------------
import hashlib  # noqa: E402  (kept local to this section for clarity)

def _synth_cf(address: str, name: str) -> str:
    base = re.sub(r"\s+", " ", (address or name or "").strip().lower())
    h = int(hashlib.sha1(base.encode("utf-8")).hexdigest(), 16) % 1_000_000
    return f"CF{h:06d}"

def _int_or(s, default=None):
    try:
        return int(float(re.sub(r"[^\d.\-]", "", str(s))))
    except (ValueError, TypeError):
        return default

def parse_draw_log(text: str) -> list:
    """Parse the Draw #/Name/Address/AA/.../Released On layout into the same
    loan/draw structure import_loans() consumes. Rows without a usable address
    (blank lines, the header, a TOTAL row) are skipped."""
    if not text or not text.strip():
        return []
    lines = [ln for ln in text.splitlines() if ln.strip()]
    sample = lines[:5]
    delim = "\t" if sum(l.count("\t") for l in sample) >= sum(l.count(",") for l in sample) else ","

    loans_by_addr: dict[str, dict] = {}
    order: list[str] = []
    for row in csv.reader(io.StringIO(text), delimiter=delim):
        if len(row) < 3:
            continue
        cell = lambda i: (row[i].strip() if i < len(row) and row[i] is not None else "")
        name, address = cell(1), cell(2)
        # Skip header and non-data rows.
        if not address or address.lower() in ("address", "total"):
            continue
        if name.lower() == "name":
            continue

        draw_no   = _int_or(cell(0), None)
        aa        = _money(cell(3))                # AA -> requested amount
        wire_in   = _bool(cell(7))                 # Wire Received ("IN")
        went_out  = _bool(cell(8))                 # OUT checkmark
        going_out = _money(cell(9))                # Going Out -> approved amount
        released  = cell(10) or None               # Released On -> funded date
        last4     = cell(5)

        key = re.sub(r"\s+", " ", address.strip().lower())
        loan = loans_by_addr.get(key)
        if loan is None:
            loan = {
                "cf_number":       _synth_cf(address, name),
                "borrower_name":   name,
                "address":         address,
                "fci_loan_number": (f"x{last4}" if last4 and last4.isdigit() else ""),
                "budget_amount":   None,
                "budget_left":     None,
                "sheet_row_index": None,
                "draws":           [],
            }
            loans_by_addr[key] = loan
            order.append(key)

        d = {
            "draw_number":        draw_no if draw_no is not None else (len(loan["draws"]) + 1),
            "requested_amount":   aa,
            "inspection_ordered": went_out,        # it left for inspection
            "appraisal_paid":     False,
            "appraisal_received": False,
            "draw_email_sent":    False,
            "approved_amount":    going_out,
            "fci_updated":        False,
            "funded_date":        (released if (wire_in or went_out) else None),
        }
        d["current_state"] = _state_of(d)
        if d["current_state"] != "pending":
            loan["draws"].append(d)

    return [loans_by_addr[k] for k in order if loans_by_addr[k]["draws"]]


# ---------------------------------------------------------------------------
# Google Sheets fetch
# ---------------------------------------------------------------------------
def _is_local_source(path: str) -> bool:
    """True if the configured source path points at a local Excel file."""
    return bool(path) and path.lower().endswith((".xlsx", ".xlsm"))


def _read_local_xlsx(path, sheet_name):
    """Open a local .xlsx and return a TSV string parse_sheet understands.
    Falls back to the first sheet if sheet_name isn't provided or isn't found."""
    from openpyxl import load_workbook  # lazy: only loaded when used
    wb = load_workbook(path, data_only=True, read_only=True)
    try:
        ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active
        return "\n".join(
            "\t".join("" if v is None else str(v) for v in row)
            for row in ws.iter_rows(values_only=True)
        )
    finally:
        wb.close()


def _read_all_xlsx_sheets(path):
    """Return {sheet_name: tsv_string} for every non-empty tab in the workbook."""
    from openpyxl import load_workbook
    wb = load_workbook(path, data_only=True, read_only=True)
    try:
        out = {}
        for name in wb.sheetnames:
            ws = wb[name]
            tsv = "\n".join(
                "\t".join("" if v is None else str(v) for v in row)
                for row in ws.iter_rows(values_only=True)
            ).strip()
            if tsv:
                out[name] = tsv
        return out
    finally:
        wb.close()


def _read_file_as_tsv(path: str) -> str:
    """Read a user-supplied loan file (.xlsx/.xlsm or .csv/.tsv/.txt) and
    return text that parse_sheet() understands. For Excel, tries the active
    sheet first, then falls back to whichever tab yields the most CF# rows."""
    low = path.lower()
    if low.endswith((".xlsx", ".xlsm")):
        primary = _read_local_xlsx(path, None)
        if parse_sheet(primary):
            return primary
        best, best_n = primary, 0
        for _name, tsv in _read_all_xlsx_sheets(path).items():
            n = len(parse_sheet(tsv))
            if n > best_n:
                best, best_n = tsv, n
        return best
    # Plain text — parse_sheet sniffs tab vs comma itself. utf-8-sig strips BOM.
    return Path(path).read_text(encoding="utf-8-sig", errors="replace")


def _fetch_all_google_sheets(creds_path, spreadsheet_id):
    """Return {sheet_name: tsv_string} for every tab. Single batchGet round trip."""
    from google.oauth2 import service_account              # type: ignore[import-not-found]
    from googleapiclient.discovery import build            # type: ignore[import-not-found]
    creds = service_account.Credentials.from_service_account_file(
        creds_path,
        scopes=["https://www.googleapis.com/auth/spreadsheets.readonly"],
    )
    service = build("sheets", "v4", credentials=creds, cache_discovery=False)
    meta = service.spreadsheets().get(
        spreadsheetId=spreadsheet_id,
        fields="sheets.properties.title"
    ).execute()
    tabs = [s["properties"]["title"] for s in meta.get("sheets", [])]
    if not tabs:
        return {}
    result = service.spreadsheets().values().batchGet(
        spreadsheetId=spreadsheet_id, ranges=tabs
    ).execute()
    out = {}
    for tab, vr in zip(tabs, result.get("valueRanges", [])):
        rows = vr.get("values", [])
        if not rows:
            continue
        max_w = max(len(r) for r in rows)
        padded = [r + [""] * (max_w - len(r)) for r in rows]
        tsv = "\n".join("\t".join(str(c) for c in row) for row in padded).strip()
        if tsv:
            out[tab] = tsv
    return out


_CF_ANY_RE = re.compile(r"^CF[A-Z]?\d+$", re.IGNORECASE)


# ---------------------------------------------------------------------------
# Google Sheets WRITE-BACK
# ---------------------------------------------------------------------------
# The primary sheet layout is deterministic:
#   col A (1) = CF#
#   col B (2) = borrower name
#   col C (3) = address
#   col D (4) = fci loan #
#   col E (5) = budget amount
#   col F (6) = budget left
#   col G (7) onward = draws, 8 cols per draw
# Draw N starts at column (7 + (N-1)*8).
# Within each 8-col chunk, the offsets are:
#   0: requested_amount      (number)
#   1: inspection_ordered    (bool / "TRUE")
#   2: appraisal_paid        (bool)
#   3: appraisal_received    (bool)
#   4: draw_email_sent       (bool)
#   5: approved_amount       (number)
#   6: fci_updated           (bool)
#   7: funded_date           (date string)
DRAW_FIELD_OFFSETS = {
    "requested_amount":    0,
    "inspection_ordered":  1,
    "appraisal_paid":      2,
    "appraisal_received":  3,
    "draw_email_sent":     4,
    "approved_amount":     5,
    "fci_updated":         6,
    "funded_date":         7,
}

# When the user advances a draw's progress in the app, these are the sheet
# fields to flip TRUE (or fill, for funded_date). Forward-only: we never
# un-check a field the user may have set manually.
PROGRESS_TO_SHEET_FIELDS = {
    "not_started":               [],
    "order_inspection":          [],   # the script itself flips inspection_ordered
    "inspection_received":       [("inspection_ordered", True),
                                   ("appraisal_paid",     True),
                                   ("appraisal_received", True)],
    "inspection_approved":       [("draw_email_sent",    True)],
    "pending_wire_instructions": [("fci_updated",        True)],
    "wire_sent":                 [("funded_date",        "__today__")],
}


def _col_letter(col_one_based: int) -> str:
    """1 → A, 26 → Z, 27 → AA, etc. Used to build A1-style cell refs."""
    s = ""
    n = col_one_based
    while n > 0:
        n, r = divmod(n - 1, 26)
        s = chr(65 + r) + s
    return s


def _gsheet_service(creds_path, write=False):
    """Build an authenticated Sheets service. Use write=True to request the
    read/write scope (requires service account to have Editor on the sheet)."""
    from google.oauth2 import service_account              # type: ignore[import-not-found]
    from googleapiclient.discovery import build            # type: ignore[import-not-found]
    scope = ("https://www.googleapis.com/auth/spreadsheets"
             if write else "https://www.googleapis.com/auth/spreadsheets.readonly")
    creds = service_account.Credentials.from_service_account_file(
        creds_path, scopes=[scope],
    )
    return build("sheets", "v4", credentials=creds, cache_discovery=False)


def gsheet_write_cell(creds_path, spreadsheet_id, sheet_name, row_one_based,
                       col_one_based, value):
    """Write a single value into a single cell.
    `value` may be str / bool / number / None. Booleans become TRUE/FALSE
    strings so they render as checkboxes correctly."""
    if isinstance(value, bool):
        value_str = "TRUE" if value else "FALSE"
    elif value is None:
        value_str = ""
    else:
        value_str = str(value)
    cell = f"{_col_letter(col_one_based)}{row_one_based}"
    a1 = f"'{sheet_name}'!{cell}"
    service = _gsheet_service(creds_path, write=True)
    return service.spreadsheets().values().update(
        spreadsheetId=spreadsheet_id,
        range=a1,
        valueInputOption="USER_ENTERED",
        body={"values": [[value_str]]},
    ).execute()


def gsheet_batch_write(creds_path, spreadsheet_id, sheet_name, updates):
    """One-shot batch write. `updates` is a list of
    {row, col, value} dicts (1-based row/col). Single API round trip."""
    if not updates:
        return {"updatedCells": 0}
    data = []
    for u in updates:
        v = u["value"]
        if isinstance(v, bool):
            v = "TRUE" if v else "FALSE"
        elif v is None:
            v = ""
        else:
            v = str(v)
        cell = f"{_col_letter(int(u['col']))}{int(u['row'])}"
        data.append({
            "range":  f"'{sheet_name}'!{cell}",
            "values": [[v]],
        })
    service = _gsheet_service(creds_path, write=True)
    return service.spreadsheets().values().batchUpdate(
        spreadsheetId=spreadsheet_id,
        body={"valueInputOption": "USER_ENTERED", "data": data},
    ).execute()


def _push_progress_to_sheet(loan_row, draw_number, new_progress_state):
    """Best-effort write of one draw's progress to the primary sheet.
    Returns a result dict; never raises (errors come back as {error: ...})."""
    s  = _load_settings()
    gs = s["connections"]["google_sheets"]
    if not gs.get("write_enabled"):
        return {"skipped": "write_enabled is off"}
    path = gs.get("credentials_path", "")
    if _is_local_source(path):
        return {"skipped": "source is a local file (only Google Sheets supports write-back)"}
    if not (path and os.path.exists(path) and gs.get("spreadsheet_id")):
        return {"skipped": "Google Sheets not configured"}

    row_idx = loan_row["sheet_row_index"] if loan_row else None
    if not row_idx:
        return {"skipped": "loan has no sheet_row_index — run Sync once to populate"}

    fields = PROGRESS_TO_SHEET_FIELDS.get(new_progress_state, [])
    if not fields:
        return {"skipped": "no sheet fields mapped for this progress state"}

    today = datetime.now().strftime("%m/%d/%Y")
    draw_base_col = 7 + (int(draw_number) - 1) * 8   # 1-based, col G for draw 1
    updates = []
    for field_name, value in fields:
        if value == "__today__":
            value = today
        offset = DRAW_FIELD_OFFSETS[field_name]
        updates.append({
            "row":   row_idx,
            "col":   draw_base_col + offset,
            "value": value,
        })

    try:
        result = gsheet_batch_write(
            path, gs["spreadsheet_id"], gs.get("sheet_name", "Draw"), updates,
        )
        return {"ok": True, "updates": updates,
                "updated_cells": result.get("totalUpdatedCells")}
    except Exception as e:
        msg = f"{type(e).__name__}: {e}"
        return {"error": msg}


def parse_extra_sheet(tsv):
    """Generic parser for any sheet. Detects the most plausible header row
    (the row with the most short non-empty cells in the first 12), then
    yields (row_index, cf_number_or_None, {col_header: value}) for each
    data row below. CF# is extracted from any column that matches CF\\d+
    or CFD\\d+ style. Empty rows are skipped."""
    if not tsv or not tsv.strip():
        return []
    rows = [line.split("\t") for line in tsv.split("\n")]
    # Header detection — first 12 rows, pick the one with the most non-empty
    # short cells (real headers are short labels, not paragraphs).
    best_idx, best_score = 0, -1
    for i, row in enumerate(rows[:12]):
        score = sum(1 for c in row if c.strip() and len(c.strip()) < 60)
        if score > best_score:
            best_score, best_idx = score, i
    headers = [c.strip() or f"col_{j}" for j, c in enumerate(rows[best_idx])]
    out = []
    for i, row in enumerate(rows[best_idx + 1:], start=best_idx + 1):
        if not any((c or "").strip() for c in row):
            continue
        data = {}
        cf = None
        for j, cell in enumerate(row):
            v = (cell or "").strip()
            if not v:
                continue
            data[headers[j] if j < len(headers) else f"col_{j}"] = v
            if cf is None:
                norm = v.upper().replace(" ", "")
                if _CF_ANY_RE.match(norm):
                    cf = norm
        if data:
            out.append((i, cf, data))
    return out


def import_sheet_data(sheet_name, parsed_rows):
    """Replace this sheet's rows in sheet_data. Returns number of rows stored."""
    now = _now()
    with _db_lock, _db() as c:
        c.execute("DELETE FROM sheet_data WHERE sheet_name=?", (sheet_name,))
        for row_idx, cf, data in parsed_rows:
            c.execute("""INSERT INTO sheet_data (sheet_name, row_index, cf_number, data_json, imported_at)
                         VALUES (?, ?, ?, ?, ?)""",
                      (sheet_name, row_idx, cf, json.dumps(data, default=str), now))
        c.commit()
    return len(parsed_rows)


def fetch_sheet_values(creds_path, spreadsheet_id, range_name):
    from google.oauth2 import service_account              # type: ignore[import-not-found]
    from googleapiclient.discovery import build            # type: ignore[import-not-found]

    creds = service_account.Credentials.from_service_account_file(
        creds_path,
        scopes=["https://www.googleapis.com/auth/spreadsheets.readonly"],
    )
    service = build("sheets", "v4", credentials=creds, cache_discovery=False)
    result  = service.spreadsheets().values().get(
        spreadsheetId=spreadsheet_id, range=range_name
    ).execute()
    rows = result.get("values", [])
    if not rows:
        return ""
    max_w = max(len(r) for r in rows)
    padded = [r + [""] * (max_w - len(r)) for r in rows]
    return "\n".join("\t".join(str(c) for c in row) for row in padded)


def test_gsheet_connection(creds_path, spreadsheet_id):
    if _is_local_source(creds_path):
        from openpyxl import load_workbook
        wb = load_workbook(creds_path, data_only=True, read_only=True)
        try:
            return {"title": Path(creds_path).name, "sheets": wb.sheetnames}
        finally:
            wb.close()
    from google.oauth2 import service_account              # type: ignore[import-not-found]
    from googleapiclient.discovery import build            # type: ignore[import-not-found]

    creds = service_account.Credentials.from_service_account_file(
        creds_path,
        scopes=["https://www.googleapis.com/auth/spreadsheets.readonly"],
    )
    service = build("sheets", "v4", credentials=creds, cache_discovery=False)
    meta = service.spreadsheets().get(
        spreadsheetId=spreadsheet_id,
        fields="properties.title,sheets.properties.title",
    ).execute()
    return {
        "title":  meta["properties"]["title"],
        "sheets": [s["properties"]["title"] for s in meta.get("sheets", [])],
    }


def _days_since(iso_str):
    if not iso_str: return None
    try:
        then = datetime.fromisoformat(iso_str)
        if then.tzinfo is None:
            then = then.replace(tzinfo=timezone.utc)
        return max(0, (datetime.now(timezone.utc) - then).days)
    except Exception:
        return None


def _seconds_since(iso_str):
    if not iso_str: return None
    try:
        then = datetime.fromisoformat(iso_str)
        if then.tzinfo is None:
            then = then.replace(tzinfo=timezone.utc)
        return max(0, int((datetime.now(timezone.utc) - then).total_seconds()))
    except Exception:
        return None


# ---------------------------------------------------------------------------
# Sync manager (background thread)
# ---------------------------------------------------------------------------
class SyncManager:
    def __init__(self):
        self.stop_event = threading.Event()
        self.thread     = None

    def start(self):
        if self.thread and self.thread.is_alive():
            return
        self.stop_event.clear()
        self.thread = threading.Thread(target=self._loop, daemon=True)
        self.thread.start()

    def _loop(self):
        while not self.stop_event.is_set():
            try:
                s  = _load_settings()
                gs = s["connections"]["google_sheets"]
                if (gs.get("auto_sync") and gs.get("credentials_path")
                        and (_is_local_source(gs["credentials_path"])
                             or gs.get("spreadsheet_id"))):
                    last = gs.get("last_sync_at")
                    interval_s = max(60, int(gs.get("interval_minutes", 15)) * 60)
                    age = _seconds_since(last)
                    if age is None or age >= interval_s:
                        do_sheet_sync()
            except Exception as e:
                _sync_state["last_error"] = f"{type(e).__name__}: {e}"
            for _ in range(3):
                if self.stop_event.is_set(): return
                time.sleep(1)


def do_sheet_sync():
    """Perform one sync. Returns stats dict or error dict.
    Reads the entire workbook: the primary sheet drives loans/draws, all other
    sheets are stored as raw rows in sheet_data and surfaced as cross-references
    in the loan drawer when their CF# matches."""
    s  = _load_settings()
    gs = s["connections"]["google_sheets"]
    path = gs.get("credentials_path", "")
    if not path:
        return {"error": "Source path not set (need .xlsx or service-account .json)"}
    if not os.path.exists(path):
        return {"error": f"Source file not found: {path}"}
    local = _is_local_source(path)
    if not local and not gs.get("spreadsheet_id"):
        return {"error": "Spreadsheet ID not set"}

    try:
        if local:
            all_sheets = _read_all_xlsx_sheets(path)
        else:
            try:
                all_sheets = _fetch_all_google_sheets(path, gs["spreadsheet_id"])
            except Exception as e:
                msg = str(e)
                if "Unable to parse range" in msg or "Unable to parse" in msg:
                    try:
                        info = test_gsheet_connection(path, gs["spreadsheet_id"])
                        return {"error": f"Workbook fetch failed: {msg}",
                                "available_sheets": info.get("sheets", [])}
                    except Exception:
                        pass
                raise

        primary = gs.get("sheet_name") or "Draw"
        primary_tsv = all_sheets.get(primary)
        if not primary_tsv:
            return {"error": f"Primary sheet '{primary}' not found in workbook",
                    "available_sheets": list(all_sheets),
                    "configured_sheet": primary}

        # The primary sheet drives the loans/draws tables via the existing parser.
        parsed = parse_sheet(primary_tsv)
        if not parsed:
            _sync_state["last_error"] = "No CF# rows found in sheet range"
            return {"error": f"No CF# rows found in primary sheet '{primary}'"}
        stats = import_loans(parsed)

        # Every OTHER sheet gets stored as raw rows in sheet_data so we can
        # cross-reference them by CF# in the loan drawer.
        extra_stats = {}
        for name, tsv in all_sheets.items():
            if name == primary:
                continue
            rows = parse_extra_sheet(tsv)
            if rows:
                extra_stats[name] = import_sheet_data(name, rows)

        _save_settings({"connections": {"google_sheets": {"last_sync_at": _now()}}})
        _sync_state["counter"]    += 1
        _sync_state["last_error"]  = None
        return {"ok": True, "loans_parsed": len(parsed),
                "extra_sheets": extra_stats, **stats}
    except Exception as e:
        msg = f"{type(e).__name__}: {e}"
        _sync_state["last_error"] = msg
        return {"error": msg}


# ---------------------------------------------------------------------------
# JS API
# ---------------------------------------------------------------------------
def _draw_worker_cmd_head(s: dict):
    """Resolve how to launch the draw-request worker (frozen exe vs script).
    Returns (cmd_head_list, None) or (None, {error})."""
    if getattr(sys, "frozen", False):
        worker = APP_DIR / "draw-request.exe"
        if not worker.exists():
            return None, {"error": f"draw-request.exe not found next to {APP_DIR}"}
        return [str(worker)], None
    script = Path(s["scriptDir"]) / "draw_request.py"
    if not script.exists():
        return None, {"error": f"draw_request.py not found at {script}"}
    return [PYTHON, str(script)], None


def _build_draw_cmd(s, dr, address, profile_dir, live, csv, headless=False):
    """Assemble the worker command line for one address. Each run gets its own
    isolated browser profile (its own window) — the model that actually works.
    Returns (cmd_list, None) or (None, {error})."""
    head, err = _draw_worker_cmd_head(s)
    if err:
        return None, err
    cmd = head + [
        "--address",        address,
        "--contact-choice", dr.get("contact_choice", "auto"),
        "--step",           str(dr.get("progress_step", "25")),
        "--type",           dr.get("inspection_type", ""),
        "--profile-dir",    str(profile_dir),
        "--db",             str(DB_PATH),
    ]
    if csv:
        cmd += ["--csv", csv]
    if live:
        cmd += ["--submit"]      # LIVE: worker will click Submit
    if headless:
        cmd += ["--headless"]
    return cmd, None


class Api:
    def __init__(self):
        self._window: webview.Window | None = None
        self._maximized = False
        self._hwnd: int | None = None    # populated by _apply_windows_icon

    def set_window(self, w):
        self._window = w

    # ----- version / auto-update (consumed by titlebar indicator) -----
    def app_version(self):
        return VERSION

    def update_status(self):
        with _update_lock:
            return dict(_update_status)

    # ----- window controls -----
    def win_minimize(self):
        if self._window: self._window.minimize()

    def win_toggle_max(self):
        if not self._window: return False
        self._window.toggle_fullscreen()
        self._maximized = not self._maximized
        return self._maximized

    def win_close(self):
        if self._window: self._window.destroy()

    def win_start_drag(self):
        """Hand the drag to Win32's modal drag loop.
        WebView2 SetCaptures the mouse on the UI thread; our JS-API call
        runs on a worker thread, so a direct ReleaseCapture here is a
        no-op. AttachThreadInput briefly merges queues so ReleaseCapture
        takes effect on the UI thread, then SendMessage WM_NCLBUTTONDOWN/
        HTCAPTION enters the modal drag loop with the mouse free.
        Returns a diagnostic dict for the debug HUD."""
        if not _u32: return {"ok": False, "reason": "not_windows"}
        hwnd = self._find_our_window() or self._hwnd
        if not hwnd: return {"ok": False, "reason": "no_hwnd"}
        try:
            pt = wintypes.POINT()
            _u32.GetCursorPos(ctypes.byref(pt))
            lparam = ((pt.y & 0xFFFF) << 16) | (pt.x & 0xFFFF)
            ui_tid, my_tid = _u32.GetWindowThreadProcessId(hwnd, None), _k32.GetCurrentThreadId()
            attached = bool(ui_tid and ui_tid != my_tid
                            and _u32.AttachThreadInput(my_tid, ui_tid, True))
            try:
                _u32.ReleaseCapture()
                _u32.SendMessageW(hwnd, 0x00A1, 2, lparam)  # WM_NCLBUTTONDOWN, HTCAPTION
            finally:
                if attached: _u32.AttachThreadInput(my_tid, ui_tid, False)
            self._hwnd = hwnd
            return {"ok": True, "hwnd": int(hwnd), "attached": attached,
                    "ui_tid": int(ui_tid) if ui_tid else None}
        except Exception as e:
            return {"ok": False, "reason": repr(e)}

    def _find_our_window(self):
        """First visible top-level window in our process titled 'Draw Manager'.
        Filtering by PID avoids matching e.g. an Explorer window of the same name."""
        if not _u32: return None
        try:
            found = [None]
            def cb(hwnd, _):
                pid = wintypes.DWORD()
                _u32.GetWindowThreadProcessId(hwnd, ctypes.byref(pid))
                if pid.value != _MY_PID or not _u32.IsWindowVisible(hwnd): return True
                n = _u32.GetWindowTextLengthW(hwnd)
                if n <= 0: return True
                buf = ctypes.create_unicode_buffer(n + 1)
                _u32.GetWindowTextW(hwnd, buf, n + 1)
                if buf.value == "Draw Manager":
                    found[0] = hwnd
                    return False
                return True
            _u32.EnumWindows(_EnumProc(cb), 0)
            return found[0]
        except Exception:
            return None

    def win_debug_info(self):
        """Diagnostic snapshot for the debug HUD. No side effects."""
        h = lambda v: int(v) if v else None
        info = {"os": os.name, "frozen": bool(getattr(sys, "frozen", False)),
                "cached": h(self._hwnd)}
        if _u32:
            try:
                ow = self._find_our_window()
                fg = _u32.GetForegroundWindow()
                info.update({"by_title": h(_u32.FindWindowW(None, "Draw Manager")),
                             "by_proc":  h(ow), "fg": h(fg),
                             "is_fg":    bool(fg and (fg == self._hwnd or fg == ow))})
            except Exception as e:
                info["err"] = repr(e)
        return info

    # ----- file dialog -----
    def pick_file(self, kind="json"):
        if not self._window: return None
        try:
            ft = {
                "json":         ("JSON files (*.json)",),
                "xlsx":         ("Excel files (*.xlsx;*.xlsm)",),
                "csv":          ("CSV files (*.csv)",),
                "sheet_source": ("Sheet source (*.xlsx;*.xlsm;*.json)",
                                 "Excel (*.xlsx;*.xlsm)",
                                 "JSON (*.json)"),
                "loan_import":  ("Loan files (*.xlsx;*.xlsm;*.csv;*.tsv;*.txt)",
                                 "Excel (*.xlsx;*.xlsm)",
                                 "CSV or text (*.csv;*.tsv;*.txt)"),
            }.get(kind, ())
            r = self._window.create_file_dialog(webview.OPEN_DIALOG, file_types=ft)
            if not r: return None
            return r[0] if isinstance(r, (list, tuple)) else r
        except Exception as e:
            return {"error": str(e)}

    # ----- settings -----
    def get_settings(self):
        return _load_settings()

    def save_settings(self, settings):
        _save_settings(settings or {})
        return _load_settings()

    def get_script_settings(self, slug):
        return _load_settings()["scripts"].get(slug, {})

    def save_script_settings(self, slug, values):
        _save_settings({"scripts": {slug: values or {}}})
        return _load_settings()["scripts"].get(slug, {})

    # ----- google sheets -----
    def gsheet_test(self):
        s  = _load_settings()
        gs = s["connections"]["google_sheets"]
        if not gs.get("credentials_path"):
            return {"error": "Source path not set"}
        if not os.path.exists(gs["credentials_path"]):
            return {"error": "Source file not found"}
        if not _is_local_source(gs["credentials_path"]) and not gs.get("spreadsheet_id"):
            return {"error": "Spreadsheet ID not set"}
        try:
            info = test_gsheet_connection(gs["credentials_path"], gs.get("spreadsheet_id", ""))
            configured = gs.get("sheet_name", "")
            return {"ok": True,
                    "configured_sheet": configured,
                    "sheet_name_ok": (not configured) or (configured in info.get("sheets", [])),
                    **info}
        except Exception as e:
            return {"error": f"{type(e).__name__}: {e}"}

    def gsheet_sync_now(self):
        return do_sheet_sync()

    def get_sync_status(self):
        s  = _load_settings()
        gs = s["connections"]["google_sheets"]
        path = gs.get("credentials_path", "")
        return {
            "counter":       _sync_state["counter"],
            "last_error":    _sync_state["last_error"],
            "last_sync_at":  gs.get("last_sync_at"),
            "seconds_since": _seconds_since(gs.get("last_sync_at")),
            "auto_sync":     bool(gs.get("auto_sync")),
            "configured":    bool(path and (_is_local_source(path) or gs.get("spreadsheet_id"))),
        }

    # ----- manual import (fallback) -----
    def import_sheet(self, text):
        try:
            parsed = parse_sheet(text or "")
            if not parsed:
                return {"error": "No loan rows found (column A must contain CF###)."}
            stats = import_loans(parsed)
            _sync_state["counter"] += 1
            return {"ok": True, "loans_parsed": len(parsed), **stats}
        except Exception as e:
            return {"error": f"{type(e).__name__}: {e}"}

    def import_loan_file(self, path):
        """Import a draw-log file (.xlsx/.xlsm or .csv/.tsv) into the local DB
        so Submit Draw can be tested on loans that aren't on the synced sheet.
        Expected columns, in order:
            Draw # | Name | Address | AA | BL | Last 4 | Draw Fee |
            Wire Received | OUT | Going Out | Released On
        A header row is fine — it's skipped automatically."""
        if not path or not isinstance(path, str):
            return {"error": "No file selected."}
        if not os.path.exists(path):
            return {"error": f"File not found: {path}"}
        try:
            text = _read_file_as_tsv(path)
        except Exception as e:
            return {"error": f"Couldn't read file: {type(e).__name__}: {e}"}
        parsed = parse_draw_log(text or "")
        if not parsed:
            return {"error": "No rows found. Need columns: Draw #, Name, "
                             "Address, AA, BL, Last 4, Draw Fee, Wire Received, "
                             "OUT, Going Out, Released On — with a real Address "
                             "in column C."}
        stats = import_loans(parsed)
        _sync_state["counter"] += 1
        imported = [{"cf_number": p["cf_number"],
                     "address":   p.get("address", ""),
                     "name":      p.get("borrower_name", ""),
                     "draws":     len(p.get("draws", []))} for p in parsed]
        return {"ok": True, "loans_parsed": len(parsed),
                "imported": imported, "source": os.path.basename(path), **stats}

    # ----- queries (unchanged surface) -----
    def get_summary(self):
        with _db_lock, _db() as c:
            loans = c.execute("SELECT COUNT(*) n FROM loans").fetchone()["n"]
            in_flight = c.execute("""SELECT COUNT(*) n FROM draws
                                     WHERE current_state NOT IN ('pending','funded')""").fetchone()["n"]
            funded_count = c.execute("SELECT COUNT(*) n FROM draws WHERE current_state='funded'").fetchone()["n"]
            funded_amt = c.execute("""SELECT COALESCE(SUM(approved_amount),0) a FROM draws
                                      WHERE current_state='funded'""").fetchone()["a"]
            rows = c.execute("""SELECT current_state, state_entered_at FROM draws
                                WHERE current_state NOT IN ('pending','funded')""").fetchall()
        stale = 0
        for r in rows:
            thr = STALE_DAYS.get(r["current_state"])
            days = _days_since(r["state_entered_at"]) or 0
            if thr is not None and days >= thr: stale += 1
        return {"loans": loans, "in_flight": in_flight,
                "funded_count": funded_count, "funded_amount": funded_amt,
                "stale": stale}

    def get_kanban(self):
        with _db_lock, _db() as c:
            rows = c.execute("""SELECT d.id, d.draw_number, d.current_state, d.state_entered_at,
                                       d.requested_amount, d.approved_amount, d.funded_date,
                                       l.id as loan_id, l.cf_number, l.borrower_name, l.address
                                FROM draws d JOIN loans l ON l.id = d.loan_id
                                WHERE d.current_state != 'pending'
                                ORDER BY d.state_entered_at ASC""").fetchall()
        by_state = {s: [] for s in STATE_ORDER}
        for r in rows:
            d = dict(r)
            days = _days_since(d["state_entered_at"])
            d["days_in_state"] = days if days is not None else 0
            thr = STALE_DAYS.get(d["current_state"])
            d["stale"] = (thr is not None and d["days_in_state"] >= thr)
            by_state.setdefault(d["current_state"], []).append(d)
        return {"states": [{"key": s, "label": STATE_LABEL[s],
                            "stale_days": STALE_DAYS.get(s),
                            "cards": by_state.get(s, [])} for s in STATE_ORDER]}

    def get_progress_kanban(self):
        """Same shape as get_kanban but bucketed by the manual progress_state
        flow. Cards include enough context for the drawer and drag-and-drop."""
        with _db_lock, _db() as c:
            rows = c.execute("""SELECT d.id, d.draw_number, d.progress_state,
                                       d.current_state, d.requested_amount,
                                       d.approved_amount, d.funded_date, d.updated_at,
                                       l.id as loan_id, l.cf_number, l.borrower_name, l.address,
                                       (SELECT MAX(st.at) FROM state_transitions st
                                          WHERE st.draw_id = d.id AND st.source != 'sync') AS last_manual_at
                                FROM draws d JOIN loans l ON l.id = d.loan_id""").fetchall()
        by_state = {s: [] for s in PROGRESS_ORDER}
        for r in rows:
            d = dict(r)
            ps = d.get("progress_state") or "order_inspection"
            d["progress_state"] = ps
            d["days_in_state"] = _days_since(d.get("last_manual_at") or d.get("updated_at")) or 0
            by_state.setdefault(ps, []).append(d)
        return {"states": [{"key": s, "label": PROGRESS_LABEL[s],
                            "cards": by_state.get(s, [])} for s in PROGRESS_ORDER]}

    def get_inbox(self, limit=12):
        with _db_lock, _db() as c:
            rows = c.execute("""SELECT d.id, d.draw_number, d.current_state, d.state_entered_at,
                                       d.requested_amount, d.approved_amount,
                                       l.id as loan_id, l.cf_number, l.borrower_name, l.address
                                FROM draws d JOIN loans l ON l.id = d.loan_id
                                WHERE d.current_state NOT IN ('pending', 'funded')
                                ORDER BY d.state_entered_at ASC LIMIT ?""", (limit,)).fetchall()
        items = []
        for r in rows:
            d = dict(r)
            days = _days_since(d["state_entered_at"])
            d["days_in_state"] = days if days is not None else 0
            thr = STALE_DAYS.get(d["current_state"])
            d["stale"] = (thr is not None and d["days_in_state"] >= thr)
            d["state_label"] = STATE_LABEL.get(d["current_state"], d["current_state"])
            items.append(d)
        return items

    def get_loans(self):
        with _db_lock, _db() as c:
            rows = c.execute("""SELECT l.*,
                                       COUNT(d.id) as draws_count,
                                       SUM(CASE WHEN d.current_state='funded' THEN 1 ELSE 0 END) as funded_count,
                                       SUM(CASE WHEN d.current_state NOT IN ('pending','funded') THEN 1 ELSE 0 END) as inflight_count,
                                       COALESCE(SUM(CASE WHEN d.current_state='funded' THEN d.approved_amount ELSE 0 END),0) as drawn_amount,
                                       MAX(d.updated_at) as last_activity
                                FROM loans l LEFT JOIN draws d ON d.loan_id = l.id
                                GROUP BY l.id ORDER BY l.cf_number ASC""").fetchall()
        return [dict(r) for r in rows]

    def get_loan_detail(self, loan_id):
        with _db_lock, _db() as c:
            loan = c.execute("SELECT * FROM loans WHERE id=?", (loan_id,)).fetchone()
            if not loan: return {"error": "Loan not found"}
            draws = [dict(r) for r in c.execute(
                "SELECT * FROM draws WHERE loan_id=? ORDER BY draw_number ASC",
                (loan_id,)).fetchall()]
            if draws:
                ids = [d["id"] for d in draws]
                qmarks = ",".join("?" * len(ids))
                hist_rows = c.execute(
                    f"""SELECT draw_id, from_state, to_state, at, source
                        FROM state_transitions
                        WHERE draw_id IN ({qmarks})
                        ORDER BY at DESC""", ids).fetchall()
            else:
                hist_rows = []
            # Cross-references: every row in sheet_data with our CF#
            cf = (loan["cf_number"] or "").upper().replace(" ", "")
            xref_rows = c.execute(
                """SELECT sheet_name, data_json FROM sheet_data
                   WHERE cf_number=? ORDER BY sheet_name, row_index""",
                (cf,)).fetchall() if cf else []
        by_draw: dict[int, list] = {}
        for h in hist_rows:
            by_draw.setdefault(h["draw_id"], []).append(dict(h))
        for d in draws:
            d["progress_state"] = d.get("progress_state") or "not_started"
            d["history"] = by_draw.get(d["id"], [])
        cross_refs: dict[str, list] = {}
        for r in xref_rows:
            try:
                cross_refs.setdefault(r["sheet_name"], []).append(json.loads(r["data_json"]))
            except Exception:
                pass
        return {"loan": dict(loan), "draws": draws, "cross_refs": cross_refs}

    def draw_set_progress(self, draw_id, new_state):
        """Move a draw to a new manual progress state. Records the transition
        in state_transitions with source='manual' so the history is preserved.
        If auto_write is enabled in settings, also pushes the change back to
        the Google Sheet."""
        if new_state not in PROGRESS_ORDER:
            return {"error": f"Invalid progress state: {new_state}"}
        now = _now()
        loan_row_for_write = None
        draw_number_for_write = None
        with _db_lock, _db() as c:
            row = c.execute(
                """SELECT d.progress_state, d.draw_number,
                          l.sheet_row_index, l.cf_number
                   FROM draws d JOIN loans l ON l.id = d.loan_id
                   WHERE d.id=?""", (draw_id,)).fetchone()
            if not row: return {"error": "Draw not found"}
            old = row["progress_state"] or "order_inspection"
            if old == new_state:
                return {"ok": True, "unchanged": True, "state": new_state}
            c.execute("UPDATE draws SET progress_state=?, updated_at=? WHERE id=?",
                      (new_state, now, draw_id))
            c.execute("""INSERT INTO state_transitions
                           (draw_id, from_state, to_state, at, source)
                         VALUES (?, ?, ?, ?, 'manual')""",
                      (draw_id, old, new_state, now))
            c.commit()
            loan_row_for_write = {
                "sheet_row_index": row["sheet_row_index"],
                "cf_number":       row["cf_number"],
            }
            draw_number_for_write = row["draw_number"]

        # Optional write-back to Google Sheets. Only attempted if auto_write is
        # on. Failures are reported but do not roll back the local state change.
        sheet_result = None
        s = _load_settings()
        gs = s["connections"]["google_sheets"]
        if (gs.get("write_enabled") and gs.get("auto_write")
                and loan_row_for_write and draw_number_for_write):
            sheet_result = _push_progress_to_sheet(
                loan_row_for_write, draw_number_for_write, new_state,
            )
        out = {"ok": True, "from": old, "to": new_state, "at": now}
        if sheet_result is not None:
            out["sheet_write"] = sheet_result
        return out

    def gsheet_push_draw_progress(self, draw_id):
        """Manually push a draw's CURRENT progress state to the sheet.
        Useful as a one-shot 'sync me to the sheet' button without enabling
        auto-write."""
        with _db_lock, _db() as c:
            row = c.execute(
                """SELECT d.progress_state, d.draw_number,
                          l.sheet_row_index, l.cf_number
                   FROM draws d JOIN loans l ON l.id = d.loan_id
                   WHERE d.id=?""", (draw_id,)).fetchone()
        if not row: return {"error": "Draw not found"}
        return _push_progress_to_sheet(
            {"sheet_row_index": row["sheet_row_index"], "cf_number": row["cf_number"]},
            row["draw_number"],
            row["progress_state"] or "not_started",
        )

    def gsheet_test_write(self):
        """Round-trip write test: queries the sheet's actual grid bounds,
        writes a sentinel into the last cell (least likely to overlap real
        data), then clears it. If the service account can write, this
        succeeds; if not, the error tells us why (usually permissions)."""
        s  = _load_settings()
        gs = s["connections"]["google_sheets"]
        path = gs.get("credentials_path", "")
        if _is_local_source(path):
            return {"error": "Source is a local file. Write-back only works with Google Sheets."}
        if not (path and os.path.exists(path) and gs.get("spreadsheet_id")):
            return {"error": "Google Sheets not configured"}
        sheet = gs.get("sheet_name", "Draw")
        try:
            # First, find the sheet's actual grid dimensions so we don't write
            # past the right edge (which returns a 400 "exceeds grid limits").
            service = _gsheet_service(path, write=True)
            meta = service.spreadsheets().get(
                spreadsheetId=gs["spreadsheet_id"],
                fields="sheets(properties(title,gridProperties))",
            ).execute()
            grid = None
            for sh in meta.get("sheets", []):
                if sh["properties"]["title"] == sheet:
                    grid = sh["properties"].get("gridProperties", {})
                    break
            if grid is None:
                return {"error": f"Sheet '{sheet}' not found in workbook"}
            n_rows = int(grid.get("rowCount", 1000))
            n_cols = int(grid.get("columnCount", 26))
            # Write to the very last cell — almost certainly empty padding.
            test_row, test_col = n_rows, n_cols
            stamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S write-test")
            gsheet_write_cell(path, gs["spreadsheet_id"], sheet,
                              test_row, test_col, stamp)
            gsheet_write_cell(path, gs["spreadsheet_id"], sheet,
                              test_row, test_col, "")
            return {"ok": True,
                    "cell":    f"{_col_letter(test_col)}{test_row}",
                    "grid":    f"{n_rows} rows × {n_cols} cols"}
        except Exception as e:
            return {"error": f"{type(e).__name__}: {e}"}

    def get_progress_states(self):
        """Constants for the JS UI — kept here so we have a single source of truth."""
        return {"order": PROGRESS_ORDER, "labels": PROGRESS_LABEL}

    def draw_progress_action(self, draw_id: int, new_state: str):
        """Single entry point for advancing a draw's progress, used by the
        kanban drag-drop and the drawer's Next/Back buttons.

        Returns one of:
          {"ok": True, "action": "moved", "from": ..., "to": ...}
            — state was set without any side effect
          {"ok": True, "action": "run_script", "script": "draw-request",
           "draw_id": ..., "address": ...}
            — JS should now invoke the script; state will advance via the
              poll_task auto-advance hook on success
        """
        if new_state not in PROGRESS_ORDER:
            return {"error": f"Invalid progress state: {new_state}"}
        with _db_lock, _db() as c:
            row = c.execute("""SELECT d.progress_state, d.id as did, l.address
                               FROM draws d JOIN loans l ON l.id = d.loan_id
                               WHERE d.id=?""", (draw_id,)).fetchone()
            if not row:
                return {"error": "Draw not found"}
            old = row["progress_state"] or "order_inspection"
            old_i = PROGRESS_ORDER.index(old) if old in PROGRESS_ORDER else 0
            new_i = PROGRESS_ORDER.index(new_state)

            # Has this draw ever had the script run for it?
            had_script = c.execute("""SELECT 1 FROM state_transitions
                                      WHERE draw_id=? AND source='script' LIMIT 1""",
                                   (draw_id,)).fetchone() is not None

        # The MERGE: advancing INTO order_inspection from earlier AND no past
        # script run means the user actually wants to submit a draw request.
        # The state change itself comes later via the poll_task auto-advance.
        if (new_state == "order_inspection" and new_i > old_i and not had_script):
            return {"ok": True, "action": "run_script", "script": "draw-request",
                    "draw_id": draw_id, "address": row["address"]}

        # Otherwise it's just a plain state move.
        return self.draw_set_progress(draw_id, new_state) | {"action": "moved"}

    def set_total_draws(self, loan_id: int, total: int):
        """Set the planned/expected total number of draws on a loan."""
        try:
            total = max(0, int(total))
        except (TypeError, ValueError):
            return {"error": "Total must be a non-negative integer"}
        now = _now()
        with _db_lock, _db() as c:
            r = c.execute("UPDATE loans SET total_draws=?, updated_at=? WHERE id=?",
                          (total, now, loan_id))
            c.commit()
            if r.rowcount == 0:
                return {"error": "Loan not found"}
        return {"ok": True, "total_draws": total}

    def add_draw(self, loan_id: int):
        """Create a new draw on a loan, numbered one above the current max.
        Starts at order_inspection so the next click runs the script."""
        now = _now()
        with _db_lock, _db() as c:
            loan = c.execute("SELECT id FROM loans WHERE id=?", (loan_id,)).fetchone()
            if not loan: return {"error": "Loan not found"}
            mx = c.execute("SELECT COALESCE(MAX(draw_number), 0) as m FROM draws WHERE loan_id=?",
                           (loan_id,)).fetchone()["m"]
            new_n = mx + 1
            cur = c.execute("""INSERT INTO draws
                                 (loan_id, draw_number, current_state, progress_state,
                                  state_entered_at, imported_at, updated_at)
                               VALUES (?, ?, 'pending', 'not_started', ?, ?, ?)""",
                            (loan_id, new_n, now, now, now))
            c.commit()
        return {"ok": True, "draw_number": new_n, "draw_id": cur.lastrowid}

    def get_recent_errors(self, since=None):
        """Return recent errors from the ring buffer. If `since` is provided,
        only return errors with `at` > since (ISO timestamp string)."""
        with _err_lock:
            buf = list(_ERROR_BUF)
        if since:
            buf = [e for e in buf if e["at"] > since]
        return buf

    # ----- script execution -----
    def run_draw_request(self, address: str, draw_id: int | None = None):
        """Launch the ProxyPics automation for a property.

        If draw_id is provided, the task is linked back to that draw so its
        progress can be advanced when the script completes successfully.
        The draw-items CSV is optional: if one is configured AND exists it is
        passed through to ProxyPics; otherwise the line-items upload is skipped."""
        s  = _load_settings()
        dr = s["scripts"]["draw-request"]
        csv = (dr.get("draw_items_csv") or "").strip()
        # CSV is optional. Only use it when it actually points at a real file;
        # a stale/blank path is silently ignored rather than blocking the run.
        if csv and not os.path.exists(csv):
            print(f">>> [draw-request] Configured CSV not found, ignoring: {csv}")
            csv = ""
        if not (address or "").strip():
            return {"error": "Property address is empty."}

        live = _as_bool(dr.get("live_submit"))
        cmd, err = _build_draw_cmd(s, dr, address, APP_DIR / ".pw-profile", live, csv,
                                   headless=_as_bool(dr.get("headless")))
        if err:
            return err
        proc = subprocess.Popen(
            cmd, stdout=subprocess.PIPE, stderr=subprocess.STDOUT,
            text=True, bufsize=1,
            cwd=str(APP_DIR),
            creationflags=(subprocess.CREATE_NO_WINDOW if os.name == "nt" else 0),
        )
        q: queue.Queue = queue.Queue()
        def reader():
            for line in iter(proc.stdout.readline, ""):
                q.put(line.rstrip("\n"))
            proc.stdout.close(); proc.wait()
            q.put(None)
        threading.Thread(target=reader, daemon=True).start()

        tid = uuid.uuid4().hex[:8]
        pretty = " ".join(f'"{c}"' if " " in c else c for c in cmd)
        with _tasks_lock:
            _tasks[tid] = {"process": proc, "queue": q, "output": [],
                           "done": False, "rc": None, "cmd": pretty,
                           "draw_id": draw_id, "auto_advanced": False}
        return {"taskId": tid, "cmd": pretty}

    def run_draw_request_for_draw(self, draw_id: int):
        """Convenience wrapper used from the loan drawer. Looks up the
        loan address from the DB so the user doesn't have to retype it,
        and threads draw_id through to enable progress auto-advance."""
        with _db_lock, _db() as c:
            row = c.execute("""SELECT l.address, l.cf_number, d.draw_number
                               FROM draws d JOIN loans l ON l.id = d.loan_id
                               WHERE d.id = ?""", (draw_id,)).fetchone()
        if not row: return {"error": "Draw not found"}
        if not row["address"]: return {"error": f"No address on file for {row['cf_number']}"}
        return self.run_draw_request(row["address"], draw_id=draw_id)

    def poll_task(self, task_id: str):
        with _tasks_lock:
            t = _tasks.get(task_id)
        if not t: return {"error": "unknown task"}
        while True:
            try: line = t["queue"].get_nowait()
            except queue.Empty: break
            if line is None:
                t["done"] = True; t["rc"] = t["process"].returncode
            else:
                t["output"].append(line)
        # If the task just finished successfully and was linked to a draw,
        # advance that draw from not_started → order_inspection (the merged
        # "submit draw request" semantic). Fires once per task.
        if (t["done"] and t["rc"] == 0
                and t.get("draw_id") and not t.get("auto_advanced")):
            t["auto_advanced"] = True
            try:
                with _db_lock, _db() as c:
                    row = c.execute("SELECT progress_state FROM draws WHERE id=?",
                                    (t["draw_id"],)).fetchone()
                    if row and (row["progress_state"] or "not_started") == "not_started":
                        now = _now()
                        c.execute("UPDATE draws SET progress_state='order_inspection', updated_at=? WHERE id=?",
                                  (now, t["draw_id"]))
                        c.execute("""INSERT INTO state_transitions
                                       (draw_id, from_state, to_state, at, source)
                                     VALUES (?, 'not_started', 'order_inspection', ?, 'script')""",
                                  (t["draw_id"], now))
                        c.commit()
            except Exception as e:
                _record_error("poll_task auto-advance", e)
        return {"output": t["output"], "done": t["done"], "rc": t["rc"],
                "draw_id": t.get("draw_id")}

    def cancel_task(self, task_id: str):
        with _tasks_lock:
            t = _tasks.get(task_id)
        if t and not t["done"]:
            proc = t["process"]
            try: proc.terminate()
            except Exception: pass
            time.sleep(0.3)
            try:
                if proc.poll() is None: proc.kill()
            except Exception: pass
        return True

    # ----- parallel batch execution -----
    def run_draw_request_batch(self, addresses=None):
        """Run the draw-request worker for many addresses concurrently, capped
        at the configured worker count. If `addresses` is omitted, every loan
        address in the DB is used (i.e. everything you imported). The TEST/LIVE
        mode and worker count both come from Settings → Submit Draw Request."""
        s  = _load_settings()
        dr = s["scripts"]["draw-request"]
        live = _as_bool(dr.get("live_submit"))
        try:
            workers = int(float(dr.get("workers") or 7))
        except (TypeError, ValueError):
            workers = 7
        workers = max(1, min(workers, 20))   # sane ceiling

        csv = (dr.get("draw_items_csv") or "").strip()
        if csv and not os.path.exists(csv):
            csv = ""

        # Fail fast with a friendly message if the worker can't be found.
        _, err = _draw_worker_cmd_head(s)
        if err:
            return err

        if addresses is None:
            with _db_lock, _db() as c:
                rows = c.execute(
                    "SELECT address FROM loans "
                    "WHERE address IS NOT NULL AND TRIM(address) != '' ORDER BY id"
                ).fetchall()
            addresses = [r["address"] for r in rows]

        # Clean + dedupe (case-insensitive) while preserving order.
        seen, uniq = set(), []
        for a in (addresses or []):
            a = (a or "").strip()
            if a and a.lower() not in seen:
                seen.add(a.lower()); uniq.append(a)
        addresses = uniq
        if not addresses:
            return {"error": "No addresses to run. Import a draw-log file first."}

        bid = uuid.uuid4().hex[:8]
        items = {a: {"state": "queued", "rc": None, "output": [], "cmd": ""}
                 for a in addresses}
        with _batches_lock:
            _batches[bid] = {"items": items, "order": addresses, "workers": workers,
                             "live": live, "done": False, "started": _now(),
                             "procs": {}}
        threading.Thread(target=self._batch_worker,
                         args=(bid, addresses, workers, dr, s, live, csv),
                         daemon=True).start()
        return {"batchId": bid, "count": len(addresses),
                "workers": workers, "live": live}

    def _batch_worker(self, bid, addresses, workers, dr, s, live, csv):
        """Coordinator: each address runs in its OWN browser window (its own
        cloned, already-logged-in profile), through a pool capped at workers.
        Isolated processes = one run failing or crashing can't take down the rest."""
        import shutil
        from concurrent.futures import ThreadPoolExecutor
        headless = _as_bool(dr.get("headless"))
        base = APP_DIR / ".pw-profile"
        pool: queue.Queue = queue.Queue()
        n_profiles = max(1, min(workers, len(addresses)))
        for i in range(n_profiles):
            pd = APP_DIR / f".pw-profile-w{i}"
            try:
                if base.exists():
                    if pd.exists():
                        shutil.rmtree(pd, ignore_errors=True)
                    # Clone the logged-in profile, dropping lock/singleton files
                    # so each Chromium window can launch from its own copy.
                    shutil.copytree(base, pd, ignore=shutil.ignore_patterns(
                        "Singleton*", "*.lock", "lockfile", "RunningChromeVersion"))
                else:
                    pd.mkdir(parents=True, exist_ok=True)
            except Exception as e:
                _record_error("batch profile clone", e)
                try: pd.mkdir(parents=True, exist_ok=True)
                except Exception: pass
            pool.put(str(pd))

        try:
            with ThreadPoolExecutor(max_workers=workers) as ex:
                list(ex.map(
                    lambda a: self._run_one_address(bid, a, dr, s, live, csv,
                                                    pool, headless),
                    addresses))
        except Exception as e:
            _record_error("batch_worker", e)
        finally:
            with _batches_lock:
                if bid in _batches:
                    _batches[bid]["done"] = True

    def _run_one_address(self, bid, address, dr, s, live, csv, pool, headless):
        """Run one address in its own window. Borrows/returns a profile dir."""
        with _batches_lock:
            b = _batches.get(bid)
        if not b:
            return
        item = b["items"][address]
        profile = pool.get()
        try:
            cmd, err = _build_draw_cmd(s, dr, address, profile, live, csv,
                                       headless=headless)
            if err:
                item["state"] = "failed"; item["rc"] = 2
                item["output"].append(err["error"]); return
            item["cmd"] = " ".join(cmd)
            item["state"] = "running"
            proc = subprocess.Popen(
                cmd, stdout=subprocess.PIPE, stderr=subprocess.STDOUT,
                text=True, bufsize=1, cwd=str(APP_DIR),
                creationflags=(subprocess.CREATE_NO_WINDOW if os.name == "nt" else 0),
            )
            with _batches_lock:
                b["procs"][address] = proc
            for line in iter(proc.stdout.readline, ""):
                item["output"].append(line.rstrip("\n"))
                if len(item["output"]) > 400:
                    del item["output"][: len(item["output"]) - 400]
            proc.stdout.close(); proc.wait()
            item["rc"] = proc.returncode
            item["state"] = "done" if proc.returncode == 0 else "failed"
        except Exception as e:
            item["state"] = "failed"
            item["output"].append(f"launch error: {type(e).__name__}: {e}")
            _record_error("batch _run_one_address", e)
        finally:
            pool.put(profile)

    def poll_batch(self, batch_id: str):
        with _batches_lock:
            b = _batches.get(batch_id)
        if not b:
            return {"error": "unknown batch"}
        counts = {"queued": 0, "running": 0, "done": 0, "failed": 0}
        items = []
        for a in b["order"]:
            it = b["items"][a]
            counts[it["state"]] = counts.get(it["state"], 0) + 1
            items.append({"address": a, "state": it["state"], "rc": it["rc"],
                          "tail": it["output"][-4:]})
        return {"done": b["done"], "workers": b["workers"], "live": b["live"],
                "total": len(b["order"]), "counts": counts, "items": items}

    def cancel_batch(self, batch_id: str):
        """Force-stop a batch: terminate every worker (closing its window),
        hard-killing anything that won't exit."""
        with _batches_lock:
            b = _batches.get(batch_id)
        if not b:
            return {"error": "unknown batch"}
        procs = list(b.get("procs", {}).values())
        for proc in procs:
            try:
                if proc.poll() is None:
                    proc.terminate()
            except Exception:
                pass
        time.sleep(0.4)
        for proc in procs:                      # hard-kill stragglers
            try:
                if proc.poll() is None:
                    proc.kill()
            except Exception:
                pass
        b["done"] = True
        return {"ok": True}


# Universal error capture: wrap every public Api method so uncaught exceptions
# are logged to _ERROR_BUF and surfaced to JS as {"error": ...} instead of
# crashing the bridge.  set_window is internal (Python-side) so skip it.
def _install_error_capture(cls, skip=("set_window",)):
    import functools
    for name in list(vars(cls)):
        if name.startswith("_") or name in skip:
            continue
        attr = vars(cls)[name]
        if not callable(attr):
            continue
        @functools.wraps(attr)
        def wrapped(self, *args, _orig=attr, _name=name, **kwargs):
            try:
                return _orig(self, *args, **kwargs)
            except Exception as e:
                _record_error(f"Api.{_name}", e)
                return {"error": f"{type(e).__name__}: {e}"}
        setattr(cls, name, wrapped)
    return cls
_install_error_capture(Api)


def _apply_windows_icon(api):
    """Set AppUserModelID, apply .ico, cache HWND, force window to front.
    No-op on non-Windows."""
    if not _u32: return
    try:
        ctypes.windll.shell32.SetCurrentProcessExplicitAppUserModelID("drawmanager.casa.v0_3")
    except Exception:
        pass

    def poll_and_set():
        for _ in range(60):  # up to 15s
            hwnd = api._find_our_window() or _u32.FindWindowW(None, "Draw Manager") or None
            if not hwnd:
                time.sleep(0.25); continue
            api._hwnd = hwnd
            # Apply icon (small + big)
            if ICON_PATH.exists():
                try:
                    hicon = _u32.LoadImageW(0, str(ICON_PATH), 1, 0, 0, 0x10)  # IMAGE_ICON, LR_LOADFROMFILE
                    if hicon:
                        for i in (0, 1):  # ICON_SMALL, ICON_BIG
                            _u32.SendMessageW(hwnd, 0x0080, i, hicon)  # WM_SETICON
                except Exception: pass
            # Force foreground. SetForegroundWindow alone is blocked by
            # anti-focus-stealing rules; topmost-toggle bypasses them.
            try:
                _u32.ShowWindow(hwnd, 9)  # SW_RESTORE
                for top in (-1, -2):  # HWND_TOPMOST, HWND_NOTOPMOST
                    _u32.SetWindowPos(hwnd, top, 0, 0, 0, 0, 0x43)  # NOMOVE|NOSIZE|SHOW
                _u32.SetForegroundWindow(hwnd)
                _u32.BringWindowToTop(hwnd)
            except Exception: pass
            return

    threading.Thread(target=poll_and_set, daemon=True).start()


def main():
    _init_db()
    _save_settings({})    # ensures migrated structure is written on first run
    api = Api()
    window = webview.create_window(
        title="Draw Manager",
        url=str(HTML_PATH),
        js_api=api,
        frameless=True,
        easy_drag=False,
        width=1380, height=860,
        min_size=(1024, 640),
        background_color="#D8E4EF",
    )
    api.set_window(window)
    sync = SyncManager(); sync.start()
    _start_update_check()
    _apply_windows_icon(api)
    try:
        webview.start(debug=False, icon=str(APP_DIR / "draw_manager.ico"))
    except TypeError:
        webview.start(debug=False)


if __name__ == "__main__":
    main()