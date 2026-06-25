"""
casa_api.py — pywebview bridge for Casa Balance Updater.

Replaces the tkinter CasaApp GUI (casa_logic.CasaApp) with a JSON-speaking
Api class that the HTML front-end (casa_updater.html) drives, mirroring how
Draw Manager's app.py exposes an Api to its webview.

ALL business logic still lives in casa_logic.py (the original casa_updater.py,
functions lines 1-1666) and is imported unchanged. This file only:
  - holds session state (properties, pdf queue, detected tabs, driver, config)
  - runs long ops on background threads
  - streams log lines + progress to the UI via a poll-able event queue
  - turns the modal "manual FCI entry" dialog into a pending-prompt the UI
    renders and answers via submit_manual()

The contract reproduced here matches CasaApp's handlers:
  detect_tabs / pull / fci / save / run_all / launch_chrome / google_signin /
  test_api / add_pdfs / remove_pdf / edit_cell / get/set config / open file.
"""

import os
import queue
import threading
import traceback
from datetime import datetime

import casa_logic as L


def _fmt_amount(amt):
    return f"${amt:,.2f}" if isinstance(amt, (int, float)) else ""


def _fmt_date(d):
    if isinstance(d, datetime):
        return d.strftime("%m/%d/%Y")
    return d or ""


class CasaApi:
    """One instance per window. Methods are called from JS via pywebview."""

    # property fields the preview table lets you edit inline (col -> dict key),
    # identical to CasaApp._EDITABLE
    _EDITABLE = {"address": "address", "amount": "amount", "cf": "cf_number",
                 "loan": "loan_number", "acct": "account_number",
                 "date": "date_received"}

    def __init__(self):
        self.cfg = L.load_config()
        self.driver = None
        self.properties = []        # rows pulled from Sheets and/or PDFs
        self.existing = []          # rows already in the output xlsx
        self.pdf_paths = []         # PDFs queued for parsing
        self.detected_tabs = []     # tabs detected from the workbook

        # UI event stream (log lines + progress + status), polled by JS.
        self._events = queue.Queue()
        self._busy = False          # a background op is running
        self._lock = threading.Lock()

        # pending manual-FCI prompt: when set, the UI shows a dialog and
        # answers via submit_manual(); the worker thread waits on _manual_evt.
        self._manual_prop = None
        self._manual_result = None
        self._manual_evt = threading.Event()
        self._window = None

    # ----- window wiring (set by main after window creation) --------------
    def set_window(self, w):
        self._window = w

    # ======================================================================
    # Event stream helpers (replace tkinter log / set_progress / status)
    # ======================================================================
    def _emit(self, kind, **payload):
        payload["kind"] = kind
        self._events.put(payload)

    def log(self, msg):
        ts = datetime.now().strftime("%H:%M:%S")
        self._emit("log", line=f"[{ts}] {msg}")

    def status(self, msg):
        self._emit("status", text=msg)

    def set_progress(self, done, total, msg=None):
        pct = (float(done) / float(total) * 100.0) if total else 0.0
        self._emit("progress", pct=round(pct, 1),
                   done=done, total=total, text=(msg or ""))
        if msg:
            self.status(msg)

    def get_events(self):
        """JS polls this ~4x/sec. Drains and returns queued UI events plus a
        snapshot of busy state and any pending manual prompt."""
        out = []
        try:
            while True:
                out.append(self._events.get_nowait())
        except queue.Empty:
            pass
        pending = None
        if self._manual_prop is not None and not self._manual_evt.is_set():
            p = self._manual_prop
            pending = {
                "address": p.get("address", ""),
                "amount": _fmt_amount(p.get("amount")),
                "cf": p.get("cf_number") or "",
                "draw": p.get("draw_num") or "",
                "date": _fmt_date(p.get("date_received")),
            }
        return {"events": out, "busy": self._busy, "manual": pending}

    # ======================================================================
    # Config get / set  (mirrors CasaApp v_* vars + _save_config)
    # ======================================================================
    def get_config(self):
        c = self.cfg
        return {
            "sheet_url":        c.get("sheet_url", ""),
            "today_sheet_name": c.get("today_sheet_name", ""),
            "draw_sheet_name":  c.get("draw_sheet_name", "Draw"),
            "output_xlsx":      c.get("output_xlsx", ""),
            "fci_api_key":      c.get("fci_api_key", ""),
            "fci_auto":         bool(c.get("fci_auto", True)),
            "chrome_headless":  bool(c.get("chrome_headless", False)),
            # convenience state for the UI
            "tabs":             list(self.detected_tabs),
            "pdfs":             [os.path.basename(p) for p in self.pdf_paths],
        }

    def set_config(self, d):
        d = d or {}
        self.cfg.update({
            "sheet_url":        (d.get("sheet_url") or "").strip(),
            "today_sheet_name": (d.get("today_sheet_name") or "").strip(),
            "draw_sheet_name":  (d.get("draw_sheet_name") or "").strip() or "Draw",
            "output_xlsx":      (d.get("output_xlsx") or "").strip(),
            "fci_auto":         bool(d.get("fci_auto", True)),
            "chrome_headless":  bool(d.get("chrome_headless", False)),
            "fci_api_key":      (d.get("fci_api_key") or "").strip(),
        })
        L.save_config(self.cfg)
        return {"ok": True}

    # ======================================================================
    # Preview table snapshot + inline edit
    # ======================================================================
    def get_properties(self):
        rows = []
        for i, p in enumerate(self.properties):
            st = p.get("_status")
            status = "skip" if st == "skip" else ("done" if st == "done" else "pending")
            rows.append({
                "i": i,
                "status": status,
                "draw": p.get("draw_num") or "",
                "address": p.get("address", ""),
                "amount": _fmt_amount(p.get("amount")),
                "cf": p.get("cf_number") or "",
                "loan": p.get("loan_number") or "",
                "acct": p.get("account_number") or "",
                "date": _fmt_date(p.get("date_received")),
            })
        return rows

    def edit_cell(self, idx, col, value):
        """Inline edit from the preview table. Mirrors CasaApp._on_tree_edit
        commit logic (type coercion per field)."""
        try:
            idx = int(idx)
        except (TypeError, ValueError):
            return {"ok": False, "error": "bad row"}
        if idx < 0 or idx >= len(self.properties):
            return {"ok": False, "error": "row out of range"}
        if col not in self._EDITABLE:
            return {"ok": False, "error": "column not editable"}
        field = self._EDITABLE[col]
        p = self.properties[idx]
        val = (value or "").strip()
        if field == "amount":
            p[field] = L.parse_amount(val)
        elif field == "date_received":
            p[field] = L.parse_date(val) or (val or None)
        elif field in ("loan_number", "account_number"):
            p[field] = L._to_int_or_keep(val) if val else None
        else:
            p[field] = val or None
        return {"ok": True}

    # ======================================================================
    # PDF queue
    # ======================================================================
    def add_pdfs(self):
        """Open a native file picker (pywebview) and queue chosen PDFs."""
        if not self._window:
            return {"ok": False, "error": "no window"}
        import webview
        paths = self._window.create_file_dialog(
            webview.OPEN_DIALOG, allow_multiple=True,
            file_types=("PDF files (*.pdf)", "All files (*.*)"))
        added = 0
        for p in (paths or []):
            if p and p not in self.pdf_paths:
                self.pdf_paths.append(p)
                added += 1
        if added:
            self.status(f"{len(self.pdf_paths)} PDF(s) queued.")
        return {"ok": True, "pdfs": [os.path.basename(p) for p in self.pdf_paths]}

    def remove_pdf(self, idx):
        try:
            idx = int(idx)
            del self.pdf_paths[idx]
        except (TypeError, ValueError, IndexError):
            return {"ok": False}
        self.status(f"{len(self.pdf_paths)} PDF(s) queued.")
        return {"ok": True, "pdfs": [os.path.basename(p) for p in self.pdf_paths]}

    def browse_output(self):
        """Native save-file picker for the output xlsx."""
        if not self._window:
            return {"ok": False, "error": "no window"}
        import webview
        initial = self.cfg.get("output_xlsx") or "balance_update.xlsx"
        res = self._window.create_file_dialog(
            webview.SAVE_DIALOG,
            save_filename=os.path.basename(initial),
            file_types=("Excel workbook (*.xlsx)", "All files (*.*)"))
        path = res if isinstance(res, str) else (res[0] if res else None)
        if path:
            self.cfg["output_xlsx"] = path
            L.save_config(self.cfg)
            return {"ok": True, "path": path}
        return {"ok": False}

    # ======================================================================
    # Background-op plumbing
    # ======================================================================
    def _run_bg(self, target):
        """Start target() on a daemon thread, guarding against overlap."""
        with self._lock:
            if self._busy:
                self.log("Busy — wait for the current step to finish.")
                return {"ok": False, "busy": True}
            self._busy = True

        def wrapper():
            try:
                target()
            except Exception as e:
                self.log(f"ERROR: {e}")
                self._emit("error", message=str(e),
                           trace=traceback.format_exc())
            finally:
                self._busy = False
        threading.Thread(target=wrapper, daemon=True).start()
        return {"ok": True}

    def _ensure_driver(self, auto_launch=False):
        """Mirror of CasaApp._ensure_driver."""
        if self.driver is not None:
            return self.driver
        port = int(self.cfg["chrome_debug_port"])
        if not L._port_listening("127.0.0.1", port):
            if not auto_launch:
                raise RuntimeError(
                    "Chrome isn't open yet. Click 'Launch Chrome' first, sign "
                    "in to Google + FCI in the window that opens, then retry.")
            self.log("Chrome isn't running — launching it now…")
            sheet_url = self.cfg.get("sheet_url", "").strip()
            L.launch_chrome_debug(self.cfg,
                                  extra_tabs=[sheet_url] if sheet_url else [])
        self.log("Attaching to Chrome…")
        self.driver = L.attach_to_chrome(port)
        self.log("Attached.")
        return self.driver

    def _selected_tabs(self, tabs_from_ui):
        """tabs_from_ui: list the UI says are selected. Falls back to the
        comma override in today_sheet_name, like CasaApp._selected_tabs."""
        if tabs_from_ui:
            return [t for t in tabs_from_ui if t]
        override = self.cfg.get("today_sheet_name", "").strip()
        if override:
            return [t.strip() for t in override.split(",") if t.strip()]
        return []

    # ======================================================================
    # Step 1: Chrome
    # ======================================================================
    def launch_chrome(self):
        return self._run_bg(self._launch_chrome)

    def _launch_chrome(self):
        sheet_url = self.cfg.get("sheet_url", "").strip()
        extra = [sheet_url] if sheet_url else []
        self.log("Launching Chrome in debug mode…")
        L.launch_chrome_debug(self.cfg, extra_tabs=extra)
        self.log(f"Chrome is up on port {self.cfg['chrome_debug_port']}.")
        self.log("Sign in to Google + FCI in that Chrome window, open the "
                 "Sheet once to confirm auth, then click Pull / Parse.")
        self._ensure_driver()

    # ======================================================================
    # Google OAuth + FCI API test
    # ======================================================================
    def google_signin(self):
        return self._run_bg(self._google_signin)

    def _google_signin(self):
        self.log("Starting Google OAuth flow…")
        L.get_google_creds(interactive=True)
        self.log("✓ Signed in to Google. Token saved.")
        self.log("  Pull will now use the Sheets API.")

    def test_api(self, api_key=None):
        if api_key is not None:
            self.cfg["fci_api_key"] = (api_key or "").strip()
            L.save_config(self.cfg)
        return self._run_bg(self._test_api)

    def _test_api(self):
        key = self.cfg.get("fci_api_key", "").strip()
        if not key:
            self.log("No FCI API key set.")
            return
        self.log("Testing FCI API key…")
        try:
            data = L.fci_introspect(key)
            summary = L.fci_introspect_summary(data)
            self.log("✓ FCI API key works.")
            if summary:
                self.log(summary)
        except Exception as e:
            self.log(f"FCI API test failed: {e}")
            raise

    # ======================================================================
    # Step 2: pull / parse  (sheets + PDFs)  — mirror of _pull_thread
    # ======================================================================
    def pull(self, selected_tabs=None):
        return self._run_bg(lambda: self._pull(selected_tabs or []))

    def _pull(self, selected_tabs):
        sheet_url = self.cfg.get("sheet_url", "").strip()
        draw_name = self.cfg.get("draw_sheet_name", "").strip() or "Draw"
        sel_tabs = self._selected_tabs(selected_tabs)
        pdf_paths = list(self.pdf_paths)

        # Auto-detect: if no tabs picked and no PDFs, but we have a sheet URL,
        # detect tabs and auto-select the date ("/") tabs before pulling.
        if not sel_tabs and not pdf_paths and sheet_url:
            self.log("No tabs selected — auto-detecting…")
            self._detect_tabs()
            sel_tabs = [t for t in self.detected_tabs if "/" in t]
            if sel_tabs:
                self.log(f"Auto-selected: {sel_tabs}")

        if not sel_tabs and not pdf_paths:
            raise RuntimeError("Nothing selected. Detect & select sheet "
                               "tab(s) and/or add PDF(s), then Pull / Parse.")

        have_sheets = bool(sel_tabs and sheet_url)
        total_steps = len(sel_tabs) + len(pdf_paths) + (1 if have_sheets else 0)
        step = 0
        self.set_progress(0, total_steps, "Starting…")

        all_properties = []
        idx, flat = {}, []
        creds = drv = None
        use_api = False

        if have_sheets:
            sheet_id = L.extract_sheet_id(sheet_url)
            if not sheet_id:
                raise RuntimeError("That doesn't look like a Google Sheets URL.")
            use_api = L.google_creds_path().exists() and L.google_token_path().exists()
            if use_api:
                try:
                    creds = L.get_google_creds(interactive=False)
                    self.log("Using Google Sheets API.")
                except Exception as e:
                    self.log(f"API creds problem: {e} — falling back to Chrome.")
                    use_api = False
            if not use_api:
                drv = self._ensure_driver(auto_launch=True)
                self.log("Using Chrome session.")

            def fetch_tab(name):
                if use_api:
                    return L.fetch_sheet_via_api(creds, sheet_id, name)
                return L.fetch_sheet_via_driver(drv, sheet_id, sheet_name=name)

            self.log(f"Parsing {len(sel_tabs)} sheet tab(s): {sel_tabs}")
            for tab in sel_tabs:
                self.set_progress(step, total_steps, f"Sheet tab '{tab}'")
                try:
                    rows = fetch_tab(tab)
                    props = L.parse_today_sheet(rows)
                    for p in props:
                        p["_source_tab"] = tab
                    self.log(f"  '{tab}': {len(rows)} rows, {len(props)} properties.")
                    all_properties.extend(props)
                except Exception as e:
                    self.log(f"  '{tab}': skipped — {e}")
                step += 1

            self.set_progress(step, total_steps, f"Fetching '{draw_name}'")
            try:
                draw_rows = fetch_tab(draw_name)
                idx, flat = L.build_draw_index(draw_rows)
                self.log(f"  '{draw_name}': {len(draw_rows)} rows, "
                         f"{sum(len(v) for v in idx.values())} addresses indexed.")
            except Exception as e:
                self.log(f"  couldn't fetch Draw sheet for CF#: {e}")
            step += 1
        elif sel_tabs and not sheet_url:
            self.log("Tabs selected but no Sheet URL — ignoring tabs.")

        # ---- PDFs ----
        for path in pdf_paths:
            fname = os.path.basename(path)
            self.set_progress(step, total_steps, f"Parsing PDF '{fname}'")

            def pdf_progress(done, tot, msg, _base=step):
                self.status(msg)
            try:
                pdf_props = L.parse_pdf_draw_sheet(path, log=self.log,
                                                   progress=pdf_progress)
                all_properties.extend(pdf_props)
            except Exception as e:
                self.log(f"  PDF '{fname}' failed: {e}")
            step += 1

        self.log(f"Total properties: {len(all_properties)}")
        if not all_properties:
            self.set_progress(total_steps, total_steps, "Nothing parsed.")
            self.properties = []
            self._emit("refresh")
            return

        # ---- Enrich: CF# from Draw; date = released ----
        for p in all_properties:
            if idx or flat:
                match = L.lookup_draw(idx, flat, p["address"])
                p["cf_number"] = match["cf"] if match else p.get("cf_number")
            else:
                p.setdefault("cf_number", None)
            p["date_received"] = p.get("released_on")
            p["_date_tag"] = self._date_tag_for(p)

        # ---- Dedupe vs existing master output ----
        self.existing, _ = L.load_existing_rows(self.cfg.get("output_xlsx", "").strip())
        existing_by_key = {
            (L.addr_norm(r.get("address") or ""), L.parse_amount(r.get("amount"))): r
            for r in self.existing
        }
        for p in all_properties:
            key = (L.addr_norm(p["address"]), p.get("amount"))
            ex = existing_by_key.get(key)
            if ex and L.is_row_complete(ex):
                p["_status"] = "skip"
            else:
                if ex:
                    p.setdefault("loan_number", ex.get("loan_number"))
                    p.setdefault("account_number", ex.get("account_number"))
                p["_status"] = "pending"

        self.properties = all_properties
        self._emit("refresh")
        self.set_progress(total_steps, total_steps, "Pull / parse complete.")
        self.log("Pull / parse complete. Review the table, then Run FCI / Save.")

    # ======================================================================
    # Step 3: FCI  — mirror of _fci_thread, manual dialog -> pending prompt
    # ======================================================================
    def fci(self):
        if not self.properties:
            self.log("Nothing to look up — pull properties first.")
            return {"ok": False}
        return self._run_bg(self._fci)

    def _fci(self):
        api_key = self.cfg.get("fci_api_key", "").strip()
        if not api_key:
            raise RuntimeError("Enter your FCI API key in Settings and click "
                               "'Test API' once, then try again.")
        use_auto = bool(self.cfg.get("fci_auto", True))
        addr_field = self.cfg.get("fci_loan_address_field") or "propertyStreet"
        amt_field = self.cfg.get("fci_funding_amount_field") or "originalBalance"

        # A property needs FCI if it's not skipped and is missing EITHER number.
        to_lookup = []
        already_ok = 0
        skipped = 0
        for p in self.properties:
            if p.get("_status") == "skip":
                skipped += 1
                continue
            has_loan = bool(p.get("loan_number"))
            has_acct = bool(p.get("account_number"))
            if has_loan and has_acct:
                already_ok += 1
                continue
            to_lookup.append(p)

        self.log(f"FCI: {len(self.properties)} properties — "
                 f"{skipped} already in file, {already_ok} already have both "
                 f"numbers, {len(to_lookup)} to look up.")
        if not to_lookup:
            self.log("Nothing needs an FCI lookup. If numbers are missing in "
                     "your sheet, those rows may already be saved without them "
                     "— delete those rows from the .xlsx and re-run to refetch.")

        failed = []
        total_lk = len(to_lookup)
        done_lk = [0]
        self.set_progress(0, total_lk or 1, "FCI lookup…")

        def on_result(p, res):
            if res.get("loan_number"):
                p["loan_number"] = L._to_int_or_keep(res["loan_number"])
            if res.get("account_number"):
                p["account_number"] = L._to_int_or_keep(res["account_number"])
            if res.get("error") or not (p.get("loan_number") and p.get("account_number")):
                failed.append((p, res.get("error") or "missing fields"))
            else:
                p["_status"] = "done"
            done_lk[0] += 1
            self.set_progress(done_lk[0], total_lk or 1,
                              f"FCI: {p.get('address','')[:32]}")
            self._emit("refresh")

        if use_auto and to_lookup:
            self.log(f"FCI API: {len(to_lookup)} properties "
                     f"(address={addr_field}, amount={amt_field})…")
            try:
                L.fci_api_lookup_many(api_key, to_lookup, log=self.log,
                                      on_result=on_result,
                                      address_field=addr_field,
                                      amount_field=amt_field)
            except Exception as e:
                self.log(f"API call aborted: {e}")
                failed = [(p, str(e)) for p in to_lookup]
            ok_count = sum(1 for p in to_lookup if p.get("_status") == "done")
            self.log(f"  {ok_count}/{len(to_lookup)} resolved, "
                     f"{len(failed)} need manual entry.")
        elif not use_auto:
            failed = [(p, "auto disabled") for p in to_lookup]

        # Manual fallback: each unresolved property becomes a UI prompt.
        for p, why in failed:
            self.log(f"FCI manual > {p['address']}  ({str(why)[:80]})")
            # Pre-navigate Chrome to FCI search for this property, if attached.
            try:
                if self.driver is not None:
                    fci_url = self.cfg.get("fci_url", L.DEFAULT_CONFIG["fci_url"])
                    L.open_url_in_new_tab(self.driver, fci_url)
            except Exception:
                pass
            dlg = self._await_manual(p)
            if dlg.get("action") == "stop":
                self.log("Stopped by user.")
                break
            if dlg.get("action") == "skip":
                continue
            if dlg.get("loan"):
                p["loan_number"] = L._to_int_or_keep(dlg["loan"])
            if dlg.get("acct"):
                p["account_number"] = L._to_int_or_keep(dlg["acct"])
            if dlg.get("date"):
                p["date_received"] = L.parse_date(dlg["date"]) or dlg["date"]
            p["_status"] = "done"
            self._emit("refresh")

        self.set_progress(total_lk or 1, total_lk or 1, "FCI lookup complete.")
        self.log("FCI lookup complete.")

    def _await_manual(self, prop):
        """Park a manual prompt for the UI and block until submit_manual()."""
        self._manual_prop = prop
        self._manual_result = None
        self._manual_evt.clear()
        self._emit("manual_prompt")
        self._manual_evt.wait()
        res = self._manual_result or {"action": "skip"}
        self._manual_prop = None
        return res

    def submit_manual(self, payload):
        """Called from JS when the user answers the manual FCI dialog.
        payload: {action: 'save'|'skip'|'stop', loan, acct, date}"""
        self._manual_result = payload or {"action": "skip"}
        self._manual_evt.set()
        return {"ok": True}

    # ======================================================================
    # Step 4: save / append  — mirror of _do_save / _save_outputs / merge
    # ======================================================================
    def save(self):
        return self._run_bg(self._save)

    def _save(self):
        out = self.cfg.get("output_xlsx", "").strip()
        if not out:
            raise RuntimeError("Pick an output .xlsx file first (Browse).")
        to_write = []
        for p in self.properties:
            if p.get("_status") == "skip":
                continue
            to_write.append({
                "loan_number":    p.get("loan_number"),
                "account_number": p.get("account_number"),
                "amount":         p.get("amount"),
                "date_received":  p.get("date_received"),
                "address":        p.get("address"),
                "cf_number":      p.get("cf_number"),
                "_date_tag":      self._date_tag_for(p),
            })
        if not to_write:
            self.log("Nothing new to write.")
            return
        self.status("Saving…")
        files = self._save_outputs(out, to_write)
        self.log(f"Saved {len(to_write)} row(s) → master + "
                 f"{len(files)-1} per-date file(s).")
        self.status("Saved.")
        self._emit("saved", files=[os.path.basename(f) for f in files])

    def _date_tag_for(self, p):
        tag = L.tab_date_tag(p.get("_source_tab"))
        if tag:
            return tag
        tag = L.date_tag_from(p.get("date_received") or p.get("released_on"))
        if tag:
            return tag
        return datetime.now().strftime("%m%d")

    def _save_outputs(self, master_path, rows):
        def strip(r):
            return {k: v for k, v in r.items() if k != "_date_tag"}
        written = [master_path]
        self.log(f"[master] {os.path.basename(master_path)}")
        self._merge_and_save(master_path, [strip(r) for r in rows])
        folder = os.path.dirname(master_path) or "."
        groups = {}
        for r in rows:
            groups.setdefault(r.get("_date_tag") or "nodate", []).append(strip(r))
        for tag, grp in sorted(groups.items()):
            per_path = os.path.join(folder, f"balance_update_{tag}.xlsx")
            self.log(f"[{tag}] {os.path.basename(per_path)} ({len(grp)} row(s))")
            self._merge_and_save(per_path, grp)
            written.append(per_path)
        return written

    def _merge_and_save(self, path, new_rows):
        from openpyxl import Workbook
        new_rows = [{k: v for k, v in r.items() if k != "_date_tag"}
                    for r in new_rows]
        existing, _ = L.load_existing_rows(path)
        merged = list(existing)
        merged_by_key = {(L.addr_norm(r["address"] or ""), L.parse_amount(r["amount"])): r
                         for r in merged}
        appended = updated = 0
        for nr in new_rows:
            key = (L.addr_norm(nr["address"] or ""), L.parse_amount(nr["amount"]))
            if key in merged_by_key:
                target = merged_by_key[key]
                for k, v in nr.items():
                    if not target.get(k) and v:
                        target[k] = v
                updated += 1
            else:
                merged.append(nr)
                merged_by_key[key] = nr
                appended += 1
        wb = Workbook()
        ws = wb.active
        ws.append(L.EXCEL_HEADER)
        for r in merged:
            ws.cell(column=1, row=ws.max_row + 1, value=r.get("loan_number"))
            row_idx = ws.max_row
            ws.cell(column=2, row=row_idx, value=r.get("account_number"))
            ws.cell(column=3, row=row_idx, value=r.get("amount"))
            d = r.get("date_received")
            if isinstance(d, datetime):
                cell = ws.cell(column=4, row=row_idx, value=d)
                cell.number_format = "m/d/yyyy"
            else:
                ws.cell(column=4, row=row_idx, value=d)
            ws.cell(column=5, row=row_idx, value=r.get("address"))
            ws.cell(column=6, row=row_idx, value=r.get("cf_number"))
        for col, width in zip("ABCDEF", (14, 18, 12, 14, 32, 12)):
            ws.column_dimensions[col].width = width
        wb.save(path)
        self.log(f"  → {appended} appended, {updated} updated in place.")

    # ======================================================================
    # Detect tabs  — mirror of _detect_tabs_thread
    # ======================================================================
    def detect_tabs(self):
        return self._run_bg(self._detect_tabs)

    def _detect_tabs(self):
        self.status("Detecting workbook tabs…")
        sheet_id = L.extract_sheet_id(self.cfg.get("sheet_url", "").strip())
        if not sheet_id:
            raise RuntimeError("Paste a valid Google Sheet URL first.")
        use_api = L.google_creds_path().exists() and L.google_token_path().exists()
        tabs = []
        if use_api:
            try:
                creds = L.get_google_creds(interactive=False)
                tabs = L.list_sheet_tabs_via_api(creds, sheet_id)
                self.log("Detected tabs via Sheets API.")
            except Exception as e:
                self.log(f"API tab list failed: {e} — trying Chrome.")
                use_api = False
        if not use_api:
            drv = self._ensure_driver(auto_launch=True)
            tabs = L.list_sheet_tabs(drv, sheet_id)
            self.log("Detected tabs via Chrome.")
        self.detected_tabs = tabs
        # auto-select only date-style tabs (those containing a "/", e.g. 06/18)
        selected = [t for t in tabs if "/" in t]
        self.log(f"Found {len(tabs)} tab(s): {tabs}")
        if selected:
            self.log(f"Auto-selected date tabs: {selected}")
        self.status(f"Detected {len(tabs)} tab(s) — {len(selected)} date tab(s) selected.")
        self._emit("tabs", tabs=tabs, selected=selected)

    # ======================================================================
    # Run everything
    # ======================================================================
    def run_all(self, selected_tabs=None):
        return self._run_bg(lambda: self._run_all(selected_tabs or []))

    def _run_all(self, selected_tabs):
        if not self.cfg.get("output_xlsx", "").strip():
            raise RuntimeError("Pick an output .xlsx file first.")
        has_source = (self._selected_tabs(selected_tabs) or self.pdf_paths
                      or self.cfg.get("sheet_url", "").strip())
        if not has_source:
            raise RuntimeError("Set a Sheet URL (in settings) or add PDF(s) first.")
        self.log("=== Run everything: pull/parse → FCI → save ===")
        self._pull(selected_tabs)
        if not self.properties:
            self.log("Nothing parsed — stopping.")
            return
        self._fci()
        self._save()
        self.set_progress(1, 1, "Run everything: done.")
        self.log("Run everything: done.")

    # ======================================================================
    # Open file / folder
    # ======================================================================
    def open_excel(self):
        import platform, subprocess
        out = self.cfg.get("output_xlsx", "").strip()
        if not out or not os.path.exists(out):
            return {"ok": False, "error": "No output file yet — Save first."}
        try:
            if platform.system() == "Windows":
                os.startfile(out)  # noqa
            elif platform.system() == "Darwin":
                subprocess.Popen(["open", out])
            else:
                subprocess.Popen(["xdg-open", out])
            return {"ok": True}
        except Exception as e:
            return {"ok": False, "error": str(e)}

    def open_folder(self):
        import platform, subprocess
        out = self.cfg.get("output_xlsx", "").strip()
        folder = os.path.dirname(out) if out else str(L.APP_DIR)
        if not folder or not os.path.isdir(folder):
            folder = str(L.APP_DIR)
        try:
            if platform.system() == "Windows":
                os.startfile(folder)  # noqa
            elif platform.system() == "Darwin":
                subprocess.Popen(["open", folder])
            else:
                subprocess.Popen(["xdg-open", folder])
            return {"ok": True}
        except Exception as e:
            return {"ok": False, "error": str(e)}
