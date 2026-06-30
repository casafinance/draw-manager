"""
ProxyPics draw-inspection request automation.

Flow:
    1. Prompt for a property address (or pass --address "...").
    2. Look up the loan in Casa Finance (loans → falls back to servicing).
    3. Extract CF# and the Borrower / Guarantor contact rows.
    4. Ask which contact to use.
    5. Open ProxyPics and fill the new-draw-inspection form.
    6. Stop after clicking "Next" (does NOT submit).

First run on each site will need a manual login; sessions persist in
./.pw-profile so subsequent runs go straight through.
"""

from __future__ import annotations

import argparse
import csv as _csv
import os
import re
import sqlite3
import sys
import tempfile
import time
from pathlib import Path
from playwright.sync_api import (
    sync_playwright, Page, TimeoutError as PWTimeout, Locator,
)

# Windows consoles default to cp1252, which chokes on common Unicode like the
# right-arrow (U+2192). Reconfigure stdout/stderr to UTF-8 so prints don't
# crash mid-flow on characters like '->' rendered as '\u2192'.
for _stream in (sys.stdout, sys.stderr):
    try:
        _stream.reconfigure(encoding="utf-8", errors="replace")
    except Exception:
        pass

# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------
PROXYPICS_URL    = "https://app.proxypics.com/photo-requests/draw-inspections/new"
CASA_SERV_URL    = "https://myportal.casa.finance/admin/servicing"
CASA_LOANS_URL   = "https://myportal.casa.finance/admin/loans"

_SCRIPT_DIR      = (Path(sys.executable).parent
                    if getattr(sys, "frozen", False)
                    else Path(__file__).parent)
PROFILE_DIR      = _SCRIPT_DIR / ".pw-profile"   # may be overridden by --profile-dir
LOGIN_WAIT_S     = 300

# Sentinels used to detect "page is ready, not on login".
PROXYPICS_READY  = 'button:has-text("Add contact")'
CASA_READY       = 'input[placeholder="Search"]'

DEFAULTS = {
    "inspection_type_search": "draw inspection without floorplan",
    "csv_path":               "",
    "progress_step":          "25",
}

# Default SharePoint Clients root for the rehab-budget auto-build. Overridable
# via --clients-root or DRAW_CLIENTS_ROOT env var.
DEFAULT_CLIENTS_ROOT = (
    r"C:\Users\LukeQumsieh\Swift Funding"
    r"\Swift Lending - Casa Finance Group\Clients"
)


# ---------------------------------------------------------------------------
# Windows long-path helper. The Clients tree lives several folders deep and
# folder names like '1300 NW 42nd St, Miami, FL 33142' eat the budget fast.
# Prefixing absolute paths with \\?\ tells Win32 to skip the MAX_PATH check
# (effective limit ~32K). No-op off Windows.
# ---------------------------------------------------------------------------
def _long(p) -> str:
    s = os.path.abspath(str(p))
    if os.name != "nt":
        return s
    if s.startswith("\\\\?\\"):
        return s
    if s.startswith("\\\\"):
        return "\\\\?\\UNC\\" + s[2:]
    return "\\\\?\\" + s


# ---------------------------------------------------------------------------
# Rehab-budget auto-build: find the property folder under
#   Clients\<group>\<address>\Rehab Budget\<file>.xlsx
# and convert it to the ProxyPics draw-items CSV format.
# ---------------------------------------------------------------------------
def _norm_for_folder_match(s: str) -> str:
    """Lowercase + collapse whitespace + strip unicode whitespace. Folder
    names sometimes have curly punctuation or NBSPs from copy/paste."""
    s = (s or "").replace("\xa0", " ").replace("\u202f", " ")
    return re.sub(r"\s+", " ", s.strip().lower())


def find_property_folder(clients_root: Path, address: str) -> Path | None:
    """Walk Clients/<group>/<address> looking for a folder whose name starts
    with the address's street line. Returns the property folder, or None.

    Folder names look like '1300 NW 42nd St, Miami, FL 33142'. The caller
    typically passes just the street line ('1300 NW 42nd St') as `address`,
    so we match by prefix on the street portion."""
    if not clients_root or not Path(_long(clients_root)).exists():
        return None
    needle = _norm_for_folder_match(address).split(",", 1)[0].strip()
    if not needle:
        return None
    root = Path(_long(clients_root))
    try:
        groups = [g for g in root.iterdir() if g.is_dir()]
    except Exception as e:
        print(f"    (clients-root scan warning: {e})")
        return None
    # Two-tier: most-specific (full street line prefix), then bare street
    # number (in case folder uses a slightly different street formatting).
    parts = needle.split()
    street_number = parts[0] if parts else ""

    def _match(name_lower: str) -> int:
        # 2 = full street line prefix, 1 = bare number prefix, 0 = no match
        if name_lower.startswith(needle + ",") or name_lower.startswith(needle + " "):
            return 2
        if street_number and (name_lower.startswith(street_number + " ")
                              or name_lower.startswith(street_number + ",")):
            return 1
        return 0

    best: tuple[int, Path] | None = None
    for group in groups:
        try:
            props = [p for p in group.iterdir() if p.is_dir()]
        except Exception:
            continue
        for prop in props:
            score = _match(_norm_for_folder_match(prop.name))
            if score == 0:
                continue
            if best is None or score > best[0]:
                best = (score, prop)
                if score == 2:
                    return prop  # exact street-line prefix wins outright
    return best[1] if best else None


def extract_application_contact(property_folder: Path) -> dict:
    """Read the borrower/guarantor contact from the loan Application PDF in
    '<property>/Application/<file>.pdf'. Returns
    {name, phone, email} (empty strings if not found / no PDF).

    The application is the authoritative source for the point-of-contact.
    Prefers the Primary Guarantor block, falling back to Primary Borrower."""
    blank = {"name": "", "phone": "", "email": ""}
    if property_folder is None:
        return blank
    app_dir = Path(_long(property_folder)) / "Application"
    if not app_dir.is_dir():
        print("    (no Application folder for this property)")
        return blank
    pdf = None
    try:
        cands = [f for f in app_dir.iterdir()
                 if f.is_file() and f.suffix.lower() == ".pdf"
                 and not f.name.startswith("~")]
        if not cands:
            print("    (no PDF in Application folder)")
            return blank
        cands.sort(key=lambda p: p.stat().st_mtime, reverse=True)
        pdf = cands[0]
    except Exception as e:
        print(f"    (Application scan warning: {e})")
        return blank

    try:
        from pypdf import PdfReader
    except Exception:
        print("    (pypdf not installed — can't read the application PDF)")
        return blank
    try:
        reader = PdfReader(_long(pdf))
        text = "\n".join(p.extract_text() or "" for p in reader.pages)
    except Exception as e:
        print(f"    (couldn't read {pdf.name}: {e})")
        return blank

    print(f"    reading application: {pdf.name}")
    lines = [ln.strip() for ln in text.splitlines() if ln.strip()]

    def _after(label_variants, section_after=None):
        """Find the value following a 'Label value' line. If section_after is
        given, only consider lines appearing after that section header."""
        start = 0
        if section_after:
            for i, ln in enumerate(lines):
                if section_after.lower() in ln.lower():
                    start = i
                    break
        for ln in lines[start:]:
            low = ln.lower()
            for lab in label_variants:
                if low.startswith(lab.lower()):
                    val = ln[len(lab):].strip(" :\t")
                    if val:
                        return val
        return ""

    # Prefer the Guarantor block; fall back to Borrower.
    name = (_after(["Full Name"], section_after="Primary Guarantor")
            or _after(["Guarantor Name"])
            or _after(["Full Name"], section_after="Primary Borrower")
            or _after(["Full Name"]))
    phone = (_after(["Phone"], section_after="Primary Guarantor")
             or _after(["Phone"], section_after="Primary Borrower")
             or _after(["Phone"]))
    email = (_after(["Email"], section_after="Primary Guarantor")
             or _after(["Email"], section_after="Primary Borrower")
             or _after(["Email"]))

    # Tidy email (PDFs sometimes drop the dot in .com).
    email = email.replace(" ", "")
    print(f"    application contact → name={name!r} phone={phone!r} email={email!r}")
    return {"name": name, "phone": phone, "email": email}


def _norm_name(s: str) -> str:
    return re.sub(r"\s+", " ", (s or "").strip().lower())


def _norm_phone(s: str) -> str:
    d = re.sub(r"\D", "", s or "")
    return d[-10:] if len(d) >= 10 else d


def reconcile_contact(casa: dict, app: dict) -> tuple[dict, list[str]]:
    """Merge the Casa contact and the Application contact. Returns
    (chosen_contact, conflicts) where conflicts is a list of field names that
    genuinely differ (ignoring case/formatting). The Application is preferred
    as the authoritative source when both agree or when Casa is blank.

    chosen fields fall back across sources so we always send the best value.
    """
    casa = casa or {}
    app = app or {}
    conflicts = []

    def pick(field, norm):
        cv, av = (casa.get(field) or "").strip(), (app.get(field) or "").strip()
        if cv and av:
            if norm(cv) != norm(av):
                conflicts.append(field)
            return av  # prefer application when both present
        return av or cv

    chosen = {
        "role":  casa.get("role", "") or "application",
        "name":  pick("name", _norm_name),
        "phone": pick("phone", _norm_phone),
        "email": pick("email", lambda x: x.strip().lower()),
    }
    return chosen, conflicts


def ask_which_value(field: str, casa_val: str, app_val: str) -> str:
    """Pop up a small window asking which value is correct for a conflicting
    contact field. Falls back to a console prompt if no GUI is available.
    Returns the chosen value."""
    title = f"Contact mismatch: {field}"
    msg = (f"The {field} differs between sources:\n\n"
           f"  [1] Casa (Baseline):   {casa_val or '(blank)'}\n"
           f"  [2] Application PDF:    {app_val or '(blank)'}\n\n"
           f"Which is correct?")
    try:
        import tkinter as tk
        from tkinter import ttk
        root = tk.Tk()
        root.title(title)
        root.attributes("-topmost", True)
        choice = {"val": app_val}  # default to application

        tk.Label(root, text=msg, justify="left",
                 font=("Segoe UI", 10)).pack(padx=16, pady=12)
        btns = tk.Frame(root); btns.pack(pady=(0, 14))

        def choose(v):
            choice["val"] = v
            root.destroy()

        ttk.Button(btns, text=f"Casa: {casa_val or '(blank)'}",
                   command=lambda: choose(casa_val)).grid(row=0, column=0, padx=6)
        ttk.Button(btns, text=f"Application: {app_val or '(blank)'}",
                   command=lambda: choose(app_val)).grid(row=0, column=1, padx=6)
        root.update_idletasks()
        # Center on screen.
        w, h = root.winfo_width(), root.winfo_height()
        x = (root.winfo_screenwidth() - w) // 2
        y = (root.winfo_screenheight() - h) // 3
        root.geometry(f"+{x}+{y}")
        root.mainloop()
        return choice["val"]
    except Exception:
        # No GUI — console fallback.
        print(f"\n!! {title}")
        print(f"   1) Casa:        {casa_val or '(blank)'}")
        print(f"   2) Application: {app_val or '(blank)'}")
        try:
            sel = input("   Which is correct? [1=Casa / 2=Application, "
                        "default 2]: ").strip()
        except Exception:
            sel = ""
        return casa_val if sel == "1" else app_val


def find_rehab_budget_xlsx(property_folder: Path) -> Path | None:
    """Return the first .xlsx/.xlsm under '<property>/Rehab Budget/'."""
    if property_folder is None:
        return None
    rb = Path(_long(property_folder)) / "Rehab Budget"
    if not rb.is_dir():
        return None
    candidates = []
    try:
        for f in rb.iterdir():
            if f.is_file() and f.suffix.lower() in (".xlsx", ".xlsm"):
                # Skip temp-lock files Excel leaves behind ('~$Foo.xlsx').
                if f.name.startswith("~$"):
                    continue
                candidates.append(f)
    except Exception as e:
        print(f"    (rehab-budget scan warning: {e})")
        return None
    if not candidates:
        return None
    # If there's more than one, take the most-recently-modified.
    candidates.sort(key=lambda p: p.stat().st_mtime, reverse=True)
    return candidates[0]


def build_draw_items_csv(xlsx_path: Path, dest_csv: Path) -> Path:
    """Convert a rehab-budget xlsx into ProxyPics's draw_items CSV.

    Mapping (xlsx → CSV columns Name,Description,Budget):
      xlsx col A (line-item name)  →  CSV 'Name' and 'Description'
      xlsx col C (AMOUNT)          →  CSV 'Budget' (0 if blank/non-numeric)

    Reads from the header row 'Description' downward, stops at 'TOTAL'.
    Returns dest_csv on success; raises on read failure."""
    from openpyxl import load_workbook
    wb = load_workbook(_long(xlsx_path), data_only=True)
    ws = wb.active

    # Find the header row: first row whose col A == 'Description'.
    header_row = None
    for r in range(1, min(ws.max_row, 30) + 1):
        a = ws.cell(r, 1).value
        if isinstance(a, str) and a.strip().lower() == "description":
            header_row = r
            break
    if header_row is None:
        header_row = 11  # template default

    # Collect line items, MERGING duplicates (same name → sum the budgets).
    # Preserve first-seen order. Each CSV row is (Name, Budget) — Name is used
    # for both the Name and Description columns, but written ONCE per item
    # (no "Dumpsters, Dumpsters" duplication).
    order: list[str] = []
    merged: dict[str, int] = {}
    for r in range(header_row + 1, ws.max_row + 1):
        name = ws.cell(r, 1).value
        if name is None:
            continue
        name = str(name).strip()
        if not name:
            continue
        if name.lower() in ("total", "totals"):
            break
        # Skip footer rows the template includes after the line items.
        if name.lower().startswith(("% complete", "cumulative", "remaining ",
                                    "% drawn", "borrower certif", "i certify")):
            break
        amount = ws.cell(r, 3).value
        try:
            budget = int(round(float(amount))) if amount not in (None, "") else 0
        except (TypeError, ValueError):
            budget = 0
        if budget <= 0:
            continue
        key = name.lower()
        if key in merged:
            merged[key] += budget                      # duplicate → merge sum
            print(f"    (merged duplicate line item '{name}': "
                  f"total now {merged[key]})")
        else:
            merged[key] = budget
            order.append(name)                          # keep original casing

    Path(dest_csv).parent.mkdir(parents=True, exist_ok=True)
    with open(_long(dest_csv), "w", newline="", encoding="utf-8") as f:
        w = _csv.writer(f)
        w.writerow(["Name", "Description", "Budget"])
        for name in order:
            w.writerow([name, "", merged[name.lower()]])
    return Path(dest_csv)


def resolve_csv_for_address(address: str, clients_root: str | None,
                            explicit_csv: str | None) -> str | None:
    """High-level entry: pick the CSV path to upload to ProxyPics.

    Priority:
      1. --csv (user-supplied) wins. If the file exists, return it.
      2. Otherwise, look up the property folder under `clients_root`,
         find the Rehab Budget xlsx, build a temp CSV from it.
      3. Return None if neither path yields a CSV (caller decides).
    """
    if explicit_csv and Path(_long(explicit_csv)).exists():
        print(f">>> Using user-provided CSV: {explicit_csv}")
        return explicit_csv
    elif explicit_csv:
        print(f"    (warning: --csv {explicit_csv!r} doesn't exist, "
              "falling back to auto-build)")

    root = clients_root or os.environ.get("DRAW_CLIENTS_ROOT") or DEFAULT_CLIENTS_ROOT
    if not root:
        print(">>> No clients-root configured — skipping CSV auto-build.")
        return None
    root_p = Path(_long(root))
    if not root_p.exists():
        print(f">>> Clients root {root!r} doesn't exist — skipping CSV auto-build.")
        return None

    print(f">>> Looking for rehab budget under {root}")
    prop = find_property_folder(root_p, address)
    if prop is None:
        print(f"    no property folder matched {address!r} — no CSV will be uploaded.")
        return None
    print(f"    matched property folder: {prop.name}")
    xlsx = find_rehab_budget_xlsx(prop)
    if xlsx is None:
        print(f"    no .xlsx in '{prop.name}\\Rehab Budget' — no CSV will be uploaded.")
        return None
    print(f"    found rehab budget: {xlsx.name}")
    try:
        out_csv = Path(tempfile.gettempdir()) / f"draw_items_{int(time.time())}.csv"
        build_draw_items_csv(xlsx, out_csv)
        print(f"    built CSV: {out_csv}")
        return str(out_csv)
    except Exception as e:
        print(f"    !! failed to build CSV from {xlsx.name}: {e}")
        return None


# ===========================================================================
# Generic helpers
# ===========================================================================
def digits_only(s: str) -> str:
    """Strip everything but digits; drop leading 1 if 11 digits long."""
    d = re.sub(r"\D", "", s or "")
    if len(d) == 11 and d.startswith("1"):
        d = d[1:]
    return d


class PageClosed(Exception):
    """Raised when the Playwright page/tab has been closed mid-operation, so
    callers can recover by opening a fresh page from the context instead of
    crashing the whole run."""


def wait_for_ready(page: Page, target_url: str, ready_selector: str,
                   label: str) -> None:
    """Navigate to target_url; poll until ready_selector is visible.
    Handles login redirects by periodically re-navigating to target_url.
    Raises PageClosed if the page/tab is dead so callers can recover."""
    if not _safe_url(page):
        raise PageClosed(f"[{label}] page already closed before navigation")
    print(f">>> [{label}] Navigating to {target_url}")
    try:
        page.goto(target_url, wait_until="domcontentloaded", timeout=30_000)
    except Exception as e:
        if not _safe_url(page):
            raise PageClosed(f"[{label}] page closed during goto: {e}")
        print(f"    (initial goto warning: {e})")

    deadline       = time.time() + LOGIN_WAIT_S
    next_renav_at  = 0.0
    next_status_at = 0.0
    target_path    = target_url.split("//", 1)[1].split("/", 1)[1]

    while time.time() < deadline:
        if not _safe_url(page):
            raise PageClosed(f"[{label}] page closed while waiting for ready")
        try:
            if page.locator(ready_selector).count() > 0:
                page.locator(ready_selector).first.wait_for(
                    state="visible", timeout=5000
                )
                print(f">>> [{label}] Page ready.")
                return
        except PWTimeout:
            pass
        except Exception as e:
            print(f"    (poll warning: {e})")

        now = time.time()
        if now >= next_status_at:
            try:
                print(f"    [{label}] waiting…  url={page.url}")
            except Exception:
                print(f"    [{label}] waiting…")
            next_status_at = now + 5

        try:
            if target_path not in page.url and now >= next_renav_at:
                next_renav_at = now + 8
                try:
                    page.goto(target_url, wait_until="domcontentloaded",
                              timeout=15_000)
                except Exception:
                    pass
        except Exception:
            raise PageClosed(f"[{label}] page closed during renav check")

        try:
            page.wait_for_timeout(1000)
        except Exception:
            raise PageClosed(f"[{label}] page closed during wait")

    shot = Path(__file__).parent / f"timeout_{label}.png"
    try:
        page.screenshot(path=str(shot), full_page=True)
        print(f">>> Saved screenshot to {shot}")
    except Exception:
        pass
    raise TimeoutError(
        f"[{label}] page not ready within {LOGIN_WAIT_S}s. "
        f"Last URL: {page.url}."
    )


# ===========================================================================
# Local DB fast path
# ===========================================================================
# Casa Finance address search is brittle: minor formatting differences, closed
# loans hiding on the wrong tab, and React detail-page race conditions all make
# it fail intermittently. The Draw Manager app already maintains a synced
# sqlite DB with column A (CF#) → column C (address) mappings for every loan
# we know about. We resolve address → CF# locally first, then drive the Casa
# search by CF# (which is unique and deterministic) instead of address text.

_ADDR_PUNCT = re.compile(r"[,\.]")
_ADDR_WS    = re.compile(r"\s+")

def _norm_addr(s: str) -> str:
    """Loose normalization so '2307 Chestnut St, Houston TX' matches
    '2307 chestnut st' — lowercase, strip punctuation, collapse whitespace."""
    if not s:
        return ""
    s = s.lower().strip()
    s = _ADDR_PUNCT.sub(" ", s)
    s = _ADDR_WS.sub(" ", s)
    return s.strip()


def _db_lookup_out_status(db_path: str, address: str) -> dict | None:
    """Read-only lookup against draws.db: for the loan matching `address`,
    return {cf, draw_number, out, loan_address} for the most recent draw.

    `out` is True if the OUT column (Google Sheet → draws.inspection_ordered)
    was checked on that draw. Returns None if the address isn't in the DB.

    This is what the Draw Manager GUI's app.py also consults to decide
    repeat-vs-new — we duplicate it here so running this script standalone
    can SHOW the same sheet-derived decision in its output."""
    if not db_path or not Path(db_path).exists():
        return None
    needle = _norm_addr(address)
    if not needle:
        return None
    try:
        conn = sqlite3.connect(f"file:{db_path}?mode=ro", uri=True)
        conn.row_factory = sqlite3.Row
        try:
            loans = conn.execute(
                "SELECT id, cf_number, address FROM loans "
                "WHERE address IS NOT NULL AND address != ''"
            ).fetchall()
            cand = next((r for r in loans if _norm_addr(r["address"]) == needle), None)
            if not cand:
                cand = next((r for r in loans
                             if _norm_addr(r["address"]).startswith(needle + " ")), None)
            if not cand:
                cand = next((r for r in loans
                             if needle.startswith(_norm_addr(r["address"]) + " ")), None)
            if not cand:
                return None
            d = conn.execute(
                "SELECT draw_number, inspection_ordered "
                "FROM draws WHERE loan_id=? "
                "ORDER BY draw_number DESC LIMIT 1",
                (cand["id"],),
            ).fetchone()
        finally:
            conn.close()
    except Exception as e:
        print(f"    (DB lookup warning: {e})")
        return None
    return {
        "cf":           cand["cf_number"],
        "loan_address": cand["address"],
        "draw_number":  d["draw_number"]      if d else None,
        "out":          bool(d["inspection_ordered"]) if d else False,
    }


def _db_lookup_cf(db_path: str, address: str) -> str | None:
    """Look up a CF# in the local Draw Manager DB by fuzzy address match.
    Returns the CF# (e.g. 'CF008') or None. Read-only, never raises."""
    if not db_path or not Path(db_path).exists():
        return None
    needle = _norm_addr(address)
    if not needle:
        return None
    try:
        # `mode=ro` so we never lock the file the app might also be reading.
        conn = sqlite3.connect(f"file:{db_path}?mode=ro", uri=True)
        conn.row_factory = sqlite3.Row
        try:
            rows = conn.execute(
                "SELECT cf_number, address FROM loans "
                "WHERE address IS NOT NULL AND address != ''"
            ).fetchall()
        finally:
            conn.close()
    except Exception as e:
        print(f"    (DB lookup warning: {e})")
        return None

    # Three-tier match, most-strict first. Returning early on first hit means
    # closer matches win even when the DB has multiple loans on the same street.
    normalized = [(r["cf_number"], _norm_addr(r["address"])) for r in rows]

    # 1. Exact normalized equality.
    for cf, a in normalized:
        if a == needle:
            return cf
    # 2. Stored address starts with what the user typed
    #    (user gave street; DB has 'street, city, st zip').
    for cf, a in normalized:
        if a.startswith(needle + " "):
            return cf
    # 3. What the user typed starts with the stored address
    #    (user pasted in extras; DB has just the street).
    for cf, a in normalized:
        if a and needle.startswith(a + " "):
            return cf
    return None


# ===========================================================================
# Casa Finance lookup
# ===========================================================================
def casa_lookup(page: Page, address: str, cf_hint: str | None = None,
                match_address: str = "") -> dict:
    """Find the loan in Casa, return {cf_number, contacts, address, _page}.
    `address` is what we type into Casa's search box (often just "1300").
    `match_address` is the FULL address used to pick the RIGHT row.

    Resilient: if a section closes the page, recover a fresh page from the
    context and try the next section. Never raises. The (possibly new) live
    page is returned under the '_page' key so the caller keeps a valid handle.
    """
    for section_url, label in [(CASA_SERV_URL,   "casa-servicing"),
                               (CASA_LOANS_URL,  "casa-loans")]:
        # Ensure we have a live page before each section.
        if not _safe_url(page):
            try:
                page = page.context.new_page()
                print("    (recovered a fresh tab for Casa)")
            except Exception as e:
                print(f"    !! Could not recover a page: {e}")
                break
        try:
            result = _try_casa_section(page, section_url, label, address,
                                       cf_hint, match_address=match_address or address)
        except PageClosed as e:
            print(f"    (section {label} closed the page: {e})")
            result = None
        if result is not None:
            result["_page"] = page
            return result
    print(f"    !! Casa lookup failed for {address!r}. Proceeding without "
          f"Casa-derived info (CF# = {cf_hint or '(empty)'!r}, no contacts).")
    return {"cf_number": cf_hint or "", "contacts": [], "address": "",
            "_page": page if _safe_url(page) else None}


def _try_casa_section(page: Page, section_url: str, label: str,
                      address: str, cf_hint: str | None = None,
                      match_address: str = "") -> dict | None:
    """Search Casa by address, click the matching row, extract CF# + contacts.
    If cf_hint is provided (from the local DB), it OVERRIDES whatever the
    page extraction would produce — Casa's search box only accepts addresses,
    so we still search by address, but we can skip the brittle CF# extraction
    step when we already know the answer."""
    wait_for_ready(page, section_url, CASA_READY, label)

    queries = []
    # When the local DB resolved a CF#, search by it FIRST. CF#s are unique and
    # the loans-table filter matches them deterministically — no risk of cross-
    # entity collision like address text has (where "1359 Wright St" can match
    # a borrower's address field in the global search).
    if cf_hint:
        queries.append(cf_hint)
    queries.append(address)
    street_only = address.split(",", 1)[0].strip()
    if street_only and street_only != address:
        queries.append(street_only)

    for query in queries:
        # Make sure we're back on the section page (we may have navigated to a
        # row's detail on a previous attempt).
        if section_url.rsplit("/", 1)[-1] not in page.url:
            print(f"    Returning to {section_url}")
            try:
                page.goto(section_url, wait_until="domcontentloaded")
                page.locator(CASA_READY).first.wait_for(
                    state="visible", timeout=15_000
                )
            except Exception:
                pass

        print(f">>> [{label}] Searching: {query!r}")
        if not _casa_search_and_click(page, query,
                                      match_address=match_address or address):
            continue

        # If the loan opened in a new tab, operate on that tab from here on.
        if _SWAP.get("page") is not None:
            page = _SWAP["page"]
            _SWAP["page"] = None

        # We're now on the loan landing page (/admin/loans/<id>/...). The CF#
        # and the property address are visible HERE, on the first page — so
        # extract them before navigating anywhere else.
        loan_base = re.match(r"(https://[^/]+/admin/loans/\d+)", page.url)
        loan_base_url = loan_base.group(1) if loan_base else None

        # CF#: prefer the DB-known value (cf_hint); else read it off this page.
        if cf_hint:
            cf = cf_hint
            print(f">>> [{label}] Using DB-known CF# (skipped page extraction): {cf}")
        else:
            cf = _extract_cf_number(page)
            print(f">>> [{label}] CF#: {cf or '(blank)'}")

        # Full property address off the loan page (so ProxyPics gets the real
        # address, not a Google-autocomplete guess from a partial query).
        full_address = _extract_loan_address(page)
        if full_address:
            print(f">>> [{label}] Loan address: {full_address}")

        # Contacts live on the /contacts sub-page.
        contacts: list[dict] = []
        if loan_base_url:
            try:
                page.goto(loan_base_url + "/contacts",
                          wait_until="domcontentloaded", timeout=20_000)
                page.wait_for_timeout(800)
                contacts = _extract_contacts(page)
            except Exception as e:
                print(f"    (couldn't open contacts: {e})")
        else:
            contacts = _extract_contacts(page)
        return {"cf_number": cf, "contacts": contacts,
                "address": full_address}
    return None


def _extract_loan_address(page: Page) -> str:
    """Read the property address off the Casa loan page. Tries the same
    _summary_/_key_/_value_ structure used for the CF#, looking for a
    'Property' / 'Address' labelled value, then a breadcrumb fallback."""
    # Strategy A: summary block labelled Property / Address.
    for label in ("Property Address", "Property", "Address", "Subject Property"):
        try:
            summary = page.locator(
                f'div[class*="_summary_"]:has-text("{label}")'
            ).first
            if summary.count() > 0:
                val = summary.locator('div[class*="_value_"]').first
                txt = (val.text_content() or "").strip()
                # Sanity: looks like a street address (starts with a number).
                if txt and re.match(r"^\s*\d", txt):
                    return re.sub(r"\s+", " ", txt.replace("\xa0", " ")).strip()
        except Exception:
            continue
    # Strategy B: breadcrumb (Casa shows the address there on loan pages).
    try:
        bc = page.locator('[data-testid="breadcrumb-1"]').first
        if bc.count() > 0:
            txt = (bc.text_content() or "").strip()
            if txt and re.match(r"^\s*\d", txt):
                return re.sub(r"\s+", " ", txt.replace("\xa0", " ")).strip()
    except Exception:
        pass
    return ""


def _casa_search_and_click(page: Page, query: str,
                           match_address: str = "") -> bool:
    """Search Casa, find the row matching the FULL address, click it, and land
    on the loan page. Returns True if a loan page rendered.

    Row selection: among the results, pick the one whose visible text best
    matches `match_address` (e.g. '1300 NW 42nd St, Miami, FL 33142') rather
    than the bare query ('1300'), so multi-result searches click the RIGHT
    property. We click the row (Casa routes the SPA from there); whatever it
    navigates to, we resolve down to /admin/loans/<id>.
    """
    inputs = page.locator('input[placeholder="Search"]')
    search = inputs.last if inputs.count() > 1 else inputs.first
    search.click()
    search.fill("")
    search.fill(query)
    page.wait_for_timeout(1500)

    rows = page.locator('tr[data-row-id]')
    n = rows.count()
    print(f"    {n} row(s) returned.")
    if n == 0:
        return False

    # Score rows against the full address; pick the best match.
    full = _norm_addr(match_address or query)
    street = full.split(",", 1)[0].strip()              # "1300 nw 42nd st"
    street_num = (street.split() or [""])[0]            # "1300"
    best_idx, best_score = 0, -1
    for i in range(n):
        try:
            txt = _norm_addr(rows.nth(i).text_content() or "")
        except Exception:
            continue
        score = 0
        if street and street in txt:
            score = 3                                   # full street line
        elif street_num and street_num in txt:
            # number matches; boost if a distinctive street word also appears
            score = 1
            words = [w for w in street.split()[1:] if len(w) > 2]
            if any(w in txt for w in words):
                score = 2
        if score > best_score:
            best_idx, best_score = i, score
    print(f"    Best-matching row: index {best_idx} (score {best_score}) "
          f"for {street!r}")

    # Read the loan id from the matched row. Casa puts the LOAN ID in the row's
    # `id` attribute (e.g. <tr id="15999778" data-row-id="0">). Note:
    # data-row-id is a useless table index ("0"), and there's no /admin/loans/
    # href in the markup — the real id is the `id` attribute. Read that and
    # navigate directly; do NOT click the row (its Borrower cell is an <a> that
    # routes to the borrower page).
    target_row = rows.nth(best_idx)
    loan_id = ""
    try:
        rid_attr = (target_row.get_attribute("id") or "").strip()
        if rid_attr.isdigit() and len(rid_attr) >= 4:
            loan_id = rid_attr
    except Exception as e:
        print(f"    (row id read warning: {e})")
    if not loan_id:
        # Fallbacks: an explicit loans href, then /loans/<id> anywhere in row.
        try:
            row_html = target_row.evaluate("el => el.outerHTML")
            m = (re.search(r"/admin/loans/(\d{4,})", row_html)
                 or re.search(r"/loans/(\d{4,})", row_html))
            if m:
                loan_id = m.group(1)
        except Exception as e:
            print(f"    (row HTML read warning: {e})")

    if loan_id:
        loan_url = f"https://myportal.casa.finance/admin/loans/{loan_id}"
        print(f"    Loan id {loan_id} (from row id attr) → {loan_url}")
        try:
            page.goto(loan_url, wait_until="domcontentloaded", timeout=30_000)
        except Exception as e:
            print(f"    (goto loan warning: {e})")
        if _wait_loan_ready(page):
            print(f"    On loan page: {_safe_url(page)}")
            return True
        print(f"    Loan id {loan_id} didn't resolve to a loan page.")
        return False

    # Could not find the loan id in the row markup directly. Click the row
    # (Casa may route us to the borrower page) and recover the loan id from
    # wherever we land by scanning the DOM for /admin/loans/<id>.
    print("    No loan id in row markup; clicking row and recovering loan id "
          "from the destination page.")
    try:
        target_row.click(timeout=5000)
    except Exception as e:
        print(f"    (row click warning: {e})")
        return False
    page.wait_for_timeout(1500)

    if _wait_loan_ready(page, timeout_ms=6000):
        print(f"    On loan page directly: {_safe_url(page)}")
        return True

    cur = _safe_url(page)
    print(f"    Now on: {cur}")

    # On the borrower page the loan is a clickable SPA row (no href, no loan
    # id in the HTML). Click candidate rows and watch for the URL to become
    # /admin/loans/<id>. Return to the borrower page between tries.
    if cur and "/admin/borrowers/" in cur:
        borrower_url = cur
        # Casa's clickable rows use a class like _clickable_… (seen in the
        # contacts table). Try those first, then any table row.
        for sel in ['tr[class*="_clickable_"]', '[class*="_clickable_"]',
                    'tbody tr', 'tr[data-row-id]']:
            try:
                cands = page.locator(sel)
                cnt = cands.count()
            except Exception:
                continue
            if cnt == 0:
                continue
            print(f"    Trying {cnt} candidate row(s) via {sel!r}")
            for i in range(min(cnt, 10)):
                try:
                    cands.nth(i).click(timeout=4000)
                except Exception:
                    continue
                page.wait_for_timeout(1200)
                if _wait_loan_ready(page, timeout_ms=5000):
                    print(f"    Reached loan via {sel}[{i}]: {_safe_url(page)}")
                    return True
                # Not the loan — go back to the borrower page for the next try.
                nowu = _safe_url(page)
                if "/admin/loans/" in nowu:
                    return True
                if nowu != borrower_url:
                    try:
                        page.goto(borrower_url, wait_until="domcontentloaded",
                                  timeout=20_000)
                        page.wait_for_timeout(700)
                    except Exception:
                        pass
            # If we get here, this selector's rows didn't lead to a loan.

    loan_id = _find_loan_id_on_page(page)
    if loan_id:
        loan_url = f"https://myportal.casa.finance/admin/loans/{loan_id}"
        print(f"    Recovered loan id {loan_id} → {loan_url}")
        try:
            page.goto(loan_url, wait_until="domcontentloaded", timeout=30_000)
        except Exception as e:
            print(f"    (goto recovered loan warning: {e})")
        if _wait_loan_ready(page):
            print(f"    On loan page: {_safe_url(page)}")
            return True

    # Diagnostics: print what loan-ish ids ARE present so we can pin it.
    try:
        html = page.content()
        found = sorted(set(re.findall(r"/loans/(\d{4,})", html)))
        print(f"    [diag] loan ids seen in page HTML: {found or 'none'}")
        rids = page.eval_on_selector_all(
            '[data-row-id]', "els => els.map(e => e.getAttribute('data-row-id'))")
        print(f"    [diag] data-row-id values: {sorted(set(rids or []))[:20]}")
    except Exception as e:
        print(f"    [diag] page scan failed: {e}")
    print(f"    Could not reach a loan page from {query!r}.")
    return False


def _find_loan_id_on_page(page) -> str:
    """Scan a (borrower) page's DOM for a loan id. Tries, in order:
      1. Any anchor href like /admin/loans/<id>.
      2. Any element attribute containing /admin/loans/<id>.
      3. Clickable loan rows whose data-row-id is a long numeric id.
    Returns the id string, or ''."""
    # 1 & 2: search the full HTML for the loans URL pattern.
    try:
        html = page.content()
        m = re.search(r"/admin/loans/(\d{5,})", html)
        if m:
            return m.group(1)
    except Exception:
        pass
    # 3: rows carry the loan id in the `id` attribute (data-row-id is a
    #    table index). Read row ids that look like loan ids.
    try:
        ids = page.eval_on_selector_all(
            'tr[id], [data-row-id]',
            "els => els.map(e => e.getAttribute('id') || e.getAttribute('data-row-id'))",
        )
        for rid in ids or []:
            if rid and str(rid).isdigit() and len(str(rid)) >= 5:
                return str(rid)
    except Exception:
        pass
    return ""


# When a click opens the loan in a new tab, stash it here so the caller can
# switch the page it operates on.
_SWAP: dict = {"page": None}


def _safe_url(page) -> str:
    """Return page.url or '' if the page is closed/dead."""
    try:
        return page.url
    except Exception:
        return ""


def _wait_loan_ready(page: Page, timeout_ms: int = 15_000) -> bool:
    """Wait until we're on a /admin/loans/<id> page. Polls page.url in Python
    (robust to SPA context swaps) and handles the case where a click opened
    the loan in a NEW tab. Never raises on a closed page."""
    deadline = time.time() + timeout_ms / 1000.0
    while time.time() < deadline:
        # A row click may have spawned a new tab — check every context page.
        try:
            ctx = page.context
            for p in ctx.pages:
                try:
                    if "/admin/loans/" in p.url:
                        if p is not page:
                            print("    (loan opened in a new tab — switching)")
                            try:
                                p.bring_to_front()
                            except Exception:
                                pass
                        p.wait_for_timeout(900)
                        return True
                except Exception:
                    continue
        except Exception:
            pass
        try:
            if "/admin/loans/" in page.url:
                page.wait_for_timeout(900)
                return True
        except Exception:
            # Page object is dead/closed — give up cleanly.
            return False
        try:
            page.wait_for_timeout(400)
        except Exception:
            return False
    return False


CF_REGEX = re.compile(r"\bCF\s*\d+\b", re.IGNORECASE)


def _normalize_cf(text: str) -> str | None:
    """Pull a CF### out of `text`, normalize to e.g. 'CF630'."""
    if not text:
        return None
    cleaned = text.replace("\xa0", " ")
    m = CF_REGEX.search(cleaned)
    if not m:
        return None
    return re.sub(r"\s+", "", m.group(0)).upper()


def _extract_cf_number(page: Page) -> str:
    """Find the loan's CF# on the detail page. Tries several strategies."""
    try:
        page.wait_for_load_state("networkidle", timeout=8000)
    except PWTimeout:
        pass

    # Strategy 0 (most precise): the real DOM is
    #   <div class="_summary_…"><div class="_key_…">…Casa Finance File Number…</div>
    #                            <div class="_value_…">CF956</div></div>
    # Find the _summary_ block that contains the label text, then read its
    # _value_ child. This is the structure observed on the live page.
    try:
        summary = page.locator(
            'div[class*="_summary_"]:has-text("Casa Finance File Number")'
        ).first
        summary.wait_for(state="attached", timeout=5000)
        val = summary.locator('div[class*="_value_"]').first
        cf = _normalize_cf(val.text_content() or "")
        if cf:
            return cf
    except Exception:
        pass

    # Strategy 1: structural — the page uses CSS-modules with the pattern
    #   <div class="_summary_xxx"><div class="_key_xxx">…label…</div>
    #                            <div class="_value_xxx">CF772</div></div>
    # The hashes change but the prefixes don't.
    xpath_candidates = [
        # Element containing label → climb to summary container → grab value child
        '//*[contains(normalize-space(.), "Casa Finance File Number")]'
        '/ancestor::*[contains(@class, "_summary_")][1]'
        '//*[contains(@class, "_value_")]',
        # Key div with label → next sibling div (the value)
        '//div[contains(@class, "_key_") and .//*[contains(text(), "Casa Finance File Number")]]'
        '/following-sibling::div[1]',
        # Same but label is direct text, not nested
        '//div[contains(@class, "_key_") and contains(normalize-space(.), "Casa Finance File Number")]'
        '/following-sibling::div[1]',
        # Original fallbacks (in case some pages still use contenteditable)
        '//*[contains(normalize-space(.), "Casa Finance File Number")]'
        '/following::*[@contenteditable][1]',
        '//*[contains(normalize-space(.), "File Number")]'
        '/following::*[contains(@class, "_value_")][1]',
    ]
    for xp in xpath_candidates:
        try:
            loc = page.locator(f"xpath={xp}").first
            loc.wait_for(state="attached", timeout=3000)
            cf = _normalize_cf(loc.text_content() or "")
            if cf:
                return cf
        except Exception:
            continue

    # Strategy 2: brute-force — scan every `_value_*` div on the page.
    try:
        values = page.locator('[class*="_value_"]')
        for i in range(min(values.count(), 60)):
            cf = _normalize_cf(values.nth(i).text_content() or "")
            if cf:
                return cf
    except Exception:
        pass

    # Strategy 3: full body text regex sweep.
    try:
        body_text = page.locator("body").inner_text(timeout=5000)
        cf = _normalize_cf(body_text)
        if cf:
            return cf
    except Exception:
        pass

    print("    !! No CF# found on the loan detail page (will leave loan "
          "number blank for manual entry).")
    return ""


def _extract_contacts(page: Page) -> list[dict]:
    """Read Borrower and Guarantor rows from the loan's Contacts sub-page.

    The caller has already navigated to /admin/loans/<id>/contacts, so the
    table is present directly. If a Contacts tab happens to be present and
    the table isn't yet, click it as a fallback. Returns [] if no rows.
    """
    # If the contacts table rows aren't visible yet but a Contacts tab is,
    # click it (covers the case where we're on the loan landing page).
    if page.locator('tr[data-row-id]').count() == 0:
        tab = page.locator('[data-testid="Contacts"]')
        if tab.count() > 0:
            try:
                tab.first.click()
                page.wait_for_timeout(1200)
            except Exception:
                pass
    # Give the rows a moment to render.
    try:
        page.locator('tr[data-row-id]').first.wait_for(state="visible",
                                                        timeout=8000)
    except PWTimeout:
        print("    No contact rows rendered — returning empty contacts list.")
        return []

    contacts: list[dict] = []
    for role in ["Borrower", "Guarantor"]:
        # A contact row whose Role cell contains the role text.
        row = page.locator(
            f'tr[data-row-id]:has(td[data-label="Role"]:has-text("{role}"))'
        ).first
        if row.count() == 0:
            print(f"    No {role} row found.")
            continue
        name  = _cell_text(row, "Name")
        email = _cell_text(row, "Email")
        phone = _cell_text(row, "Phone")
        print(f"    {role}: name={name!r} email={email!r} phone={phone!r}")
        contacts.append({
            "role":  role,
            "name":  name,
            "email": "" if email in ("", "-") else email,
            "phone": "" if phone in ("", "-") else phone,
        })
    return contacts


def _cell_text(row: Locator, data_label: str) -> str:
    cell = row.locator(f'td[data-label="{data_label}"]')
    if cell.count() == 0:
        return ""
    return (cell.first.text_content() or "").strip()


# ===========================================================================
# ProxyPics form filling
# ===========================================================================
def fill_proxypics(page: Page, data: dict, submit: bool = False) -> None:
    wait_for_ready(page, PROXYPICS_URL, PROXYPICS_READY, "proxypics")

    # Each step is wrapped so one missing field/selector can't abort the whole
    # run — for a demo we want it to fill what it can and still reach "Next".
    def step(label, fn):
        print(f">>> {label}")
        try:
            fn()
        except Exception as e:
            print(f"    (skipped {label!r}: {type(e).__name__}: {e})")

    def _inspection_type():
        field = page.locator(".ant-select-auto-complete").first
        field.click()
        field.locator("input.ant-select-selection-search-input").fill(
            data["inspection_type_search"]
        )
        page.locator(".ant-select-item-option-content").first.wait_for(
            state="visible", timeout=10_000
        )
        page.locator(".ant-select-item-option-content").first.click()
    step("Filling: inspection type", _inspection_type)

    step("Filling: loan number",
         lambda: page.locator('input[id$="_loanNumber"]').fill(data["loan_number"]))

    # Borrower name: the top-level "Borrower Name" field on the form
    # (id ends in _borrowerName). We fill it with the contact name we
    # resolved from Casa (e.g. the Guarantor 'Angelica Stuart').
    if data.get("contact_name"):
        step("Filling: borrower name",
             lambda: page.locator('input[id$="_borrowerName"]').fill(
                 data["contact_name"]))

    def _address():
        addr = page.locator('input[id$="_address"]')
        addr.click()
        addr.fill("")
        # Type the full address with real keystrokes so Google Places fires.
        addr.press_sequentially(data["address"], delay=40)
        page.wait_for_timeout(600)  # let the Places debounce request go out
        street_num = (data["address"].split() or [""])[0]
        try:
            page.wait_for_selector(".pac-item", timeout=8000)
        except PWTimeout:
            print("    (no Places dropdown appeared; leaving typed value)")
            return
        # Pick the suggestion matching our street number; else the first.
        items = page.locator(".pac-item")
        target = None
        for i in range(min(items.count(), 6)):
            txt = (items.nth(i).text_content() or "")
            if street_num and street_num in txt:
                target = items.nth(i)
                break
        if target is None and items.count() > 0:
            target = items.first
        # Click via the bounding box (a real mouse click) — .pac-item is a
        # Google overlay that sometimes ignores synthetic clicks. Fall back
        # to keyboard ArrowDown+Enter.
        committed = False
        if target is not None:
            try:
                box = target.bounding_box()
                if box:
                    page.mouse.click(box["x"] + box["width"] / 2,
                                     box["y"] + box["height"] / 2)
                    committed = True
            except Exception:
                pass
            if not committed:
                try:
                    target.click(timeout=3000)
                    committed = True
                except Exception:
                    pass
        if not committed:
            page.keyboard.press("ArrowDown")
            page.keyboard.press("Enter")
        page.wait_for_timeout(500)
        # Verify the input now holds a full (comma-containing) address.
        try:
            val = addr.input_value()
            if "," in val:
                print(f"    address committed: {val}")
            else:
                print(f"    address value after pick: {val!r}")
        except Exception:
            pass
    step("Filling: address", _address)

    # Contact: only add one if we actually have any contact data. With no
    # contacts on the loan we skip this entirely rather than creating a blank.
    has_contact = any((data.get("contact_name"), data.get("contact_phone"),
                       data.get("contact_email")))
    if has_contact:
        def _contact():
            page.get_by_role("button", name="Add contact").click()
            if data.get("contact_name"):
                page.get_by_placeholder("Contact Name").fill(data["contact_name"])
            if data.get("contact_phone"):
                phone_input = page.locator('input[type="tel"]').last
                phone_input.click()
                phone_input.press("End")
                phone_input.type(data["contact_phone"])
            if data.get("contact_email"):
                page.get_by_placeholder("Email address").fill(data["contact_email"])
        step("Filling: contact", _contact)
    else:
        print(">>> No contact info — skipping the Add contact step.")

    if data.get("csv_path"):
        def _csv():
            with page.expect_file_chooser() as fc_info:
                page.get_by_role("button", name="Import line items").click()
            fc_info.value.set_files(data["csv_path"])
        step("Uploading: line items CSV", _csv)
    else:
        print(">>> No line-items CSV provided — skipping import.")

    def _progress():
        wrapper = page.locator('input[id$="_drawProgressStep"]').locator(
            'xpath=ancestor::div[contains(@class,"ant-select")][1]'
        )
        wrapper.click()
        page.locator(f'.ant-select-item-option[title="{data["progress_step"]}"]').click()
    step("Selecting: progress step", _progress)

    step("Clicking: Next",
         lambda: page.get_by_role("button", name="Next").click())

    if submit:
        def _submit():
            # The Submit button on the review screen:
            #   <button type="submit" class="... ant-btn-primary"><span>Submit</span></button>
            btn = page.get_by_role("button", name="Submit", exact=True)
            try:
                btn.first.wait_for(state="visible", timeout=20_000)
            except Exception:
                btn = page.locator('button[type="submit"].ant-btn-primary'
                                   ':has(span:text-is("Submit"))')
                btn.first.wait_for(state="visible", timeout=10_000)
            btn.first.click()
        step("Clicking: Submit (LIVE)", _submit)
        print(">>> LIVE submit attempted.")
    else:
        print(">>> TEST MODE — form filled and Next clicked; stopping before Submit.")


# ===========================================================================
# Repeat-draw flow: search the existing photo-requests list, open the most
# recent match for this address, click "Request New Draw Inspection", then
# confirm. Used when an address already has a draw out.
# ===========================================================================
PROXYPICS_LIST_URL = "https://app.proxypics.com/photo-requests"


def _proxypics_search_attempt(page: Page, search, fragment: str,
                              timeout_ms: int) -> bool:
    """Type `fragment` into the ProxyPics address filter and wait for the
    table to either show photo-request rows or the empty state.
    Returns True if rows rendered, False if 'no data' appeared.

    Uses press_sequentially (real key events) instead of fill() because
    ant-design's debounced filter only re-runs on input/keydown events
    from genuine keystrokes; fill() sets the value directly and the
    filter doesn't fire."""
    # Clear by selecting all + deleting (cheaper than fill("") + still
    # triggers the React state to reset).
    search.click()
    page.keyboard.press("Control+A")
    page.keyboard.press("Delete")
    page.wait_for_timeout(150)
    search.press_sequentially(fragment, delay=40)
    # Let the debounce settle. ant's default is ~300ms; give it a beat.
    page.wait_for_timeout(900)
    try:
        page.wait_for_function(
            """() => {
                const has  = document.querySelector(
                    "tr.ant-table-row td a[href^='/photo-requests/'][href$='/show']");
                const none = document.querySelector("div.ant-empty");
                return Boolean(has || none);
            }""",
            timeout=timeout_ms,
        )
    except PWTimeout:
        return False
    rows = page.locator(
        "tr.ant-table-row td a[href^='/photo-requests/'][href$='/show']"
    )
    return rows.count() > 0


def find_existing_draw(page: Page, address: str, timeout_ms: int = 12_000) -> str | None:
    """Open the ProxyPics list, search by address, and return the FIRST row's
    detail URL (e.g. '/photo-requests/1259864/show') if a result appears, or
    None if no match.

    Cascades through increasingly forgiving search fragments because
    ProxyPics' address filter is finicky: a full street line like
    "1359 Wright St" can return 0 results while "1359" returns the same
    property's 4 historical draws. We try the most specific fragment first
    (low collision risk) and fall back to the bare street number."""
    print(f">>> Searching ProxyPics for existing draws at: {address!r}")
    page.goto(PROXYPICS_LIST_URL, wait_until="domcontentloaded", timeout=30_000)

    search = page.locator("#addressICont")
    try:
        search.wait_for(state="visible", timeout=timeout_ms)
    except PWTimeout:
        print("    !! search box never appeared (login expired?).")
        return None

    # Build a cascade of fragments from most specific → least specific.
    # Order matters: first hit wins, so put narrow ones up front to avoid
    # cross-property collisions on the bare number.
    a = (address or "").strip()
    street_line = a.split(",", 1)[0].strip()           # "1359 Wright St"
    parts = street_line.split()
    street_number = parts[0] if parts else ""          # "1359"
    street_num_word = " ".join(parts[:2]) if len(parts) >= 2 else street_number
    fragments = []
    for f in (street_line, street_num_word, street_number):
        if f and f not in fragments:
            fragments.append(f)

    for i, fragment in enumerate(fragments, 1):
        tag = f"[{i}/{len(fragments)}]"
        # ALWAYS verify that the matched row's visible address contains the
        # original street line. ProxyPics' filter is loose enough that even
        # the full street line can return unrelated rows (observed: typing
        # '999 test st' returned a 6714 S Faul St request). The verify costs
        # nothing on legitimate matches — '1359 Wright St' is trivially a
        # substring of '1359 Wright St, Daytona Beach, FL ...'.
        print(f"    {tag} trying fragment: {fragment!r}"
              + (f"  (verifying rows contain {street_line!r})"
                 if street_line else ""))
        if not _proxypics_search_attempt(page, search, fragment, timeout_ms):
            print(f"    {tag} no rows.")
            continue
        # Pull every result row's address link and pick the first that
        # passes verification. Rows are returned newest-first by ProxyPics.
        links = page.locator(
            "tr.ant-table-row td a[href^='/photo-requests/'][href$='/show']"
        )
        n = links.count()
        target_lower = (street_line or "").lower()
        chosen_href = None
        for k in range(n):
            link = links.nth(k)
            href = link.get_attribute("href") or ""
            text = (link.text_content() or "").strip()
            if target_lower and target_lower not in text.lower():
                # Cross-property collision — skip this row.
                print(f"        skip row {k+1}/{n}: {text!r} doesn't contain "
                      f"{street_line!r}")
                continue
            chosen_href = href
            print(f"    {tag} matched -> row {k+1}/{n}: {text!r}  ({href})")
            break
        if chosen_href:
            return chosen_href
        print(f"    {tag} {n} row(s) found but none matched the target address.")

    print("    no existing draw found for that address.")
    return None


def request_new_inspection_from_existing(page: Page, detail_href: str,
                                         submit: bool = False,
                                         timeout_ms: int = 15_000) -> None:
    """Open an existing photo-request detail page, click 'Request New Draw
    Inspection', and confirm (if submit=True). In test mode we stop at the
    confirmation modal without clicking OK so nothing is created."""
    url = "https://app.proxypics.com" + detail_href if detail_href.startswith("/") else detail_href
    print(f">>> Opening existing draw: {url}")
    page.goto(url, wait_until="domcontentloaded", timeout=30_000)

    # "Request New Draw Inspection" — matched by exact button text.
    req_btn = page.get_by_role("button", name="Request New Draw Inspection")
    try:
        req_btn.wait_for(state="visible", timeout=timeout_ms)
    except PWTimeout:
        raise RuntimeError(
            "'Request New Draw Inspection' button never appeared on the detail "
            "page. The draw may be in a state that doesn't allow re-inspection, "
            "or the page didn't finish loading."
        )
    print(">>> Clicking 'Request New Draw Inspection'…")
    req_btn.click()

    # A confirmation modal appears with an OK button. In test mode we DO NOT
    # click OK — the modal sits open so the user can verify, then close.
    ok_btn = page.locator(".ant-modal .ant-btn-primary").get_by_text("OK", exact=True).first
    try:
        ok_btn.wait_for(state="visible", timeout=timeout_ms)
    except PWTimeout:
        # Maybe the button isn't in a modal — try a global lookup too.
        ok_btn = page.get_by_role("button", name="OK").first
        ok_btn.wait_for(state="visible", timeout=5_000)

    if submit:
        print(">>> LIVE — clicking OK to confirm the new draw request.")
        ok_btn.click()
        # Wait for the modal to close so we know the click registered.
        try:
            page.locator(".ant-modal").first.wait_for(state="hidden", timeout=timeout_ms)
            print(">>> Confirmed. New draw inspection requested.")
        except PWTimeout:
            print("    !! modal didn't close after OK — check the page manually.")
    else:
        print(">>> TEST MODE — confirmation modal is open but OK was NOT clicked.")
        print("    Close the modal manually if you want to abort, or re-run "
              "with --submit to actually request the new inspection.")


# ===========================================================================
# CLI
# ===========================================================================
def choose_contact(contacts: list[dict], preselect: str = "") -> dict:
    if not contacts:
        # No contacts on the loan — don't abort. Return an empty placeholder so
        # the run still proceeds, fills everything else, and stops at Next.
        print(">>> No contacts found on the loan — continuing without one.")
        return {"role": "", "name": "", "email": "", "phone": ""}

    preselect = (preselect or "").strip().lower()

    # Explicit skip: 'none' means "don't add any contact at all". Useful in
    # GUI/automation flows where you'd rather fill the contact on the
    # ProxyPics form by hand than risk grabbing the wrong one.
    if preselect == "none":
        print(">>> --contact-choice none → skipping contact selection.")
        return {"role": "", "name": "", "email": "", "phone": ""}

    # Explicit numeric index
    if preselect.isdigit():
        i = int(preselect)
        if 1 <= i <= len(contacts):
            c = contacts[i - 1]
            print(f">>> Auto-picked #{i}: {c['role']} — {c['name']}")
            return c
        print(f"    --contact-choice {preselect} out of range; falling through to auto.")
        preselect = "auto"

    # Auto-pick. We want the HUMAN contact, which on Casa loans is the
    # Guarantor (the Borrower is typically the LLC with no email). So:
    #   1. Prefer a Guarantor that has an email.
    #   2. Else any contact with both email AND phone.
    #   3. Else any contact with either email or phone.
    #   4. Else the first contact.
    if preselect == "auto":
        for c in contacts:
            if c["role"].lower() == "guarantor" and c["email"]:
                print(f">>> Auto-picked Guarantor (has email): {c['role']} — {c['name']}")
                return c
        for c in contacts:
            if c["email"] and c["phone"]:
                print(f">>> Auto-picked (has email+phone): {c['role']} — {c['name']}")
                return c
        for c in contacts:
            if c["email"] or c["phone"]:
                print(f">>> Auto-picked (has some contact info): {c['role']} — {c['name']}")
                return c
        print(f">>> Auto-picked first: {contacts[0]['role']} — {contacts[0]['name']}")
        return contacts[0]

    # Interactive
    print("\nContacts on file:")
    for i, c in enumerate(contacts, 1):
        print(f"  {i}. {c['role']:9s} {c['name']:30s}  "
              f"email={c['email'] or '—':30s}  phone={c['phone'] or '—'}")
    while True:
        raw = input(f"Choose contact [1-{len(contacts)}]: ").strip()
        if raw.isdigit() and 1 <= int(raw) <= len(contacts):
            return contacts[int(raw) - 1]
        print("  ??? try again.")


def ensure_chromium_installed() -> None:
    """Verify Playwright's Chromium is on disk; install it if not.
    Handles both dev mode (sys.executable is python) and the frozen .exe
    (where we have to invoke playwright's bundled driver directly because
    `python -m` isn't available)."""
    import os, subprocess
    # Common install locations playwright checks in order.
    candidates = [
        os.environ.get("PLAYWRIGHT_BROWSERS_PATH"),
        os.path.expandvars(r"%LOCALAPPDATA%\ms-playwright"),
        os.path.expanduser("~/.cache/ms-playwright"),
        os.path.expanduser("~/Library/Caches/ms-playwright"),
    ]
    for d in candidates:
        if d and os.path.isdir(d):
            try:
                if any(n.startswith("chromium") for n in os.listdir(d)):
                    return  # already installed
            except OSError:
                pass

    print(">>> Playwright Chromium not found on this machine.")
    print(">>> Installing now (one-time, ~150 MB download)…")

    if getattr(sys, "frozen", False):
        # Inside the PyInstaller .exe — sys.executable is OUR exe, not python.
        # Invoke playwright's bundled driver instead.
        try:
            from playwright._impl._driver import compute_driver_executable, get_driver_env
            driver = compute_driver_executable()
            cmd = list(driver) if isinstance(driver, (list, tuple)) else [str(driver)]
            env = get_driver_env() if callable(get_driver_env) else os.environ
            subprocess.run(cmd + ["install", "chromium"], env=env, check=True)
        except Exception as e:
            print(f"ERROR: auto-install failed inside frozen exe: {e}")
            print("Manual fix: open PowerShell and run:")
            print("    python -m playwright install chromium")
            sys.exit(2)
    else:
        # Dev mode — sys.executable IS python.
        try:
            subprocess.run([sys.executable, "-m", "playwright", "install", "chromium"],
                           check=True)
        except Exception as e:
            print(f"ERROR: auto-install failed: {e}")
            print("Manual fix: run `python -m playwright install chromium`")
            sys.exit(2)
    print(">>> Chromium installed. Continuing.")


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--address", help="Property address (prompted if omitted)")
    ap.add_argument("--csv",     default=DEFAULTS["csv_path"])
    ap.add_argument("--step",    default=DEFAULTS["progress_step"])
    ap.add_argument("--type",    default=DEFAULTS["inspection_type_search"],
                    dest="insp_type")
    ap.add_argument("--contact-choice", default="",
                    help="'auto', or 1-based index. Empty = prompt.")
    ap.add_argument("--profile-dir", default="",
                    help="Browser user data directory (for persistent login).")
    ap.add_argument("--db", default="",
                    help="Path to Draw Manager's draws.db for the local "
                         "address→CF# fast path (skips brittle Casa scraping).")
    ap.add_argument("--cf", default="",
                    help="Manually-supplied CF# (loan number). Wins over the "
                         "DB lookup and skips Casa entirely if --contact-choice "
                         "is 'none'.")
    ap.add_argument("--clients-root", default="",
                    help="Path to the SharePoint Clients folder. Used to auto-"
                         "build the draw-items CSV from each property's "
                         "Rehab Budget xlsx. Falls back to DRAW_CLIENTS_ROOT "
                         "env var, then to the hard-coded default.")
    ap.add_argument("--submit", action="store_true",
                    help="LIVE mode: actually click Submit at the end. Without "
                         "this flag the run fills everything and stops at Next.")
    ap.add_argument("--headless", action="store_true",
                    help="Run the browser with no visible window.")
    ap.add_argument("--mode", choices=("auto", "new", "repeat"), default="auto",
                    help="auto = search ProxyPics first; if the address already "
                         "has a draw, request a new inspection on it; otherwise "
                         "fall through to the new-draw flow. "
                         "new = always create a brand-new draw. "
                         "repeat = only request a new inspection on an existing "
                         "draw (fails if none found).")
    args = ap.parse_args()

    # Allow override of the browser profile location (used when launched from
    # the Draw Manager app so dev-mode and exe-mode share a profile if desired).
    global PROFILE_DIR
    if args.profile_dir:
        PROFILE_DIR = Path(args.profile_dir)

    address = (args.address or input("Property address: ")).strip()
    # Normalize unicode whitespace — copies from spreadsheets / web pages
    # often carry NBSP (U+00A0), narrow-NBSP (U+202F), zero-width spaces,
    # etc. Casa's table filter does exact substring matching and will return
    # 0 rows when "1359 Wright St" has an NBSP instead of a regular space.
    if address:
        address = re.sub(r"\s+", " ", address.replace("\xa0", " ")
                                              .replace("\u202f", " ")
                                              .replace("\u2007", " ")
                                              .replace("\u200b", "")).strip()
    if not address:
        print("No address — exiting.")
        sys.exit(1)

    # The line-items CSV is optional. If a path is given but missing, warn and
    # carry on without it rather than aborting a session that may burn a login.
    if args.csv and not Path(args.csv).exists():
        print(f"WARNING: draw-items CSV not found, continuing without it: {args.csv}")
        args.csv = ""

    ensure_chromium_installed()

    with sync_playwright() as p:
        ctx = p.chromium.launch_persistent_context(
            user_data_dir=str(PROFILE_DIR),
            headless=args.headless,
            viewport={"width": 1400, "height": 900},
        )
        page = ctx.pages[0] if ctx.pages else ctx.new_page()

        print(f">>> Mode: {args.mode}  |  {'LIVE' if args.submit else 'TEST'}")
        print(f">>> Address: {address!r}")

        # Show the sheet-derived signal the Draw Manager uses to pick mode.
        # (This is read-only — it doesn't change behavior; it just makes the
        # decision visible in the standalone-run log too.)
        if args.db:
            status = _db_lookup_out_status(args.db, address)
            if status is None:
                print(">>> [sheet] No loan in draws.db for this address "
                      "(import the draw log to populate it).")
            elif status["draw_number"] is None:
                print(f">>> [sheet] Loan {status['cf']} in DB, no draws yet "
                      "→ treating as NEW draw.")
            else:
                flag = "TRUE" if status["out"] else "FALSE"
                verdict = "renew on ProxyPics" if status["out"] else "submit new draw"
                print(f">>> [sheet] Loan {status['cf']} draw #{status['draw_number']} "
                      f"OUT={flag} → {verdict}.")

        # --- mode dispatch -------------------------------------------------
        # auto / repeat first try to find an existing draw on ProxyPics.
        # If found → click "Request New Draw Inspection" and (optionally) OK.
        existing_href = None
        if args.mode in ("auto", "repeat"):
            existing_href = find_existing_draw(page, address)
            if not existing_href and args.mode == "repeat":
                print(">>> --mode repeat: no existing draw found. Exiting.")
                ctx.close()
                sys.exit(2)
            if existing_href:
                request_new_inspection_from_existing(
                    page, existing_href, submit=args.submit
                )
                print(">>> Done. Close the browser window when you're finished reviewing.")
                try:
                    page.wait_for_event("close", timeout=0)
                except Exception:
                    pass
                ctx.close()
                return

        # --- new-draw flow (mode=new, or mode=auto with no existing match) -
        # Resolve CF# from the best available source, in priority order:
        #   1. --cf flag (explicit, wins over everything)
        #   2. local draws.db (fast, no Casa)
        #   3. Casa scrape (brittle, last resort)
        # New draws often involve loans NOT yet in the DB — that's the
        # point. So nothing here is required. If we can't resolve a CF#,
        # we proceed with empty and let the operator fill it on the form.
        cf_hint: str | None = None
        if args.cf:
            cf_hint = args.cf.strip()
            print(f">>> Using CF# from --cf flag: {cf_hint}")
        elif args.db:
            cf_hint = _db_lookup_cf(args.db, address)
            if cf_hint:
                print(f">>> Local DB match: {cf_hint} — will skip Casa CF# scrape.")
            else:
                print(">>> No local DB match.")
        else:
            print(">>> No --db provided.")

        # Resolve the full property address from the SharePoint Clients
        # folder name FIRST — this is reliable and doesn't depend on Casa's
        # flaky search. e.g. typed "1300" → folder "1300 NW 42nd St, Miami,
        # FL 33142". We reuse this for both the ProxyPics address field and
        # the rehab-budget CSV lookup.
        clients_root = (args.clients_root
                        or os.environ.get("DRAW_CLIENTS_ROOT")
                        or DEFAULT_CLIENTS_ROOT)
        folder_address = ""
        prop_folder = None
        try:
            if clients_root and Path(_long(clients_root)).exists():
                prop_folder = find_property_folder(Path(_long(clients_root)), address)
                if prop_folder is not None:
                    folder_address = prop_folder.name.strip()
                    print(f">>> Property folder address: {folder_address}")
        except Exception as e:
            print(f"    (folder-address lookup warning: {e})")

        # Decide whether we need Casa at all. Casa is the source of contacts
        # (borrower/guarantor name/email/phone) and the CF#. If the user said
        # --contact-choice none AND we already have a CF#, skip Casa entirely.
        need_contacts = (args.contact_choice or "").lower() != "none"
        info = {"cf_number": cf_hint or "", "contacts": [], "address": ""}
        if need_contacts or not (cf_hint or args.cf):
            # Wrap Casa entirely: its search is flaky and can crash/close the
            # page. A failure here must NOT kill the run — we proceed with
            # the folder address and whatever CF# we have (possibly blank).
            try:
                info = casa_lookup(page, address, cf_hint=cf_hint,
                                   match_address=folder_address or address)
                # casa_lookup may have swapped to a fresh tab; keep that handle.
                if info.get("_page") is not None:
                    page = info["_page"]
            except Exception as e:
                print(f"    !! Casa lookup raised ({type(e).__name__}: {e}). "
                      "Proceeding without Casa info.")
                info = {"cf_number": cf_hint or "", "contacts": [], "address": ""}
        else:
            print(">>> Have CF# and --contact-choice none → skipping Casa.")

        contact = choose_contact(info.get("contacts", []),
                                 preselect=args.contact_choice)

        # Point-of-contact reconciliation: the loan Application PDF is the
        # authoritative source. Pull it, cross-check against the Casa
        # (Baseline) contact, and on any genuine mismatch (ignoring case /
        # formatting) ask which is correct via a popup.
        if (args.contact_choice or "").lower() != "none" and prop_folder is not None:
            app_contact = extract_application_contact(prop_folder)
            if any(app_contact.values()):
                merged, conflicts = reconcile_contact(contact, app_contact)
                for field in conflicts:
                    casa_val = (contact.get(field) or "").strip()
                    app_val  = (app_contact.get(field) or "").strip()
                    print(f">>> Contact {field} mismatch — Casa={casa_val!r} "
                          f"App={app_val!r}; asking user.")
                    merged[field] = ask_which_value(field, casa_val, app_val)
                contact = merged
                print(f">>> Final contact → name={contact.get('name')!r} "
                      f"phone={contact.get('phone')!r} "
                      f"email={contact.get('email')!r}")

        # Address priority for ProxyPics:
        #   1. SharePoint folder name (reliable, full address)
        #   2. Casa loan-page address (if folder lookup failed)
        #   3. The raw typed address (last resort)
        casa_address = (info.get("address") or "").strip()
        effective_address = folder_address or casa_address or address
        print(f">>> Address for ProxyPics: {effective_address!r}")

        # Resolve the draw-items CSV from the rehab budget xlsx. Reuse the
        # property folder we already found, so we don't scan the tree twice.
        csv_path = None
        try:
            if prop_folder is not None:
                xlsx = find_rehab_budget_xlsx(prop_folder)
                if args.csv and Path(_long(args.csv)).exists():
                    print(f">>> Using user-provided CSV: {args.csv}")
                    csv_path = args.csv
                elif xlsx is not None:
                    print(f"    found rehab budget: {xlsx.name}")
                    out_csv = Path(tempfile.gettempdir()) / f"draw_items_{int(time.time())}.csv"
                    build_draw_items_csv(xlsx, out_csv)
                    print(f"    built CSV: {out_csv}")
                    csv_path = str(out_csv)
                else:
                    print("    no .xlsx in property's Rehab Budget folder.")
            else:
                # Fall back to the original resolver (handles --csv too).
                csv_path = resolve_csv_for_address(effective_address,
                                                   args.clients_root, args.csv)
        except Exception as e:
            print(f"    !! CSV build failed ({e}); no CSV will be uploaded.")
            csv_path = None

        data = {
            "inspection_type_search": args.insp_type,
            "loan_number":            info.get("cf_number", "") or "",
            "address":                effective_address,
            "contact_name":           contact["name"],
            "contact_phone":          digits_only(contact["phone"]),
            "contact_email":          contact["email"],
            "csv_path":               csv_path or "",
            "progress_step":          args.step,
        }
        print(f"\n>>> {'LIVE — will SUBMIT' if args.submit else 'TEST — stops before Submit'}")
        print(f">>> Submitting to ProxyPics with:")
        for k, v in data.items():
            print(f"      {k}: {v}")
        print()

        # If a Casa misstep closed the page/tab, recover a live page from the
        # context so ProxyPics filling can still proceed.
        if not _safe_url(page):
            print(">>> Page was closed during lookup — opening a fresh tab "
                  "for ProxyPics.")
            try:
                page = ctx.new_page()
            except Exception as e:
                print(f"    !! Couldn't open a new page: {e}")
                raise

        fill_proxypics(page, data, submit=args.submit)

        print(">>> Done. Close the browser window when you're finished reviewing.")
        try:
            page.wait_for_event("close", timeout=0)
        except Exception:
            pass
        ctx.close()


if __name__ == "__main__":
    main()